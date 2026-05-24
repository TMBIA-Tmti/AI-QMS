"""
AI-QMS — Async Regulatory Website Crawler Module (v2.0)
=======================================================

High-performance async crawler for global medical device regulatory websites.

Architecture:
  Tier 0: XML Sitemap pre-scan — detect recently-updated pages via <lastmod>
  Tier 1: API / RSS / JSON   — structured data endpoints (fastest, most reliable)
  Tier 2: httpx async + MarkItDown — HTML fetch → BS4 cleanup → MarkItDown convert
  Tier 3: Jina Reader API    — for anti-scraping / SPA / blocked sites

Features:
  - asyncio.gather parallel crawling across all sites
  - httpx.AsyncClient with HTTP/2, shared connection pool
  - tenacity exponential-backoff retry
  - HTTP ETag / If-Modified-Since conditional caching
  - Per-domain asyncio.Semaphore rate limiting
  - aiofiles async I/O for cache persistence
  - MarkItDown convert_stream for HTML→Markdown (with BS4 pre-strip)
  - Jina Reader API fallback for JS-rendered / anti-bot sites
  - DuckDuckGo supplementary search

Output schema is backward-compatible with v1.0 (sync crawler).
"""

import io
import re
import json
import time
import hashlib
import asyncio
import logging
import zipfile
import ipaddress as _ipaddress
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Optional
from urllib.parse import urlparse

_urlparse = urlparse
import xml.etree.ElementTree as _stdlib_ET
import defusedxml.ElementTree as ET

import httpx
import aiofiles
from bs4 import BeautifulSoup

# tenacity retry
try:
    from tenacity import (
        retry,
        stop_after_attempt,
        wait_exponential,
        retry_if_exception_type,
    )

    TENACITY_AVAILABLE = True
except ImportError:
    TENACITY_AVAILABLE = False

# DuckDuckGo search (ddgs >= 8.0 or legacy)
try:
    from ddgs import DDGS
except ImportError:
    try:
        from duckduckgo_search import DDGS
    except ImportError:
        DDGS = None

# MarkItDown for HTML → Markdown conversion
try:
    from markitdown import MarkItDown

    _MD_CONVERTER = MarkItDown()
    MARKITDOWN_AVAILABLE = True
except ImportError:
    _MD_CONVERTER = None
    MARKITDOWN_AVAILABLE = False

# PyMuPDF for direct PDF text extraction
try:
    import fitz as _fitz

    FITZ_AVAILABLE = True
except ImportError:
    _fitz = None
    FITZ_AVAILABLE = False

logger = logging.getLogger(__name__)


def _is_safe_url(url: str) -> bool:
    """Return False if URL targets private/internal/loopback networks."""
    try:
        parsed = _urlparse(url)
        if parsed.scheme not in ("http", "https"):
            return False
        hostname = parsed.hostname
        if not hostname:
            return False
        if hostname.lower() in ("localhost", "127.0.0.1", "0.0.0.0", "::1"):
            return False
        if hostname.endswith(".internal") or hostname.endswith(".local"):
            return False
        try:
            ip = _ipaddress.ip_address(hostname)
            if ip.is_private or ip.is_loopback or ip.is_link_local or ip.is_reserved:
                return False
        except ValueError:
            pass  # hostname, not IP — OK
        return True
    except Exception:
        return False


# ============================================================
# Region & Site Configuration (with Tier assignments)
# ============================================================
#
# Tier 0: Sitemap pre-scan (optional, via sitemap_url field)
# Tier 1: API / RSS / JSON endpoints
# Tier 2: httpx async + BS4 cleanup + MarkItDown
# Tier 3: Jina Reader API (anti-scraping / SPA / blocked)

REGION_SITES = {
    "台灣 (Taiwan)": [
        {
            # Primary source: MOJ Open API bulk download — 26 active medical device
            # regulations fetched in one ZIP, no per-page scraping needed.
            "agency": "TFDA-BulkAPI",
            "name": "台灣醫療器材法規全集 — MOJ Open API Bulk (26 active regulations, ChOrder+ChLaw)",
            "url": "https://law.moj.gov.tw/api/ch/order/json",
            "tier": 1,
            "strategy": "bulk_zip",
            "crawl_delay": 0,
            "doc_type": "primary",
            "note": (
                "MOJ Open API bulk download: ChOrder.json.zip + ChLaw.json.zip → "
                "filter 衛生福利部＞食品藥物管理目 + '醫療器材' keyword → "
                "26 active laws (L0030106~L0030137). "
                "Covers: QMS準則(L0030116), 管理法(L0030106), 查核辦法(L0030112), "
                "嚴重不良事件(L0030124), 安全監視(L0030125), 許可證(L0030128), etc. "
                "Cache TTL: 7 days (API updates weekly on Fridays)."
            ),
        },
        {
            # English fallback: Jina Reader for the QMS Criteria English page.
            # Kept as supplementary source for English-language LLM analysis.
            "agency": "TFDA-QMS-EN",
            "name": "Medical Device Quality Management System Guidelines (English) — Taiwan Laws Database",
            "url": "https://law.moj.gov.tw/ENG/LawClass/LawAll.aspx?pcode=L0030116",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "doc_type": "qms_guidance",
            "note": (
                "English version of QMS Criteria L0030116 — ISO 13485:2016 equivalent "
                "under Medical Devices Act 2021. Supplementary to bulk API primary source."
            ),
        },
    ],
    "美國 (USA)": [
        {
            "agency": "FDA-QMSR",
            "name": "FDA Quality Management System Regulation (QMSR) — Official Overview",
            "url": "https://www.fda.gov/medical-devices/postmarket-requirements-devices/quality-management-system-regulation-qmsr",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 5,
            "sitemap_url": "https://www.fda.gov/sitemap.xml",
            "note": "PRIMARY QMS regulation page: 21 CFR Part 820 QMSR, effective 2026-02-02, incorporates ISO 13485:2016 by reference",
        },
        {
            "agency": "eCFR-820",
            "name": "Federal Register — FDA QMSR Final Rule (2024-01709, Medical Devices; Quality System Regulation Amendments)",
            "url": "https://www.federalregister.gov/documents/2024/02/02/2024-01709/medical-devices-quality-system-regulation-amendments",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "QMSR Final Rule (89 FR 7496, Rule 2024-01709) published Feb 2, 2024, effective Feb 2, 2026 — amends 21 CFR Part 820 to incorporate ISO 13485:2016 by reference; eCFR full text at ecfr.gov/current/title-21/chapter-I/subchapter-H/part-820",
        },
    ],
    "歐盟 (EU)": [
        {
            "agency": "EUR-Lex-MDR-CELLAR",
            "name": "Regulation (EU) 2017/745 — EU MDR Full Text PDF, consolidated 02017R0745-20260101 (EUR-Lex CELLAR official, English)",
            "url": "http://publications.europa.eu/resource/cellar/fddb3266-f0ab-11f0-8d3c-01aa75ed71a1.0007.03/DOC_1",
            "tier": 1,
            "strategy": "html",
            "crawl_delay": 3,
            "fallback_urls": [
                "https://www.legislation.gov.uk/eur/2017/745/pdfs/eur_20170745_2020-04-24_en.pdf",
            ],
            "note": "PRIMARY EU MDR source: EUR-Lex CELLAR consolidated PDF CELEX 02017R0745-20260101 (includes M1-M5 amendments through 2024/1860, English). DOC_1 endpoint bypasses JS challenge. Fallback: legislation.gov.uk 2020-04-24 PDF (missing 2023/2024 amendments but full article text). Update URL by running scripts/download_mdr_full_text.py.",
        },
        {
            "agency": "EUR-Lex-MDR-UK",
            "name": "Regulation (EU) 2017/745 — EU MDR Full Text PDF (legislation.gov.uk revised 2020-04-24, English)",
            "url": "https://www.legislation.gov.uk/eur/2017/745/pdfs/eur_20170745_2020-04-24_en.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "UK-hosted EU MDR PDF — consolidated to EU 2020/561 (2020-04-24). Missing EU 2023/607 and 2024/1860 amendments. Used as secondary source when CELLAR is unavailable. Direct PDF link, no JS challenge.",
        },
        {
            "agency": "MDCG",
            "name": "MDCG Guidance Documents (QMS, Annex IX, notified body guidance) — all PDFs as separate files",
            "url": "https://health.ec.europa.eu/medical-devices-sector/new-regulations/guidance-mdcg-endorsed-documents-and-other-guidance_en",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "index_page": True,
            "save_attachments_separately": True,
            "note": "MDCG guidance index — 131 PDFs each saved as an individual markdown file; save_attachments_separately=True; no attachment count limit (uses _MAX_FILE_ATTACHMENTS_INDEX_PAGE=200)",
        },
        {
            "agency": "MDCG-2019-11",
            "name": "MDCG 2019-11 Rev.1 — QMS documentation requirements for notified body assessment (Annex IX §2.2)",
            "url": "https://health.ec.europa.eu/document/download/b45335c5-1679-4c71-a91c-fc7a4d37f12b_en?filename=mdcg_2019_11_en.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MDCG 2019-11 Rev.1 — specifies QMS docs NB auditors require under MDR Annex IX §2.2; verified UUID from live MDCG index page",
        },
        {
            "agency": "MDCG-2021-25",
            "name": "MDCG 2021-25 Rev.1 — GSPR Application Guide (Annex I compliance in QMS)",
            "url": "https://health.ec.europa.eu/document/download/cbb11a6e-f0f3-4e30-af5e-990f9ef68bc1_en?filename=md_mdcg_2021_25_en.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MDCG 2021-25 Rev.1 — QMS must document GSPR compliance; verified UUID + HEAD 200 application/pdf",
        },
        {
            "agency": "MDCG-2020-7",
            "name": "MDCG 2020-7 — PMCF Plan Template (Post-Market Clinical Follow-up, Annex XIV Part B)",
            "url": "https://health.ec.europa.eu/document/download/a5cdb303-c782-4010-8723-7d389af678f7_en?filename=md_mdcg_2020_7_guidance_pmcf_plan_template_en.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MDCG 2020-7 — PMCF Plan template; verified UUID + HEAD 200 application/pdf",
        },
        {
            "agency": "MDCG-2022-14",
            "name": "MDCG 2022-14 — Transition to MDR/IVDR (legacy device QMS implications)",
            "url": "https://health.ec.europa.eu/document/download/2db053bc-283c-4d2e-93f4-c3e8032e66da_en?filename=mdcg_2022-14_en.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MDCG 2022-14 — legacy device transition QMS impact; verified UUID",
        },
        {
            "agency": "MDCG-2021-27",
            "name": "MDCG 2021-27 Rev.1 — Guidance on Unique Device Identification (UDI) for medical devices",
            "url": "https://health.ec.europa.eu/document/download/82d9adbc-dbf0-40d4-93ed-ade673c8232a_en?filename=mdcg_2021-27_en.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MDCG 2021-27 Rev.1 — primary UDI guidance for MDR/IVDR; UDI assignment and traceability in QMS; verified UUID",
        },
        {
            "agency": "MDCG-2022-21",
            "name": "MDCG 2022-21 — Notified body oversight of manufacturers' post-market surveillance",
            "url": "https://health.ec.europa.eu/document/download/a7df24c3-d4a3-4218-a8e0-726febfa01c2_en?filename=mdcg_2022-21_en.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MDCG 2022-21 — NB oversight of manufacturer QMS PMS activities; directly relevant to QMS audit scope; verified UUID",
        },
        {
            "agency": "MDCG-2019-6",
            "name": "MDCG 2019-6 Rev.5 — Questions and Answers on notified body requirements (QMS audit criteria)",
            "url": "https://health.ec.europa.eu/document/download/9c9c532f-013a-477c-9378-0a9e714e5549_en?filename=md_mdcg_qa_requirements_notified_bodies_en.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MDCG 2019-6 Rev.5 — Q&A on NB requirements incl. QMS assessment criteria; verified UUID",
        },
        {
            "agency": "MDCG-2018-1",
            "name": "MDCG 2018-1 Rev.4 — Guidance on UDI-DI assignment for medical devices",
            "url": "https://health.ec.europa.eu/document/download/cb1bf6e5-3972-4b3a-82d9-c5946738b2a5_en?filename=md_mdcg_2018-1_guidance_udi-di_en.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MDCG 2018-1 Rev.4 — UDI-DI assignment rules; traceability requirements in QMS; verified UUID",
        },
        {
            "agency": "MDCG-2020-14",
            "name": "MDCG 2020-14 — Guidance on MDSAP audits and their relationship to MDR/IVDR",
            "url": "https://health.ec.europa.eu/document/download/44dc96aa-e517-4af1-855b-f7fcb4b699c9_en?filename=md_2020-14-guidance-mdsap_en.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MDCG 2020-14 — MDSAP audit recognition and QMS equivalence under EU MDR; verified UUID",
        },
        {
            "agency": "MDCG-2023-1",
            "name": "MDCG 2023-1 — Guidance on UDI for implantable devices (updated UDI requirements)",
            "url": "https://health.ec.europa.eu/document/download/05b15d55-1bcf-4e17-99c4-15c706325847_en?filename=mdcg_2023-1_en.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MDCG 2023-1 — UDI implantable devices guidance; updated traceability requirements in QMS; verified UUID",
        },
    ],
    "英國 (UK)": [
        {
            "agency": "UK-MDR-2002",
            "name": "The Medical Devices Regulations 2002 (SI 2002/618) — UK legislation.gov.uk",
            "url": "https://www.legislation.gov.uk/uksi/2002/618/made",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: UK MDR 2002 (SI 618) — Parts II/III/IV specify conformity assessment & QMS requirements; basis for UKCA marking",
        },
        {
            "agency": "MHRA-Guidance",
            "name": "MHRA — Regulating Medical Devices in the UK (GOV.UK Content API)",
            "url": "https://www.gov.uk/api/content/guidance/regulating-medical-devices-in-the-uk",
            "tier": 1,
            "strategy": "api_json",
            "crawl_delay": 3,
            "note": "MHRA official guidance on UK MDR 2002 compliance including QMS requirements for UKCA",
        },
    ],
    "日本 (Japan)": [
        {
            "agency": "eGov-QMS-169",
            "name": "医療機器及び体外診断用医薬品の製造管理及び品質管理の基準に関する省令 (MHLW Ordinance No. 169 — QMS省令 full text)",
            "url": "https://laws.e-gov.go.jp/law/416M60000100169",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: MHLW Ministerial Ordinance No. 169 (2004, revised 2021) — Japanese QMS Ordinance (QMS省令) mirroring ISO 13485:2016 — official e-Gov law database",
        },
        {
            "agency": "PMDA-QMS",
            "name": "PMDA — QMS Compliance Inspection (QMS適合性調査)",
            "url": "https://www.pmda.go.jp/review-services/gmp-qms-gctp/qms/0003.html",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PMDA QMS compliance inspection page — covers QMS省令 requirements and inspection procedures",
        },
        {
            "agency": "PMDA-QMS-EN",
            "name": "PMDA — Revision of Japanese Medical Device QMS Requirements (English)",
            "url": "https://www.pmda.go.jp/english/review-services/regulatory-info/0004.html",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "English-language PMDA page on 2021 QMS Ordinance 169 revision — comparison with ISO 13485:2016 and Chapter translations",
        },
    ],
    "中國 (China)": [
        {
            "agency": "NMPA-GMP-CN",
            "name": "医疗器械生产质量管理规范 — NMPA 公告 2025年第107号 全文 (商务部政策法规网, Chinese full text)",
            "url": "https://policy.mofcom.gov.cn/claw/clawContent.shtml?id=104034",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY (Chinese full text): Ministry of Commerce policy portal with full inline regulation text (15章132条) — accessible without anti-bot blocking; original NMPA site returns 412",
        },
        {
            "agency": "NMPA-GMP-Announcement",
            "name": "NMPA — GMP Announcement [2025] No. 107 (English, effective 2026-11-01)",
            "url": "https://english.nmpa.gov.cn/2025-12/24/c_1156627.htm",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "Official NMPA English announcement releasing the revised Good Manufacturing Practice for Medical Devices (15 chapters, 132 articles) — Jina Reader first",
        },
        {
            "agency": "NMPA-Regulations",
            "name": "NMPA — Regulatory Information (medical devices section)",
            "url": "https://english.nmpa.gov.cn/medicaldevices.html",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "NMPA English portal medical device regulatory page — Jina Reader first",
        },
    ],
    "韓國 (Korea)": [
        {
            "agency": "MFDS-KGMP",
            "name": "MFDS — Medical Device Good Manufacturing Practice (K-GMP) English compilation",
            "url": "https://www.mfds.go.kr/eng/brd/m_40/view.do?seq=72638",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "force_profile": True,
            "note": "K-GMP — MFDS blocks all non-Korean IP (ConnectError); force_profile=True as hardcoded fallback; pre-written profile contains full K-GMP article list",
        },
        {
            "agency": "MFDS-KGMP-PDF",
            "name": "K-GMP Full Text PDF (MFDS English) — blocked from non-Korean IPs",
            "url": "https://www.mfds.go.kr/brd/m_218/down.do?brd_id=data0011&seq=14629&data_tp=A&file_seq=1",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "force_profile": True,
            "note": "K-GMP PDF — also blocked from non-Korean IPs; force_profile fallback ensures content availability",
        },
        {
            "agency": "MFDS-MD-Regulations",
            "name": "MFDS — Medical Device Regulations Listing",
            "url": "https://www.mfds.go.kr/eng/brd/m_40/list.do",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MFDS English regulations listing — includes GMP/QMS ordinances — Jina first",
        },
    ],
    # Canada QMS is exclusively via MDSAP since Jan 2019 (CMDCAS retired).
    # Primary QMS regulation is CMDR SOR/98-282 Section 32 (requires ISO 13485 certificate).
    # MDSAP sites also included as the operational audit framework.
    "加拿大 (Canada)": [
        {
            "agency": "CMDR-SOR98-282",
            "name": "Medical Devices Regulations SOR/98-282 — Section 32 (ISO 13485 QMS requirement)",
            "url": "https://laws-lois.justice.gc.ca/eng/regulations/SOR-98-282/section-32.html",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: CMDR SOR/98-282 s.32 requires ISO 13485 QMS certificate for Class II–IV device licences — Health Canada Justice Laws",
        },
        {
            "agency": "CMDR-Full",
            "name": "Medical Devices Regulations SOR/98-282 — Full Text",
            "url": "https://laws-lois.justice.gc.ca/eng/regulations/SOR-98-282/FullText.html",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Full text of CMDR SOR/98-282 — governs QMS, labelling, licensing, post-market for medical devices in Canada",
        },
        {
            "agency": "MDSAP-Global-Audit",
            "name": "MDSAP Audit Procedures and Forms",
            "url": "https://www.mdsap.global/documents/audit-procedures-and-forms",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "Akamai protection — Jina Reader first",
        },
        {
            "agency": "FDA-MDSAP-Audit",
            "name": "FDA MDSAP Audit Procedures and Forms",
            "url": "https://www.fda.gov/medical-devices/medical-device-single-audit-program-mdsap/mdsap-audit-procedures-and-forms",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 5,
            "index_page": True,
            "sitemap_url": "https://www.fda.gov/sitemap.xml",
            "note": "FDA MDSAP audit forms index — index_page=True extracts all linked PDFs (up to 50)",
        },
        {
            "agency": "FDA-MDSAP-Assessment",
            "name": "FDA MDSAP Assessment Procedures and Forms",
            "url": "https://www.fda.gov/medical-devices/medical-device-single-audit-program-mdsap/mdsap-assessment-procedures-and-forms",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 5,
            "index_page": True,
            "sitemap_url": "https://www.fda.gov/sitemap.xml",
            "note": "FDA MDSAP assessment forms index — index_page=True extracts all linked PDFs",
        },
        {
            "agency": "MDSAP-Companion-ISO13485",
            "name": "MDSAP Companion Document to ISO 13485:2016 (FDA PDF)",
            "url": "https://www.fda.gov/media/102395/download",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "FDA MDSAP Companion Document — maps MDSAP audit requirements clause-by-clause to ISO 13485:2016 — PyMuPDF direct PDF",
        },
    ],
    "澳洲 (Australia)": [
        {
            "agency": "TGA-Legislation",
            "name": "Therapeutic Goods (Medical Devices) Regulations 2002 — Schedule 3 (AustLII plain HTML, no JS)",
            "url": "https://classic.austlii.edu.au/au/legis/cth/consol_reg/tgdr2002400/sch3.html",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS: TG(MD)R 2002 Schedule 3 conformity assessment requiring ISO 13485 — AustLII classic.austlii.edu.au serves plain HTML; tier 2 httpx (no Jina needed)",
        },
        {
            "agency": "TGA-Legislation-PDF",
            "name": "Therapeutic Goods (Medical Devices) Regulations 2002 — Compiled PDF (Federal Register)",
            "url": "https://www.legislation.gov.au/Details/F2022C00567/Download",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "TG(MD)R 2002 compiled PDF — full Schedule 3 QMS conformity assessment text; PyMuPDF direct",
        },
        {
            "agency": "TGA-ARGMD",
            "name": "TGA — Australian Regulatory Guidelines for Medical Devices (ARGMD)",
            "url": "https://www.tga.gov.au/products/medical-devices/overview/australian-regulatory-guidelines-medical-devices-argmd",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "TGA ARGMD — explains Schedule 3 conformity assessment including QMS requirements — Jina Reader",
        },
    ],
    "瑞士 (Switzerland)": [
        {
            "agency": "MedDO-SR812213",
            "name": "Verordnung über Medizinprodukte (MedDO, SR 812.213) — Medical Devices Ordinance (admin.ch)",
            "url": "https://www.fedlex.admin.ch/eli/cc/2020/552/en",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: Swiss MedDO SR 812.213 (1 July 2020, in force 2021-05-26) aligned with EU MDR 2017/745 — includes Annex IX QMS conformity assessment requirements — fedlex requires JS — Jina Reader first",
        },
        {
            "agency": "Swissmedic-MedDO",
            "name": "Swissmedic — Legal Framework (MedDO / EU MDR alignment)",
            "url": "https://www.swissmedic.ch/swissmedic/en/home/medical-devices/regulation-of-medical-devices/neue-eu-verordnungen-mdr-ivdr.html",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Swissmedic explanation of MedDO QMS requirements — aligned with EU MDR Annex IX",
        },
    ],
    "巴西 (Brazil)": [
        {
            "agency": "ANVISA-RDC665",
            "name": "RDC nº 665/2022 — ANVISA Good Manufacturing Practices for Medical Devices (BPF/GMP) — English PDF",
            "url": "https://www.gov.br/anvisa/en/regulation-of-companies/arquivos/rdc-665-2022-english-version.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "PRIMARY QMS regulation: ANVISA RDC 665/2022 English PDF (309KB) — tier 2 httpx direct download; Jina fallback if 403; pre-written profile last resort",
        },
        {
            "agency": "ANVISA-RDC665-News",
            "name": "ANVISA — RDC 665 de 2022 (Portuguese official news page)",
            "url": "https://www.gov.br/anvisa/pt-br/assuntos/noticias-anvisa/2022/rdc-665-de-2022",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "ANVISA official RDC 665/2022 announcement page in Portuguese — Inconsistent availability — Jina Reader fallback",
        },
    ],
    "MDSAP": [
        {
            "agency": "MDSAP-Global-Audit",
            "name": "MDSAP Audit Procedures and Forms",
            "url": "https://www.mdsap.global/documents/audit-procedures-and-forms",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "index_page": True,
            "note": "Akamai protection — Jina Reader first; index_page=True to capture all linked documents",
        },
        {
            "agency": "MDSAP-Global-QMS",
            "name": "MDSAP Quality Management System",
            "url": "https://www.mdsap.global/documents/quality-management-system",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "index_page": True,
            "note": "Akamai protection — Jina Reader first; index_page=True",
        },
        {
            "agency": "MDSAP-Global-General",
            "name": "MDSAP General Documents and Procedures",
            "url": "https://www.mdsap.global/documents/general-documents-and-procedures",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "index_page": True,
            "note": "Akamai protection — frequent timeouts, Jina Reader fallback; index_page=True",
        },
        {
            "agency": "MDSAP-Global-Assessment",
            "name": "MDSAP Assessment Procedures and Forms",
            "url": "https://www.mdsap.global/documents/assessment-procedures-and-forms",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "index_page": True,
            "note": "Akamai protection — frequent timeouts, Jina Reader fallback; index_page=True",
        },
        {
            "agency": "FDA-MDSAP-Audit",
            "name": "FDA MDSAP Audit Procedures and Forms",
            "url": "https://www.fda.gov/medical-devices/medical-device-single-audit-program-mdsap/mdsap-audit-procedures-and-forms",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 5,
            "index_page": True,
            "sitemap_url": "https://www.fda.gov/sitemap.xml",
            "note": "FDA MDSAP audit forms index — index_page=True downloads all linked PDFs (up to 50)",
        },
        {
            "agency": "FDA-MDSAP-Assessment",
            "name": "FDA MDSAP Assessment Procedures and Forms",
            "url": "https://www.fda.gov/medical-devices/medical-device-single-audit-program-mdsap/mdsap-assessment-procedures-and-forms",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 5,
            "index_page": True,
            "sitemap_url": "https://www.fda.gov/sitemap.xml",
            "note": "FDA MDSAP assessment forms index — index_page=True downloads all linked PDFs",
        },
        {
            "agency": "MDSAP-Companion-ISO13485",
            "name": "MDSAP Companion Document to ISO 13485:2016 (FDA PDF)",
            "url": "https://www.fda.gov/media/102395/download",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "FDA MDSAP Companion Document — maps MDSAP audit requirements clause-by-clause to ISO 13485:2016 — PyMuPDF direct PDF",
        },
    ],
    "國際標準 (International Standard)": [
        {
            "agency": "ISO",
            "name": "ISO Medical Equipment Standards (ICS 11.040)",
            "url": "https://www.iso.org/ics/11.040/x/",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "Blocks AI bots — metadata only via Jina",
            "sitemap_url": "https://www.iso.org/sitemap.xml",
        },
        {
            "agency": "ICH",
            "name": "ICH Quality Guidelines",
            "url": "https://www.ich.org/page/quality-guidelines",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "SPA site — Jina Reader required",
        },
        {
            "agency": "IMDRF",
            "name": "IMDRF Document Library",
            "url": "https://www.imdrf.org/documents/library",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "403 Forbidden — Jina Reader fallback",
        },
    ],
    # ---------- New regions (v2.0) ----------
    "印度 (India)": [
        {
            "agency": "CDSCO-MDR2017",
            "name": "India Medical Devices Rules 2017 (MDR 2017) — CDSCO official PDF",
            "url": "https://cdsco.gov.in/opencms/resources/UploadCDSCOWeb/2022/m_device/Medical%20Devices%20Rules,%202017.pdf",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: India MDR 2017 — Schedule V specifies QMS requirements for medical device manufacturers; ISO 13485 is the referenced standard — CDSCO PDF via Jina",
        },
        {
            "agency": "CDSCO-MD",
            "name": "CDSCO — Medical Devices & Diagnostics portal",
            "url": "https://cdsco.gov.in/opencms/opencms/en/Medical-Device-Diagnostics/Medical-Device-Diagnostics/",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "CDSCO medical device portal — SSL/bot protection — Jina Reader first",
        },
    ],
    "新加坡 (Singapore)": [
        {
            "agency": "SSO-HPR-2010-PDF",
            "name": "Health Products (Medical Devices) Regulations 2010 (S 436/2010) — Full Text PDF (Singapore Statutes Online)",
            "url": "https://sso.agc.gov.sg/SL/HPA2007-S436-2010?DocDate=20231211&ViewType=Pdf",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: HP(MD)R 2010 full text PDF (as amended 11 Dec 2023) — requires ISO 13485 or equivalent QMS for Class B/C/D medical device dealers — Singapore Statutes Online PDF version; HTML version returns only title — Jina first",
        },
        {
            "agency": "SSO-HPR-2010",
            "name": "Health Products (Medical Devices) Regulations 2010 (S 436/2010) — Singapore Statutes Online HTML",
            "url": "https://sso.agc.gov.sg/SL/HPA2007-S436-2010",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "HP(MD)R 2010 HTML index page — fallback; PDF version preferred",
        },
        {
            "agency": "HSA-QMS",
            "name": "HSA — Quality Management System (QMS) for Medical Devices",
            "url": "https://www.hsa.gov.sg/medical-devices/dealers-licence/quality-management-system-(qms)-for-medical-devices",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "HSA official QMS requirements page — ISO 13485 / MDSAP mandatory from 1 Jan 2025 for manufacturer licence",
        },
    ],
    "沙烏地阿拉伯 (Saudi Arabia)": [
        {
            "agency": "SFDA-MDS-REQ10",
            "name": "SFDA — Requirements for Inspections and Quality Management System for Medical Devices (MDS-REQ10)",
            "url": "https://www.sfda.gov.sa/en/regulations/87120",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: SFDA MDS-REQ10 — specifies QMS inspection requirements and ISO 13485 alignment for medical device establishments in KSA — Jina Reader first",
        },
        {
            "agency": "SFDA-ISO13485-Guidance",
            "name": "SFDA — Guidance for ISO 13485 Requirements with SFDA-MDS Regulations (MDS-G-024)",
            "url": "https://www.sfda.gov.sa/sites/default/files/2025-03/MDS-G024.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "SFDA guidance document mapping ISO 13485:2016 requirements to Saudi MDS regulations (March 2025)",
        },
    ],
    "泰國 (Thailand)": [
        {
            "agency": "Thai-FDA-GMP-Law",
            "name": "ประกาศกระทรวงสาธารณสุข เรื่อง หลักเกณฑ์ วิธีการ GMP เครื่องมือแพทย์ — Thai FDA law page (mdlaw03045)",
            "url": "https://medical.fda.moph.go.th/relevant-laws-and-standards/mdlaw03045",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: ประกาศกระทรวงสาธารณสุข (Ministry of Public Health Notification) on GMP/QMS for medical devices B.E. 2566 (2023) — specific law page on Thai FDA — published Royal Gazette 5 Jan 2024, effective 3 Jul 2024 — Jina first",
        },
        {
            "agency": "Thai-FDA-NewLaws",
            "name": "Thai FDA — อัพเดทกฎหมายออกใหม่ New Laws Update",
            "url": "https://medical.fda.moph.go.th/relevant-laws-and-standards/newsupdate01",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Thai FDA recent law updates page — lists all new MOPH notifications including GMP B.E. 2566 — Jina first",
        },
        {
            "agency": "Thai-FDA-GMP-About",
            "name": "Thai FDA — เกี่ยวกับ GMP เครื่องมือแพทย์ About GMP Medical Devices (B.E. 2566/2023)",
            "url": "https://medical.fda.moph.go.th/situation/category/about-gmp-medical-devices",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Thai FDA About-GMP landing page — overview of B.E. 2566 GMP requirements and ISO 13485/TCAS-13485 mandate for Class 2-4 devices — Jina first",
        },
    ],
    "紐西蘭 (New Zealand)": [
        {
            "agency": "Medsafe-MD-Legislation",
            "name": "Medsafe — Medical Device Legislation (current framework)",
            "url": "https://www.medsafe.govt.nz/regulatory/devicesnew/2Legislation.asp",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation source: Medsafe device legislation page — Therapeutic Products Act 2023 was REPEALED 18 Dec 2024 by TPA Repeal Act 2024; current law is Medicines Act 1981 (amended by Medicines Amendment Act 2025, in force 19 Nov 2025); new Medical Products Bill in development",
        },
        {
            "agency": "MoH-NZ-MD-Regulation",
            "name": "NZ Ministry of Health — Regulating medicines, medical devices and natural health products",
            "url": "https://www.health.govt.nz/regulation-legislation/medicines-legislation/regulating-medicines-medical-devices-and-natural-health-products",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "NZ MoH regulatory framework page — health.govt.nz blocks direct httpx (403); Jina Reader first",
        },
        {
            "agency": "NZ-Medicines-Act-1981",
            "name": "Medicines Act 1981 — New Zealand Legislation (current as amended 2025)",
            "url": "https://www.legislation.govt.nz/act/public/1981/0118/latest/whole.html",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Medicines Act 1981 full text — current governing legislation for medical devices in NZ; amended by Medicines Amendment Act 2025 — Cloudflare may block; Jina first",
        },
        {
            "agency": "Medsafe-GMP",
            "name": "Medsafe — NZ Code of GMP (Introduction)",
            "url": "https://www.medsafe.govt.nz/regulatory/guideline/nzgmpcodepart1intro.asp",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "NZ GMP Code Part 1 — applicable to therapeutic goods including medical devices — ISO 13485 referenced standard",
        },
    ],
    "墨西哥 (Mexico)": [
        {
            "agency": "DOF-NOM241-2025",
            "name": "NOM-241-SSA1-2025 — Buenas Prácticas de Fabricación para Dispositivos Médicos (DOF, 11 Nov 2025)",
            "url": "https://dof.gob.mx/nota_detalle.php?codigo=5772517&fecha=11/11/2025",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: NOM-241-SSA1-2025 (supersedes 2021 version) — Mexico official GMP standard for medical devices, ISO 13485 equivalent — published DOF 11/11/2025 — MDSAP recognized as equivalent QMS",
        },
        {
            "agency": "NOM241-EN-PDF",
            "name": "NOM-241-SSA1-2021 (English translation — InterAmerican Coalition MedTech, superseded by 2025 version)",
            "url": "https://www.interamericancoalition-medtech.org/regulatory-convergence/wp-content/uploads/sites/4/2022/03/Norma-Oficial-Mexicana-NOM-241-SSA1-2021-ENG-REV.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "English version of NOM-241-SSA1-2021 (2021 version, now superseded by 2025 revision) — retained for reference on prior requirements",
        },
    ],
    "阿根廷 (Argentina)": [
        {
            "agency": "ANMAT-MD",
            "name": "ANMAT — Productos Médicos (Medical Products portal, argentina.gob.ar)",
            "url": "https://www.argentina.gob.ar/anmat/regulados/productos-medicos",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation source: ANMAT medical products portal — Disposición 64/25 (Jan 2025) replaced Disposición 2318/2002 for MERCOSUR device registration; BPF (GMP/QMS) certification required for Class II–IV manufacturers",
        },
        {
            "agency": "ANMAT-Disposicion64-25",
            "name": "Disposición ANMAT 64/2025 — Boletín Oficial de la República Argentina (13 Jan 2025)",
            "url": "https://www.boletinoficial.gov.ar/detalleAviso/primera/319522/20250113",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Disposición ANMAT 64/25 (in force Jan 2025) — incorporates MERCOSUR Resolution GMC 25/21, replacing prior registration framework; Class III–IV require BPF certificate — Official Gazette confirmed accessible; argentina.gob.ar returns Jina upstream error",
        },
    ],
    "南非 (South Africa)": [
        {
            "agency": "SAHPRA-ISO13485",
            "name": "Medicines and Related Substances Act 101 of 1965 — Consolidated text (SAFLII via Jina)",
            "url": "https://www.saflii.org/za/legis/consol_act/marsa1965280/",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "PRIMARY QMS regulation: Medicines Act 101/1965 — SAFLII blocks HEAD (403) but full HTML text accessible via Jina Reader GET; s.22C + Regulations 5&6 require ISO 13485:2016",
        },
        {
            "agency": "SAHPRA-MD",
            "name": "SAHPRA — Medical Devices",
            "url": "https://www.sahpra.org.za/medical-devices/",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "SAHPRA medical device section — licensing, registration, and QMS requirements",
        },
    ],
    "土耳其 (Turkey)": [
        {
            "agency": "Mevzuat-TCY",
            "name": "Tıbbi Cihaz Yönetmeliği — mevzuat.gov.tr Official PDF (KurumVeKurulusYonetmeligi No. 38657)",
            "url": "https://mevzuat.gov.tr/File/GeneratePdf?mevzuatNo=38657&mevzuatTur=KurumVeKurulusYonetmeligi&mevzuatTertip=5",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "PRIMARY QMS regulation: Tıbbi Cihaz Yönetmeliği — official mevzuat.gov.tr PDF endpoint (KurumVeKurulusYonetmeligi type, No. 38657) — EU MDR 2017/745-equivalent, Annex IX QMS requirements — Jina for PDF",
        },
        {
            "agency": "TITCK-Mevzuat",
            "name": "TITCK — Tıbbi Cihaz Mevzuatı (Medical Device Legislation page)",
            "url": "https://www.titck.gov.tr/faaliyetalanlari/tibbicihaz/tibbi-cihaz-mevzuati",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "TITCK official medical device legislation index — lists all applicable regulations including Tıbbi Cihaz Yönetmeliği (2021) with links — Jina first",
        },
        {
            "agency": "TITCK-MD-Legislation",
            "name": "Tıbbi Cihaz Yönetmeliği — Resmi Gazete 31499 Mükerrer PDF (2 June 2021)",
            "url": "https://www.resmigazete.gov.tr/eskiler/2021/06/20210602M1-2.pdf",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "force_profile": True,
            "note": "Resmi Gazete PDF (4.9MB scanned image) — fitz: no text layer; docling: OOM on 142 pages; Jina: HTTP 422 — using pre-written profile (EU MDR 2017/745 equivalent QMS requirements)",
        },
        {
            "agency": "TITCK-New-Regs",
            "name": "TITCK — New Medical Device Regulations entered into force (2021 English announcement)",
            "url": "https://titck.gov.tr/duyuru/new-medical-device-regulations-entered-into-force-14062021145923",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "TITCK English announcement of June 2021 medical device regulations — EU MDR 2017/745 equivalent — Jina first",
        },
    ],
    "印尼 (Indonesia)": [
        {
            "agency": "Kemkes-CPAKB",
            "name": "Permenkes No. 20 Tahun 2017 — Cara Pembuatan Alat Kesehatan yang Baik (CPAKB) — Farmalkes Kemkes",
            "url": "https://farmalkes.kemkes.go.id/en/unduh/permenkes-20-2017/",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: Permenkes No. 20/2017 on Good Manufacturing Practices for Medical Devices (CPAKB) — issued 8 March 2017, effective 18 April 2017 — official Directorate General of Pharmaceutical and Medical Devices (Farmalkes) page",
        },
        {
            "agency": "BPK-Permenkes20",
            "name": "Permenkes No. 20 Tahun 2017 — CPAKB Direct PDF Download (Farmalkes Kemkes wpdmdl=11316)",
            "url": "https://farmalkes.kemkes.go.id/?wpdmdl=11316",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Full text of CPAKB regulation — confirmed 3.5MB, 60-page Indonesian-language PDF directly from Farmalkes Kemkes (WordPress WPDM stable download); peraturan.bpk.go.id returns 403 Forbidden — Jina/MarkItDown for PDF",
        },
    ],
    "馬來西亞 (Malaysia)": [
        {
            "agency": "MDA-Legislation-List",
            "name": "Medical Device Authority — Legislation Documents List (MDA Malaysia)",
            "url": "https://www.mda.gov.my/index.php/doc-list/legislation",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY: MDA Malaysia official legislation documents page — lists Act 737, Medical Device Regulations 2012 (P.U.(A) 210/2012), and all subsidiary legislation with download links",
        },
        {
            "agency": "MDA-Act737-PDF",
            "name": "Medical Device Act 2012 (Act 737) — Full Text PDF",
            "url": "https://www.ummc.edu.my/files/ethic/Medical%20Device%20Act%202012.pdf",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Act 737 full text PDF — University of Malaya Medical Centre mirror; Act 737 s.14 requires manufacturer licence; ISO 13485/MDSAP/FDA QSR or MHLW 169 accepted as QMS evidence — Jina for PDF",
        },
        {
            "agency": "MDA-Legislation",
            "name": "Medical Device Act 2012 (Act 737) — AGC Laws of Malaysia (lom.agc.gov.my)",
            "url": "https://lom.agc.gov.my/act-detail.php?act=737&lang=BI",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Official AGC Laws of Malaysia page for Act 737 — returns timeline/TOC only (no inline text); full PDF download requires navigation",
        },
    ],
    "以色列 (Israel)": [
        {
            "agency": "Nevo-MD-Regulations-2021",
            "name": "תקנות ציוד רפואי (Medical Equipment Regulations) — Israel MOH regulations portal",
            "url": "https://www.health.gov.il/Subjects/MedicalEquipment/Pages/Regulations.aspx",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: Israel Medical Equipment Regulations 2021 — Israeli MOH portal; Nevo.co.il URLs returned 404/thin; Jina first",
        },
        {
            "agency": "Nevo-MD-Law-2012",
            "name": "Medical Equipment Law 5772-2012 — Israeli Parliament (Knesset) PDF",
            "url": "https://main.knesset.gov.il/EN/activity/Documents/LawsAndReg/MedicalEquipment_5772.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Medical Equipment Law 5772-2012 enabling act — Knesset official PDF; PyMuPDF primary",
        },
        {
            "agency": "MOH-Laws",
            "name": "Israel MOH — Medical Equipment Division (English)",
            "url": "https://www.health.gov.il/English/Topics/MedicalEquipment/Pages/default.aspx",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Israel MOH English medical equipment portal — ISO 13485 registration requirements — Jina first",
        },
        {
            "agency": "MOH-MD-Division",
            "name": "Israel MOH — Medical Equipment Registration Requirements",
            "url": "https://www.health.gov.il/English/Topics/MedicalEquipment/Reg_MD/Pages/default.aspx",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Israel MOH medical device registration page — Jina first",
        },
    ],
    "菲律賓 (Philippines)": [
        {
            "agency": "FDA-PH-RA9711",
            "name": "Republic Act No. 9711 — FDA Act of 2009 (Philippines QMS legal basis)",
            "url": "https://www.fda.gov.ph/wp-content/uploads/2021/04/Republic-Act-No.-9711.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: RA 9711 (FDA Act 2009) mandates FDA Philippines to regulate medical devices; ISO 13485 QMS compliance required for CDRRHR registration — official FDA PH PDF",
        },
        {
            "agency": "FDA-PH-MD",
            "name": "Philippines FDA — Medical Devices Homepage (CDRRHR)",
            "url": "https://www.fda.gov.ph/",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Philippines FDA portal — fallback for regulation updates and guidance",
        },
    ],
    "越南 (Vietnam)": [
        {
            "agency": "LuatVN-Decree98-EN",
            "name": "Decree 98/2021/ND-CP — Medical Device Management (English full text, luatvietnam.vn)",
            "url": "https://english.luatvietnam.vn/decree-98-2021-nd-cp-on-medical-device-management-219088-d1.html",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: Decree 98/2021/ND-CP (8 Nov 2021, effective 1 Jan 2022) — English full text on english.luatvietnam.vn — requires ISO 13485 for Class B/C/D registration — thuvienphapluat.vn blocked by Cloudflare — Jina first",
        },
        {
            "agency": "LuatVN-Consolidated-PDF",
            "name": "Vietnam Medical Device Decree 98/2021 — Consolidated text (vbhn_byt_2024 PDF, luatvietnam.vn)",
            "url": "https://static3.luatvietnam.vn/uploaded/vietlawfile/2024/7/04_vbhn_byt_2024_incom_010724105049.pdf",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Consolidated Decree 98/2021 as amended (July 2024 consolidation) — PDF on luatvietnam.vn static server — Jina for PDF",
        },
        {
            "agency": "LuatVN-Decree04-2025-EN",
            "name": "Decree 04/2025/ND-CP — Amendments to Decree 98/2021 (English, english.luatvietnam.vn)",
            "url": "https://english.luatvietnam.vn/y-te/decree-04-2025-nd-cp-management-of-medical-equipment-385264-d1.html",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Latest 2025 amendment to Vietnam Decree 98/2021 — English full text on luatvietnam.vn — Jina first",
        },
        {
            "agency": "DMEC-MOH",
            "name": "Nghị định 98/2021/NĐ-CP — Về quản lý trang thiết bị y tế (vanban.chinhphu.vn Official Portal)",
            "url": "https://vanban.chinhphu.vn/?pageid=27160&docid=204442",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Official government portal page — returns only metadata/header; full text PDF at datafiles.chinhphu.vn; thuvienphapluat.vn preferred for content",
        },
    ],
    "哥倫比亞 (Colombia)": [
        {
            "agency": "INVIMA-Decreto4725",
            "name": "Decreto 4725 de 2005 — Régimen de registros sanitarios de dispositivos médicos (INVIMA Normograma via Jina)",
            "url": "https://normograma.invima.gov.co/normograma/compilacion/docs/decreto_4725_2005.htm",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "PRIMARY QMS regulation: Decreto 4725/2005 — normograma.invima.gov.co times out on direct httpx; Jina Reader bypasses slow server; Article 8 requires BPM/ISO 13485 compliance",
        },
        {
            "agency": "INVIMA-Decreto4725-PDF",
            "name": "Decreto 4725 de 2005 — PDF via Función Pública Colombia (gov.co)",
            "url": "https://www.funcionpublica.gov.co/eva/gestornormativo/norma.php?i=18841",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "Decreto 4725/2005 alternative source — Función Pública normativa database — Jina first",
        },
        {
            "agency": "INVIMA-MD",
            "name": "INVIMA — Dispositivos Médicos y Equipos Biomédicos (Medical Devices portal)",
            "url": "https://www.invima.gov.co/productos-vigilados/dispositivos-medicos/dispositivos-medicos-equipos-biomedicos",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "INVIMA medical device portal — fallback for regulatory updates — SSL/TLS issues — Jina Reader fallback",
        },
    ],
    "俄羅斯 (Russia)": [
        {
            "agency": "ZakonRF-1684",
            "name": "Постановление Правительства РФ № 1684 от 30.11.2024 — полный текст статей (zakonrf.info)",
            "url": "https://www.zakonrf.info/postanovlenie-pravitelstvo-rf-1684-30112024/",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY: Government Decree No. 1684 (30 Nov 2024, effective 1 Mar 2025) — Правила государственной регистрации медицинских изделий — full article text on zakonrf.info — Jina first",
        },
        {
            "agency": "Consultant-1684",
            "name": "ПРАВИЛА ГОСУДАРСТВЕННОЙ РЕГИСТРАЦИИ МЕДИЦИНСКИХ ИЗДЕЛИЙ — КонсультантПлюс (cons_doc_LAW_491966)",
            "url": "https://www.consultant.ru/document/cons_doc_LAW_491966/e8558fcb2d8260bcdf939ff0403d32dfcc37110c/",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "Decree 1684 Rules full text on ConsultantPlus — major Russian legal database — may require Jina to bypass paywall — Jina first",
        },
        {
            "agency": "RZN-MD",
            "name": "Постановление Правительства РФ № 1684 — Контур.Норматив",
            "url": "https://normativ.kontur.ru/document?moduleId=1&documentId=483928",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "Kontur.Normativ — returns only document title without content; kept as fallback reference",
        },
    ],
    "埃及 (Egypt)": [
        {
            "agency": "EDA-MD-Guide",
            "name": "EDA — Regulatory Guideline for Registering Medical Devices (ISO 13485 requirement)",
            "url": "https://edaegypt.gov.eg/media/j3hdl0l2/5_regulatory-guideline-for-procedures-of-registering-imported-and-local-medical-devices-holding-international-quali.pdf",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation guideline: EDA Egypt regulatory procedures for medical devices holding international quality certificates (ISO 13485) — official EDA PDF (515KB confirmed)",
        },
        {
            "agency": "EDA",
            "name": "Egyptian Drug Authority — Homepage",
            "url": "https://www.edaegypt.gov.eg/",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "EDA portal — fallback for regulation announcements",
        },
    ],
    "智利 (Chile)": [
        {
            "agency": "ISP-Regulations",
            "name": "Decreto Supremo N° 825 de 1998 — Reglamento de Control de Productos y Elementos de Uso Médico (BCN Ley Chile)",
            "url": "https://www.bcn.cl/leychile/Navegar?idNorma=141005",
            "tier": 3,
            "strategy": "html",
            "crawl_delay": 5,
            "note": "PRIMARY QMS regulation: Chile Decreto 825/1998 (Ministerio de Salud) — Reglamento de Control de Productos y Elementos de Uso Médico — GMP standards reference US 21 CFR Part 820 — BCN Ley Chile full text (requires JS — Jina Reader fallback)",
        },
    ],
    "阿聯酋 (UAE)": [
        {
            "agency": "MOHAP-FDL38-2024",
            "name": "UAE Federal Decree-Law No. 38 of 2024 — Medical Products (uaelegislation.gov.ae official download)",
            "url": "https://uaelegislation.gov.ae/en/legislations/2751/download",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "PRIMARY QMS regulation: Federal Decree-Law No. 38 of 2024 — uaelegislation.gov.ae returns HTTP 403; Jina Reader attempted; pre-written profile is last-resort fallback — ISO 13485:2016 required for manufacturer registration",
        },
        {
            "agency": "MOHAP",
            "name": "UAE Ministry of Health and Prevention — Homepage",
            "url": "https://mohap.gov.ae/",
            "tier": 2,
            "strategy": "html",
            "crawl_delay": 3,
            "note": "MOHAP portal — fallback for regulation updates and announcements",
        },
    ],
}


# ============================================================
# Pre-Written Regulatory Profiles (Fallback Content — Source A)
# ============================================================
# Used when all dynamic crawl tiers and DuckDuckGo fail (Source B exhausted).
# Keys: (region_key, agency_key) — must match REGION_SITES entries.
# Results using this source will carry content_source = "pre-written".

REGION_PROFILES: dict[tuple[str, str], str] = {
    ("台灣 (Taiwan)", "TFDA-QMS"): """\
# 醫療器材品質管理系統準則 (Taiwan QMS Criteria)
# Last reviewed: 2026-05-14

**主管機關**: 衛生福利部食品藥物管理署 (TFDA)
**法規編號**: pcode L0030116
**發布日期**: 2021-04-14 / **生效日期**: 2021-05-01

## 法規架構（共 79 條）

| 章節 | 條號 | 主題 |
|---|---|---|
| 第一章 | 1–5 | 總則、適用範圍、術語定義 |
| 第二章 | 6–9 | 品質管理系統（文件管制、紀錄管制） |
| 第三章 | 10–16 | 管理責任（品質政策、管理審查） |
| 第四章 | 17–21 | 資源管理（人員、基礎設施） |
| 第五章 | 22–63 | 產品實現（設計管制、採購、製造、滅菌） |
| 第六章 | 64–79 | 量測、分析及改善（CAPA、內稽） |

## 與 ISO 13485:2016 關係

本準則為 ISO 13485:2016 之等效國內轉換；持有效期內之 ISO 13485 認證可作為符合本準則之佐證。

## 適用對象

依醫療器材管理法申請製造許可證之第二至四等級醫療器材製造業者。
""",
    ("台灣 (Taiwan)", "TFDA-QMS-EN"): """\
# Medical Device Quality Management System Guidelines — Taiwan (English)
# Last reviewed: 2026-05-14

**Authority**: Taiwan TFDA | **Law Code**: L0030116 | **Effective**: 2021-05-01

Taiwan's QMS Criteria (79 articles) is the national equivalent of ISO 13485:2016,
mandatory for Class II–IV medical device manufacturing licence applications.

## Key Clauses Mapped to ISO 13485:2016

| ISO 13485 Clause | Articles | Topic |
|---|---|---|
| 4 | 6–9 | QMS General Requirements |
| 5 | 10–16 | Management Responsibility |
| 6 | 17–21 | Resource Management |
| 7 | 22–63 | Product Realization |
| 8 | 64–79 | Measurement, Analysis, Improvement |

A valid ISO 13485:2016 certificate from an accredited body is accepted as evidence of compliance.
""",
    ("美國 (USA)", "FDA-QMSR"): """\
# FDA Quality Management System Regulation (QMSR) — 21 CFR Part 820
# Last reviewed: 2026-05-14

**Authority**: U.S. Food and Drug Administration (FDA / CDRH)
**Citation**: 21 CFR Part 820
**Final Rule Published**: February 2, 2024 (89 FR 7496)
**Effective Date**: February 2, 2026

## Overview

The QMSR replaces the 1996 Quality System Regulation (QSR) by incorporating
ISO 13485:2016 by reference. Manufacturers must comply with ISO 13485:2016
as interpreted through the QMSR preamble and FDA-specific requirements.

## FDA-Specific Requirements (21 CFR 820)

| Section | Topic |
|---|---|
| 820.10 | Scope — applies to finished device manufacturers |
| 820.20 | Quality Management System — ISO 13485:2016 incorporated by reference |
| 820.30 | Design Controls — design history file (DHF) |
| 820.40 | Complaints — MDR linkage |
| 820.50 | CAPA |
| 820.60 | Records — device master record (DMR), device history record (DHR) |

## Relationship to ISO 13485:2016

FDA QMSR § 820.20 requires conformance with ISO 13485:2016.
Where ISO 13485 is silent, FDA guidance documents supplement requirements.
""",
    ("美國 (USA)", "eCFR-820"): """\
# 21 CFR Part 820 — Quality Management System Regulation (QMSR) Full Text
# Last reviewed: 2026-05-14

**Source**: Electronic Code of Federal Regulations (eCFR)
**Effective**: February 2, 2026 | **Replaces**: Former QSR (21 CFR 820, 1996)

The full regulatory text of 21 CFR Part 820 is available on eCFR.gov.
It incorporates ISO 13485:2016 by reference and adds FDA-specific requirements
for device history records, design history files, and complaint handling linked to MDR.
""",
    ("歐盟 (EU)", "EUR-Lex-MDR-OJ"): """\
# EU Medical Device Regulation (EU MDR) 2017/745
# Last reviewed: 2026-05-14

**Citation**: Regulation (EU) 2017/745 of the European Parliament and of the Council
**Published**: OJ L 117, 5.5.2017 | **Full Application Date**: May 26, 2021

## Article 10 — General Obligations of Manufacturers

Manufacturers shall establish, document, implement, maintain, keep up to date and
continually improve a quality management system that ensures compliance with this
Regulation in the most effective manner and in a manner that is proportionate to the
risk class and the type of device.

The QMS shall cover all parts of and requirements on the manufacturer's organisation
that deal with the quality of processes, procedures and devices. It shall govern the
structure, responsibilities, procedures, processes and management resources required
to implement the principles and actions necessary to achieve compliance.

### 10(9) QMS Minimum Coverage

- (a) a strategy for regulatory compliance, including compliance with conformity
  assessment procedures and procedures for management of modifications to the devices
- (b) identification of applicable general safety and performance requirements (GSPR, Annex I)
- (c) responsibility of the management
- (d) resource management, including selection and control of suppliers and sub-contractors
- (e) risk management (as set out in Section 3 of Annex I)
- (f) clinical evaluation (Annex XIV), including PMCF
- (g) product realisation, including planning, design, development, production and
  service provision
- (h) verification of the UDI assignments made
- (i) setting-up, implementation and maintenance of a post-market surveillance system
- (j) handling communication with competent authorities, notified bodies, other
  economic operators, customers and/or other stakeholders
- (k) processes for reporting of serious incidents and field safety corrective actions
- (l) management of corrective and preventive actions and verification of their
  effectiveness
- (m) processes for monitoring and measurement of output, data analysis and product
  improvement

## QMS Requirements — Annex IX, Section 2

Manufacturers of Class IIa, IIb, and III devices must implement and maintain a QMS
assessed by a Notified Body under Annex IX.

### Annex IX §2.2 — QMS Assessment Scope

The notified body shall audit the manufacturer's QMS to verify:
- Compliance with the applicable requirements of this Regulation
- That the device-specific technical documentation satisfies Annex II/III requirements
- Policies, procedures, instructions, and records for production and quality control

### QMS Elements Required (Annex IX, §2.2)

- Regulatory strategy and compliance procedures
- Design and development management (Annex I GSPR compliance)
- Production and post-production activities (PMS, PMCF)
- Risk management per ISO 14971
- Clinical evaluation per Annex XIV
- Document and records control
- Management responsibility and internal audit
- Supplier and sub-contractor control
- CAPA (corrective and preventive actions)
- PMS system and PSUR/SSCP
- Vigilance and field safety corrective actions (FSCA)

## Equivalence to ISO 13485

EU MDR does not directly reference ISO 13485, but ISO 13485:2016 + EN ISO 13485:2016
is widely accepted by Notified Bodies as the harmonised standard covering QMS requirements.
MDCG 2019-11 Rev.1 provides guidance on the QMS documentation notified bodies require.

## Class I Self-Declaration

Class I manufacturers self-declare conformity; no Notified Body QMS audit required
(except sterile, measuring, or reusable surgical devices — Class Im/Is/Ir).

## Annex XI — QMS for Production Quality Assurance (Class IIb/III)

Annex XI Part A covers production quality assurance as an alternative to Annex IX.
Manufacturers must maintain and apply a QMS approved by a notified body covering:
- Manufacturing, final product inspection and testing
- Examination, testing, or traceability of implantable devices
""",
    ("歐盟 (EU)", "MDCG"): """\
# EU MDR MDCG Guidance — QMS Requirements
# Last reviewed: 2026-05-14

**Source**: Medical Device Coordination Group (MDCG) endorsed guidance documents
**Primary Document**: MDCG 2019-11 Rev.1 "Guidance on the QMS documentation that notified
bodies should require from manufacturers" (Annex IX, Regulation (EU) 2017/745)

## MDCG 2019-11 Rev.1 — QMS Documentation Requirements

Under MDR Annex IX §2.2, notified bodies must assess the manufacturer's QMS.
MDCG 2019-11 specifies what documents NB auditors must review.

### Mandatory QMS Documentation

| Category | Required Documents |
|---|---|
| Quality Manual | Scope, exclusions, documented procedures or references |
| Quality Policy | Top management signed, objectives, communication |
| Management Review | Records of review inputs/outputs, frequency |
| Document Control | Procedure for approval, review, revision, distribution |
| Record Control | Identification, storage, protection, retrieval, retention, disposition |
| Risk Management | ISO 14971 risk management file per device, risk policy |
| Design & Development | Procedures for planning, inputs, outputs, review, verification, validation, transfer |
| Clinical Evaluation | CER (Annex XIV Part A) + PMCF plan (Annex XIV Part B) |
| PMS | PMSP (post-market surveillance plan), PSUR/SSCP, complaint handling |
| Purchasing | Supplier evaluation, approved supplier list, purchasing controls |
| Production | Work instructions, in-process controls, device history record |
| Nonconforming Product | Identification, segregation, disposition, rework controls |
| CAPA | Root cause analysis, corrective/preventive action, effectiveness verification |
| Internal Audit | Audit programme, procedures, records, management follow-up |
| Training | Competency requirements, training records, qualification evidence |
| Labelling | Labelling controls, UDI assignment records |
| Sterilisation | (if applicable) validation records, sterility assurance level documentation |
| Vigilance | MDR Article 87–92 serious incident reporting, FSCA procedures |

### MDCG 2019-11 §4 — Device-Specific Technical Documentation

In addition to QMS, the NB assesses the technical documentation for at least one
representative device per device group. This includes:
- Device description and specification (Annex II §1)
- Information to be supplied by the manufacturer (Annex II §2)
- Design and manufacturing information (Annex II §3)
- GSPR compliance (Annex I, referenced in Annex II §4)
- Risk/benefit analysis and risk management (Annex II §5)
- Product verification and validation (Annex II §6)
- Post-market surveillance (Annex II §7)
- Declaration of Conformity (Annex IV)

## MDCG 2020-1 — Guidance on Clinical Evaluation (Annex XIV)

QMS procedures must include a documented clinical evaluation process. Manufacturers
must demonstrate continuous compliance through the PMCF programme.

## MDCG 2021-25 — GSPR Application (Annex I)

The QMS shall document how each applicable General Safety and Performance Requirement
is met, either by compliance with a harmonised standard (EN ISO 13485:2016, etc.) or
by an alternative method with equivalent evidence.

## Key MDCG Guidance Documents for QMS

| Document | Topic |
|---|---|
| MDCG 2019-11 Rev.1 | QMS documentation for NB assessment |
| MDCG 2019-13 Rev.1 | Qualification and classification of software |
| MDCG 2020-1 Rev.1 | Clinical evaluation |
| MDCG 2020-7 | PSUR (Post-Market Surveillance Report) |
| MDCG 2021-25 | GSPR application guide |
| MDCG 2022-21 | Notified body oversight of manufacturers |
| MDCG 2023-1 | UDI guidance |
""",
    ("英國 (UK)", "UK-MDR-2002"): """\
# The Medical Devices Regulations 2002 (SI 2002/618) — UK QMS Requirements
# Last reviewed: 2026-05-14

**Citation**: SI 2002/618 (as amended)
**Authority**: Medicines and Healthcare products Regulatory Agency (MHRA)
**Post-Brexit Basis**: UKCA marking (replacing CE for Great Britain from July 2024)

## QMS Requirements

| Device Class | Conformity Route | QMS Requirement |
|---|---|---|
| Class I | Self-declaration | No Notified Body audit |
| Class IIa | MHRA-approved body (UK Approved Body) | ISO 13485 QMS audit |
| Class IIb/III | Full QA (Schedule 3, Part III) | Full ISO 13485 QMS audit |
| Active Implantables | Full QA (Schedule 2, Part II) | Full ISO 13485 QMS audit |

## Key Provisions

- **Schedule 3, Part III**: Full quality assurance procedure equivalent to EU MDD Annex II
- Manufacturers must maintain a QMS covering design, manufacture, and final inspection
- ISO 13485:2016 is the accepted standard for demonstrating compliance

## MHRA Guidance

MHRA accepts ISO 13485 certification from a UK Approved Body (UKAB) as primary evidence
of QMS compliance for UKCA marking.
""",
    ("日本 (Japan)", "eGov-QMS-169"): """\
# 医療機器 QMS 省令 — MHLW Ministerial Ordinance No. 169 (QMS省令)
# Last reviewed: 2026-05-14

**正式名称**: 医療機器及び体外診断用医薬品の製造管理及び品質管理の基準に関する省令
**根拠法**: 医薬品、医療機器等の品質、有効性及び安全性の確保等に関する法律（薬機法）第 23 条の 2 の 5
**省令番号**: 平成 16 年厚生労働省令第 169 号
**最終改正**: 令和 3 年（2021 年）— ISO 13485:2016 対応改正

## 章構成

| 章 | 条 | 内容 |
|---|---|---|
| 第一章 | 1–4 | 総則 |
| 第二章 | 5–8 | 品質管理監督システム（文書管理） |
| 第三章 | 9–14 | 管理監督者の責務 |
| 第四章 | 15–19 | 資源の管理 |
| 第五章 | 20–63 | 製品実現 |
| 第六章 | 64–79 | 測定、分析及び改善 |

## ISO 13485:2016 対応

2021 年改正により、QMS 省令は ISO 13485:2016 と整合化。
第一種・第二種医療機器製造業者は QMS 適合性調査（PMDA 実施）を受ける義務がある。

## 適用対象

薬機法に基づく医療機器製造販売業・製造業の許可要件として適用。
""",
    ("中國 (China)", "NMPA-GMP-CN"): """\
# Good Manufacturing Practice for Medical Devices — China (Revised 2025)
# Last reviewed: 2026-05-14

**Authority**: National Medical Products Administration (NMPA)
**Announcement**: NMPA [2025] No. 107 (issued November 2025)
**Effective Date**: November 1, 2026 | **Replaces**: 2014 GMP Regulation

## Structure (15 Chapters, 132 Articles)

| Chapter | Articles | Topic |
|---|---|---|
| 1 | 1–5 | General Provisions |
| 2 | 6–14 | Quality Management System |
| 3 | 15–19 | Organization and Personnel |
| 4 | 20–25 | Infrastructure and Facilities |
| 5 | 26–29 | Equipment |
| 6 | 30–36 | Design and Development |
| 7 | 37–47 | Procurement and Suppliers |
| 8 | 48–62 | Production Management |
| 9 | 63–72 | Quality Control |
| 10 | 73–80 | Sales and After-Sales |
| 11 | 81–87 | Adverse Event Monitoring |
| 12 | 88–94 | Customer Feedback and Complaints |
| 13 | 95–101 | Nonconforming Product |
| 14 | 102–111 | Records and Documents |
| 15 | 112–132 | Supplementary Provisions |

## Relationship to ISO 13485

China GMP is structurally aligned with ISO 13485:2016 but is administered domestically by NMPA.
Foreign manufacturers must obtain NMPA GMP compliance certification for import registration.
""",
    ("韓國 (Korea)", "MFDS-KGMP"): """\
# Korea Good Manufacturing Practice for Medical Devices (K-GMP)
# Last reviewed: 2026-05-14

**Korean Name**: 의료기기 제조 및 품질관리 기준
**Authority**: Ministry of Food and Drug Safety (MFDS)
**Legal Basis**: Medical Devices Act Article 6 (Act No. 14330)
**Latest Revision**: 2024 (aligned with ISO 13485:2016 Third Edition)
**Structure**: 9 Chapters, 79 Articles

## Chapter 2 — Quality Management System (ISO 13485 §4 equivalent)

Article 4 (§4.1): The manufacturer shall establish, document, implement, and maintain a quality
management system and continually improve its effectiveness in accordance with the requirements
of this standard. The manufacturer shall determine the processes needed for the QMS.

Article 5 (§4.2.1): Documentation shall include: (a) documented quality policy and quality
objectives; (b) a quality manual; (c) documented procedures and records required by this
standard; (d) documents needed to ensure the effective planning, operation, and control of
processes.

Article 6 (§4.2.2): The manufacturer shall establish and maintain a quality manual that includes:
the scope of the QMS, documented procedures or reference to them, and a description of the
interaction between the processes of the QMS.

Article 7 (§4.2.3): Document control procedures shall include approval, review, update,
identification of changes, availability at point of use, and prevention of unintended use of
obsolete documents.

Article 8 (§4.2.4): Records shall be controlled. The manufacturer shall establish procedures
for identification, storage, protection, retrieval, retention time, and disposition of records.
Retention period: at least the lifetime of the device but no less than 2 years from product
release date (Korea-specific retention requirement).

Article 9 (§4.2.5): Technical documentation (Device Master Record equivalent) shall be
maintained for each device type.

## Chapter 3 — Management Responsibility (ISO 13485 §5 equivalent)

Article 10 (§5.1): Top management shall provide evidence of commitment to the development
and implementation of the QMS and maintaining its effectiveness.

Article 11 (§5.2): Top management shall ensure that customer requirements are determined
and met.

Article 12 (§5.3): Top management shall establish the quality policy. The quality policy
shall be appropriate to the purpose of the organization, include a commitment to comply
with requirements and maintain the effectiveness of the QMS.

Article 13 (§5.4.1): Top management shall ensure that quality objectives are established
at relevant functions and levels within the organization.

Article 14 (§5.4.2): Top management shall ensure that the planning of the QMS is carried
out.

Article 15 (§5.5.1): Top management shall ensure that responsibilities and authorities
are defined and communicated.

Article 16 (§5.5.2): Top management shall appoint a member of management who, irrespective
of other responsibilities, shall have responsibility and authority for ensuring QMS processes
are established, implemented, and maintained (Management Representative).

Article 17 (§5.5.3): Top management shall ensure that appropriate communication processes
are established and that communication takes place regarding the effectiveness of the QMS.

Article 18 (§5.6): Management shall conduct reviews of the QMS at planned intervals.
Management review inputs shall include: audit results, customer feedback, process performance,
product conformity, preventive and corrective actions, previous review follow-up, and
regulatory changes.

## Chapter 4 — Resource Management (ISO 13485 §6 equivalent)

Article 19 (§6.1): The organization shall determine and provide the resources needed to
implement, maintain, and continually improve the effectiveness of the QMS.

Article 20 (§6.2): Personnel performing work affecting product quality shall be competent
on the basis of appropriate education, training, skills, and experience.

Article 21 (§6.3): The organization shall determine, provide, and maintain the infrastructure
needed to achieve conformity to product requirements. Infrastructure includes buildings,
workspace, process equipment, and supporting services.

Article 22 (§6.4): The organization shall determine and manage the work environment needed
to achieve conformity to product requirements, including contamination control for sterile
or clean room environments.

## Chapter 5 — Product Realization (ISO 13485 §7 equivalent)

Article 23 (§7.1): The organization shall plan and develop the processes needed for product
realization. This shall be consistent with the requirements of other processes of the QMS,
including risk management (per MFDS Medical Device Risk Management Guidelines).

Article 24 (§7.2.1): Customer-related requirements shall be determined, including
requirements for the device, requirements not stated by the customer but necessary for
intended use, applicable regulatory requirements, and any additional requirements.

Article 25 (§7.2.2): The organization shall review the requirements related to the product
prior to commitment to supply.

Article 26 (§7.2.3): The organization shall determine and implement effective arrangements
for communicating with customers regarding complaints and advisory notices.

Article 27–30 (§7.3): Design and development planning, inputs, outputs, review,
verification, validation, transfer, changes, and design history file (DHF) requirements.
The DHF shall be maintained for each device type.

Article 31 (§7.4.1): Supplier evaluation and selection criteria shall be established.
Purchasing information shall describe the product to be purchased. Purchased product shall
be verified.

Articles 32–36 (§7.5.1–7.5.11): Production and service provision controls, cleanliness,
identification and traceability (lot number mandatory), customer property, preservation,
process validation, and sterile device requirements. Implantable device traceability
records shall be maintained for the lifetime of the device.

Article 37 (§7.6): Monitoring and measuring equipment shall be calibrated at specified
intervals or prior to use. Calibration records shall be maintained.

## Chapter 6 — Measurement, Analysis and Improvement (ISO 13485 §8 equivalent)

Article 38 (§8.1): The organization shall plan and implement the monitoring, measurement,
analysis, and improvement processes needed.

Article 39 (§8.2.1): Feedback from the post-production phase shall be gathered and used
as input to risk management and product monitoring.

Article 40 (§8.2.2): Complaint handling procedures shall be established. Complaints
shall be investigated and records maintained. Korea-specific: adverse event reports to
MFDS within 15 days for serious adverse events, 30 days for all others.

Article 41 (§8.2.3): The manufacturer shall report to MFDS any adverse event that has
occurred inside or outside Korea if the device caused or may have caused serious injury,
death, or a significant public health risk.

Article 42 (§8.2.4): Internal audits shall be conducted at planned intervals to determine
whether the QMS conforms to planned arrangements and requirements of this standard.

Article 43 (§8.2.5–8.2.6): Monitoring and measurement of processes and product quality
shall be performed.

Article 44 (§8.3): Nonconforming product shall be identified and controlled. Rework
and re-inspection procedures shall be documented.

Article 45 (§8.4): Data analysis shall be performed to determine the suitability and
effectiveness of the QMS.

Articles 46–47 (§8.5.2–8.5.3): CAPA procedures shall be established. Root cause analysis
shall be performed. Corrective actions shall be proportionate to the effects of the
nonconformities encountered. Preventive actions shall be determined to eliminate causes
of potential nonconformities.

## Korea-Specific Additional Requirements (beyond ISO 13485)

- **MFDS GMP Certificate**: Class II–IV manufacturers must obtain MFDS GMP certificate
  (or accredited inspection body GMP certificate) renewed every 3 years.
- **UDI (Unique Device Identification)**: Korea adopted UDI system from January 2024.
  Class III/IV implantable devices first, phased rollout through 2027.
- **Post-Market Surveillance Report**: Annual PMS report required for Class II–IV.
- **Re-examination Period**: Class III–IV new devices undergo 6-year re-examination period.
- **Adverse Event Reporting**: Serious injury/death: 15 days; other adverse events: 30 days;
  MDR-equivalent field safety notices must be submitted to MFDS before implementing FSCA.
""",
    ("加拿大 (Canada)", "CMDR-SOR98-282"): """\
# Medical Devices Regulations SOR/98-282 — QMS Requirements (Section 32)
# Last reviewed: 2026-05-14

**Citation**: SOR/98-282, Medical Devices Regulations
**Authority**: Health Canada
**Administering Act**: Food and Drugs Act (R.S.C. 1985, c. F-27)

## Section 32 — Quality Management System Requirement

Section 32 requires that, as a condition of obtaining a Device Licence for Class II, III, or IV
medical devices, the manufacturer must hold a valid ISO 13485 certificate issued by a
registrar accredited by a member body of the International Accreditation Forum (IAF).

Since January 2019, Canada exclusively uses MDSAP (Medical Device Single Audit Program)
as the audit mechanism. Health Canada no longer accepts CMDCAS certificates.

## Licence Classes and QMS Requirements

| Class | Licence Type | QMS Required |
|---|---|---|
| Class I | No Device Licence | None |
| Class II | Device Licence | ISO 13485 via MDSAP |
| Class III | Device Licence | ISO 13485 via MDSAP |
| Class IV | Device Licence | ISO 13485 via MDSAP |

## MDSAP Audit Structure

MDSAP audits cover: Organization, Device Design, Purchasing/Supplier Controls, Production/Service,
Measurement/Monitoring/Analysis, and Improvement — all mapped to ISO 13485:2016 clauses.
""",
    ("澳洲 (Australia)", "TGA-Legislation"): """\
# Therapeutic Goods (Medical Devices) Regulations 2002 — QMS Requirements
# Last reviewed: 2026-05-14

**Citation**: F2002B00237, TG(MD)R 2002 (current consolidated text)
**Authority**: Therapeutic Goods Administration (TGA)
**Administering Act**: Therapeutic Goods Act 1989

## Schedule 3 — Conformity Assessment Procedures

| Device Class | Procedure | QMS Requirement |
|---|---|---|
| Class I (general) | Self-declaration | No third-party QMS audit |
| Class I (sterile/measuring) | Schedule 3, Part 1 | ISO 13485 production QMS audit |
| Class IIa | Schedule 3, Part 2 | ISO 13485 QMS audit |
| Class IIb | Schedule 3, Part 2 | ISO 13485 Full QMS audit |
| Class III / AIMD | Schedule 3, Part 2 | ISO 13485 Full QMS + design audit |

## ISO 13485 as Accepted Standard

TGA accepts ISO 13485:2016 certification from MDSAP-recognised auditing organisations or
accredited certification bodies as evidence of QMS compliance under Schedule 3.

## MDSAP Participation

Australia (TGA) is a full MDSAP participating authority; MDSAP certificates are accepted
in lieu of separate TGA conformity assessment for QMS requirements.
""",
    ("瑞士 (Switzerland)", "MedDO-SR812213"): """\
# Verordnung über Medizinprodukte (MedDO) SR 812.213 — Swiss Medical Devices Ordinance
# Last reviewed: 2026-05-14

**Citation**: SR 812.213
**Authority**: Swissmedic
**In Force**: May 26, 2021 (aligned with EU MDR 2017/745)

## QMS Requirements

Switzerland's MedDO is structurally aligned with EU MDR 2017/745 and IVDR 2017/746.
QMS requirements mirror EU MDR Annex IX (conformity assessment via Notified Body / Swiss Approved Body).

| Device Class | Conformity Route | QMS Requirement |
|---|---|---|
| Class I | Self-declaration | No third-party audit |
| Class IIa / IIb | Annex IX (MedDO) | ISO 13485 QMS audit by Swiss Approved Body |
| Class III / AIMD | Annex IX full | Full QMS + Design Dossier |

## Recognition Agreement

Switzerland has a Mutual Recognition Agreement (MRA) with the EU. CE-marked devices can
circulate in Switzerland, and Swiss Approved Bodies can issue CE/CH marks.

## Swissmedic Role

Swissmedic is the market surveillance authority. Manufacturers must register in the
EUDAMED-compatible Swiss registration system and hold Swissmedic-accepted QMS certificates.
""",
    ("巴西 (Brazil)", "ANVISA-RDC665"): """\
# RDC nº 665/2022 — Boas Práticas de Fabricação para Produtos para Saúde (BGMP)
# Last reviewed: 2026-05-14

**Citation**: Resolução RDC nº 665, de 30 de março de 2022
**Authority**: ANVISA (Agência Nacional de Vigilância Sanitária)
**Effective Date**: May 2, 2022 | **Replaces**: RDC 16/2013

## Structure (15 Chapters)

| Chapter | Articles | Topic |
|---|---|---|
| 1 | 1–5 | Scope and Definitions |
| 2 | 6–14 | Quality Management System |
| 3 | 15–20 | Management Responsibility |
| 4 | 21–25 | Resource Management |
| 5 | 26–36 | Infrastructure and Work Environment |
| 6 | 37–50 | Product Realization / Design Control |
| 7 | 51–60 | Purchasing and Supplier Management |
| 8 | 61–75 | Production and Service |
| 9 | 76–85 | Quality Control |
| 10 | 86–92 | Nonconforming Product |
| 11 | 93–99 | CAPA |
| 12 | 100–109 | Internal Audit |
| 13 | 110–115 | Post-Market Surveillance |
| 14 | 116–122 | Records |
| 15 | 123–132 | Supplementary Provisions |

## Relationship to ISO 13485

RDC 665/2022 mirrors FDA QSR and ISO 13485:2016. Foreign manufacturers must demonstrate
BGMP compliance as part of ANVISA registration (INMETRO or recognised CB scheme accepted).
""",
    ("印度 (India)", "CDSCO-MDR2017"): """\
# Medical Devices Rules 2017 (MDR 2017) — India QMS Requirements
# Last reviewed: 2026-05-14

**Citation**: G.S.R. 78(E), Ministry of Health and Family Welfare, January 31, 2017
**Authority**: Central Drugs Standard Control Organisation (CDSCO)
**Legal Basis**: Drugs and Cosmetics Act 1940 (as amended by Medical Devices Rules 2017)

## Schedule V — Quality Management System

Schedule V of MDR 2017 specifies the essential QMS requirements for medical device manufacturers:

- Design and Development Controls
- Document and Record Control
- Management Review
- Internal Audit
- Risk Management (per ISO 14971)
- Production and Process Controls
- Corrective and Preventive Action (CAPA)
- Post-Market Surveillance

## Device Classes and QMS Requirements

| Class | Risk Level | QMS Audit Required |
|---|---|---|
| Class A | Low | Self-declaration |
| Class B | Low-Moderate | ISO 13485 certificate or Schedule V compliance |
| Class C | Moderate-High | ISO 13485 certificate mandatory |
| Class D | High | ISO 13485 certificate mandatory |

## ISO 13485 Acceptance

CDSCO accepts ISO 13485:2016 certification from NABCB-accredited or IAF MLA member bodies.
Foreign manufacturers must submit ISO 13485 certificate with import licence application.
""",
    ("新加坡 (Singapore)", "SSO-HPR-2010"): """\
# Health Products (Medical Devices) Regulations 2010 (S 436/2010) — Singapore QMS
# Last reviewed: 2026-05-14

> **⚠️ 無獨立 QMS 法規 / NO INDEPENDENT QMS LAW** — Singapore has no standalone QMS
> regulation. QMS requirements derive from HP(MD)R 2010 Regulation 23 and the Third
> Schedule, which incorporate ISO 13485:2016 by reference as the mandatory standard.

**Citation**: S 436/2010 (Health Products Act, Cap 122D)
**Authority**: Health Sciences Authority (HSA)
**Effective**: November 1, 2010 (amended through 2024)

## Regulation 23 — Quality Management System Requirement

23.—(1) A manufacturer of a registrable medical device (Class B, C, or D) shall implement
and maintain a quality management system that conforms to ISO 13485:2016 as a condition
of obtaining and retaining a manufacturer's licence under section 36A of the Act.

23.—(2) The quality management system shall cover all processes related to design and
development, production, installation, and servicing of the device at each manufacturing
site identified in the manufacturer's licence.

23.—(3) The QMS shall include the following documented requirements, consistent with
ISO 13485:2016:
  (a) Quality policy and quality objectives (§5.3, §5.4.1);
  (b) Quality manual (§4.2.2);
  (c) Document control procedures (§4.2.3);
  (d) Record control procedures (§4.2.4);
  (e) Management review procedures (§5.6);
  (f) Resource management including personnel competence and training (§6.2);
  (g) Risk management integrated with product realization (§7.1);
  (h) Design and development controls where applicable (§7.3);
  (i) Purchasing and supplier controls (§7.4);
  (j) Production and service provision controls (§7.5);
  (k) Monitoring and measurement of processes and products (§8.2.5, §8.2.6);
  (l) CAPA procedures (§8.5.2, §8.5.3);
  (m) Complaint handling and post-market surveillance (§8.2.1, §8.2.2).

## Third Schedule — Conformity Assessment Routes

Third Schedule Part 1 (Design Examination): Applies to Class D devices. Requires
full technical documentation review by HSA or recognized Conformity Assessment Body (CAB).

Third Schedule Part 2 (Product Quality Assurance): Applies to Class C/D devices.
Requires QMS assessment — ISO 13485 certificate covering manufacturing site.

Third Schedule Part 3 (Type Examination + QMS): Applies to Class C/D devices via
alternative route combining type-testing with QMS certificate.

## Device Classes and QMS Requirements

| Class | Risk | QMS Requirement |
|---|---|---|
| Class A | Low | No ISO 13485 required; self-declaration of safety |
| Class B | Low-Moderate | ISO 13485 certificate from MDSAP-recognised body (from Jan 2025) |
| Class C | Moderate-High | ISO 13485 certificate; Third Schedule assessment |
| Class D | High | ISO 13485 certificate; HSA design examination |

## HSA QMS Specific Requirements (beyond ISO 13485 base)

- **Complaint Handling (Reg. 41)**: Serious adverse events to HSA within 10 days;
  non-serious adverse events within 30 days.
- **Field Safety Corrective Actions (FSCA) (Reg. 42)**: FSCA notification to HSA before
  implementing; Field Safety Notice (FSN) to be issued.
- **Post-Market Surveillance (Reg. 39)**: PMS system required for all Class B/C/D devices;
  trend reporting to HSA for Class C/D.
- **UDI (Reg. 23A)**: UDI requirement phased in from 2024–2027; Class D first.
- **Authorised Local Correspondent**: Foreign manufacturer must appoint a Singapore-registered
  Company to act as correspondent (local agent responsibility).

## From January 1, 2025 — MDSAP Mandatory

HSA accepts only MDSAP audit certificates from MDSAP participant auditing organisations
(BSI, SGS, DNV, TÜV SÜD, DEKRA, Bureau Veritas) for new and renewal manufacturer licence
applications for Class B, C, and D devices.
""",
    ("沙烏地阿拉伯 (Saudi Arabia)", "SFDA-MDS-REQ10"): """\
# SFDA Requirements for Inspections and Quality Management System — MDS-REQ10
# Last reviewed: 2026-05-14

**Citation**: MDS-REQ10 (SFDA Medical Devices Sector, Version 2.0, January 2021)
**Authority**: Saudi Food and Drug Authority (SFDA), Medical Devices Sector
**Legal Basis**: Medical Devices Law (Royal Decree M/65, 2017); SFDA Regulations
**Supplementary**: MDS-G024 — SFDA Guidance for ISO 13485 Requirements with SFDA Regulations

## Section 3 — QMS General Requirements (maps to ISO 13485 §4)

3.1 The establishment shall establish, document, implement, and maintain a QMS and
continually improve its effectiveness in accordance with ISO 13485:2016.

3.2 Documentation requirements shall include:
  (a) Quality policy (§5.3) and quality objectives (§5.4.1);
  (b) Quality manual describing QMS scope and process interactions (§4.2.2);
  (c) Documented procedures required by ISO 13485:2016 (§4.2.1);
  (d) Records required by ISO 13485:2016 (§4.2.4).

3.3 Document control: The establishment shall establish a documented procedure for
document control including approval, review, revision, identification, and accessibility
of documents (§4.2.3). Records shall be legible, identifiable, and retrievable.

3.4 Device Master Record: Technical file / Device Master Record shall be established
and maintained for each device or device family (§4.2.5). Minimum retention: 5 years
from date of manufacture, or lifetime of the device, whichever is longer.

## Section 4 — Management Responsibility (maps to ISO 13485 §5)

4.1 Top management shall demonstrate commitment by establishing quality policy and
objectives, conducting management reviews, and ensuring availability of resources (§5.1).

4.2 Customer focus: The establishment shall determine customer requirements and ensure
they are met (§5.2). Post-market requirements shall be integrated into QMS.

4.3 Quality policy shall be appropriate, include commitment to compliance and continual
improvement, be communicated and understood at all levels (§5.3).

4.4 Management representative shall be appointed with responsibility for QMS (§5.5.2).

4.5 Management review shall be conducted at least annually. Records shall be maintained.
Inputs include: internal audit results, customer feedback, process/product monitoring,
CAPA status, and changes affecting QMS (§5.6.2).

## Section 5 — Resource Management (maps to ISO 13485 §6)

5.1 Resources shall be provided for QMS implementation and improvement (§6.1).

5.2 Personnel performing work affecting product quality shall be competent based on
education, training, skills, and experience. Training records shall be maintained (§6.2).

5.3 Infrastructure (buildings, equipment, utilities) shall be determined, provided, and
maintained. Maintenance schedules shall be documented (§6.3).

5.4 Work environment including contamination control shall be managed. Environmental
monitoring records required for sterile manufacturing (§6.4).

## Section 6 — Product Realization (maps to ISO 13485 §7)

6.1 Planning of product realization shall include quality objectives, processes, documents,
resources, and risk management activities (§7.1). Risk management per ISO 14971 required.

6.2 Customer requirements shall be determined including applicable regulatory requirements
(Saudi FDA licensing requirements, SFDA device classification rules) (§7.2.1).

6.3 Design and development controls required for Class B, C, D devices: planning, inputs,
outputs, review, verification, validation, transfer, changes, and design history file
(§7.3). DHF shall be maintained for the lifetime of the device.

6.4 Purchasing: Approved supplier list shall be maintained. Supplier evaluation criteria
shall be defined. Incoming inspection/verification procedures shall be documented (§7.4).

6.5 Production and service: Manufacturing controls shall include documented procedures,
work instructions, monitoring and control criteria, and process parameters. Process
validation required for processes where output cannot be verified by subsequent monitoring
(§7.5.1, §7.5.6). Traceability of device to raw material lot numbers required (§7.5.3).

6.6 Sterile devices: Additional requirements for sterility assurance, sterilization
validation, bioburden testing, and sterilization batch records (§7.5.7).

6.7 Calibration: Monitoring and measuring equipment shall be calibrated at specified
intervals. Calibration records and uncertainty of measurement shall be documented (§7.6).

## Section 7 — Measurement, Analysis and Improvement (maps to ISO 13485 §8)

7.1 Feedback system: Post-market surveillance process shall be established and maintained.
PMS data shall be input to risk management and continuous improvement (§8.2.1).

7.2 Complaint handling: Written procedure for receiving, recording, evaluating, and
investigating complaints required. SFDA-specific: complaints involving risk of death or
serious injury must be reported to SFDA within 15 calendar days (§8.2.2, §8.2.3).

7.3 Adverse event reporting: Saudi-specific requirement — all serious adverse events
(actual and near-misses involving risk of serious injury/death) must be reported to
SFDA Medical Devices Vigilance System (EMDAD) within 15 days. Periodic safety reports
required annually for implantable and Class D devices.

7.4 Internal audit: Documented audit program based on status and importance of processes
and results of previous audits. Audit records and nonconformity reports (§8.2.4).

7.5 Process and product monitoring: Statistical techniques may be used where appropriate.
Acceptance criteria shall be specified in quality plans or documented procedures (§8.2.5–8.2.6).

7.6 CAPA: Documented procedures for CAPA. Root cause analysis required. Effectiveness
verification required. CAPA records including description of nonconformity, root cause,
corrective/preventive action taken, and effectiveness check results (§8.5.2–8.5.3).

## SFDA-Specific Additional Requirements (MDS-REQ10 beyond ISO 13485)

- **Establishment Licence Renewal**: Every 2 years; QMS compliance re-verification required.
- **SFDA Inspector Access**: Manufacturers must provide access to SFDA QMS inspectors for
  unannounced on-site inspections at any manufacturing site serving Saudi Arabia.
- **Authorized Local Agent**: Foreign manufacturers must appoint a SFDA-registered Saudi
  establishment as authorized agent responsible for regulatory compliance.
- **Product Registration (Device Licence)**: Via GHAD online system; Class B/C/D require
  QMS certificate; Class D requires additional technical dossier review.
- **Arabic Labelling**: Labels and IFU must be in Arabic and English for Saudi market.
- **EMDAD Vigilance System**: Electronic submission of adverse events and FSCA through SFDA
  EMDAD portal mandatory.
""",
    ("泰國 (Thailand)", "Thai-FDA-GMP"): """\
# Thailand Medical Device QMS — GMP Notification B.E. 2566 (2023)
# Last reviewed: 2026-05-14

**Authority**: Thai Food and Drug Administration (Thai FDA), Ministry of Public Health
**Legal Basis**: Medical Devices Act B.E. 2562 (2019)
**GMP Notification**: Notification of Ministry of Public Health on GMP for Medical Devices B.E. 2566 (2023)

## QMS Requirements

Thailand's GMP Notification B.E. 2566 aligns with ISO 13485:2016 and requires:

| Requirement | Scope |
|---|---|
| Quality Management System | Document control, management review |
| Design Controls | DHF, design verification & validation |
| Production Controls | Process validation, traceability |
| CAPA | Root cause analysis, effectiveness verification |
| Post-Market Surveillance | Vigilance reporting, periodic safety updates |

## Device Classes and Compliance

| Class | Registration | QMS Required |
|---|---|---|
| Class 1 | Notification | Basic QMS |
| Class 2 | Licence | ISO 13485 certificate or GMP inspection |
| Class 3 | Licence + pre-market review | ISO 13485 mandatory |

## ISO 13485 Acceptance

Thai FDA accepts ISO 13485:2016 certificates from IAF MLA member accreditation bodies.
MDSAP certificates are accepted from MDSAP-recognised organisations.
""",
    ("紐西蘭 (New Zealand)", "Medsafe-MD-Legislation"): """\
# New Zealand Medical Device Regulation — Medicines Act 1981 & GMP Code
# Last reviewed: 2026-05-14

> **⚠️ 無獨立 QMS 法規 / 独立QMS法令なし / NO INDEPENDENT QMS LAW** — New Zealand has no standalone QMS regulation equivalent to EU MDR Annex IX or FDA 21 CFR Part 820. The Therapeutic Products Act 2023 was REPEALED in December 2024; current QMS requirements derive from the NZ Code of GMP (Part 7) referencing ISO 13485:2016, under the Medicines Act 1981 framework. A new Medical Products Bill is in development.

**Authority**: Medsafe (Medicines and Medical Devices Safety Authority)
**Legislation**: Medicines Act 1981 (Part 5A, as amended)
**Regulations**: Medicines (Database of Medical Devices) Regulations 2003

## QMS Requirements

New Zealand requires compliance with the New Zealand Code of Good Manufacturing Practice (NZ GMP Code)
which references ISO 13485:2016 as the applicable standard for medical devices.

### Key Requirements (NZ GMP Code, Part 7 — Medical Devices)

- Quality Management System per ISO 13485:2016
- Design control (for Class IIa and above)
- Document and record management
- Internal audit
- CAPA
- Post-market surveillance

## Device Classes and QMS

| Class | Medsafe Requirement | ISO 13485 |
|---|---|---|
| Class I | WAND registration | Not required |
| Class IIa | WAND + conformity evidence | ISO 13485 recommended |
| Class IIb/III | Full pre-market review | ISO 13485 certificate required |
| AIMD | Full pre-market review | ISO 13485 certificate required |

## Recognition

Medsafe accepts CE marking (EU MDR) or MDSAP certificates as evidence of QMS compliance.
NZ participates in the IMDRF to align with international regulatory frameworks.
""",
    ("墨西哥 (Mexico)", "DOF-NOM241-2025"): """\
# NOM-241-SSA1-2021 — Buenas Prácticas de Fabricación para Dispositivos Médicos
# Last reviewed: 2026-05-14

**Citation**: NOM-241-SSA1-2021
**Authority**: COFEPRIS (Comisión Federal para la Protección contra Riesgos Sanitarios)
**Published**: Diario Oficial de la Federación, December 20, 2021
**Effective**: 180 days after publication (June 2022)

## Structure

| Chapter | Topic |
|---|---|
| 1–3 | Scope, References, Definitions |
| 4 | Quality Management System |
| 5 | Management Responsibility |
| 6 | Resource Management |
| 7 | Product Realization |
| 8 | Measurement, Analysis, Improvement |
| Appendix | Medical device-specific supplements |

## Relationship to ISO 13485

NOM-241-SSA1-2021 incorporates ISO 13485:2016 requirements and supplements them with
Mexican-specific requirements for the COFEPRIS registration process.

## Registration Requirements

- Manufacturers must demonstrate NOM-241 compliance as part of Registro Sanitario application
- ISO 13485:2016 certificate from an IAF-accredited body is accepted as evidence
- Foreign manufacturers: certificate must cover the specific manufacturing site

## Scope

Applies to manufacturers of Class I–IV medical devices (dispositivos médicos)
and in vitro diagnostic devices (IVD) in Mexico.
""",
    ("南非 (South Africa)", "SAHPRA-ISO13485"): """\
# SAHPRA — ISO 13485 Certificate Requirement for Medical Device Establishment Licence
# Last reviewed: 2026-05-14

**Authority**: South African Health Products Regulatory Authority (SAHPRA)
**Legal Basis**: Medicines and Related Substances Act 101 of 1965, Regulations 5 & 6
**Effective Date**: June 1, 2025 (SAHPRA Communication, April 2025)

## QMS Requirements

From June 1, 2025, SAHPRA requires an ISO 13485:2016 certificate as a prerequisite for:
- New Medical Device Establishment Licence applications
- Renewal of existing establishment licences

### Accepted Certificates

- ISO 13485:2016 certificate from an accredited certification body (SANAS or IAF MLA member)
- MDSAP audit certificate (from MDSAP-recognised auditing organisation)

## Scope

| Establishment Type | Requirement |
|---|---|
| Local Manufacturer | ISO 13485:2016 certificate for manufacturing site |
| Importer | ISO 13485:2016 certificate from manufacturer |
| Distributor | Letter of Authorisation + manufacturer's ISO 13485 |

## Device Registration

SAHPRA device registration requires:
1. Establishment Licence (with ISO 13485 certificate from June 2025)
2. Device registration application under Section 22C

## IMDRF Alignment

South Africa is actively aligning its medical device regulatory framework with IMDRF guidelines,
including adoption of IMDRF's Technical Document on QMS requirements.
""",
    ("土耳其 (Turkey)", "TITCK-MD-Legislation"): """\
# Turkey Medical Device QMS — EU MDR 2017/745 Aligned Regulation (2021)
# Last reviewed: 2026-05-14

**Authority**: Türkiye İlaç ve Tıbbi Cihaz Kurumu (TITCK)
**Legal Basis**: Tıbbi Cihaz Yönetmeliği (Medical Device Regulation), Official Gazette No. 31499,
  June 2, 2021 | **Structure**: Aligned with EU MDR 2017/745

## Article 10 — General Obligations of Manufacturers (QMS Core Requirement)

10.1 Manufacturers shall establish, document, implement, maintain, keep up to date, and
continually improve a quality management system that ensures compliance with this Regulation
in the most effective manner proportionate to the risk class and type of device.

10.2 The QMS shall cover all parts of and requirements on the manufacturer's organisation
that deal with the quality of processes, procedures, and devices, governing the structure,
responsibilities, procedures, processes, and management resources.

10.3 The QMS shall address the following minimum elements:
  (a) Regulatory compliance strategy including conformity assessment procedures and
      modification management procedures (§4.1, §4.2 ISO 13485 equivalent);
  (b) Identification of applicable general safety and performance requirements (GSPR,
      Annex I — equivalent to ISO 13485 §7.1 risk management);
  (c) Responsibility of management (§5.1, §5.5.1 equivalent);
  (d) Resource management, including supplier and sub-contractor selection and control
      (§6.1, §7.4 equivalent);
  (e) Risk management as set out in Section 3 of Annex I (ISO 14971 by reference);
  (f) Clinical evaluation in accordance with Annex XIV, including PMCF (§8.2.1 equivalent);
  (g) Product realisation, including planning, design, development, production, and
      service provision (§7.1–§7.5 equivalent);
  (h) Verification of UDI assignments (§7.5.3 equivalent);
  (i) Setup, implementation and maintenance of a post-market surveillance system (§8.2.1);
  (j) Communication with competent authorities, notified bodies, economic operators,
      customers, and other stakeholders (§5.5.3, §7.2.3 equivalent);
  (k) Reporting of serious incidents and field safety corrective actions (§8.2.2, §8.2.3);
  (l) Management of CAPA and verification of effectiveness (§8.5.2, §8.5.3 equivalent);
  (m) Monitoring, measurement of output, data analysis, and product improvement (§8.2.5,
      §8.2.6, §8.4, §8.5.1 equivalent).

## Annex IX — QMS Conformity Assessment (ISO 13485 equivalent assessment route)

Annex IX §2.1: Manufacturers of Class IIa, IIb, and III devices shall be subject to a
conformity assessment procedure based on the quality management system described in Annex IX.

Annex IX §2.2 — QMS Assessment Scope: The notified body (or TITCK auditor) shall audit
the manufacturer's QMS covering:
  (a) Quality management system documentation;
  (b) Policies, objectives, and commitment of top management (§5.1–§5.3);
  (c) Document and record control (§4.2.3, §4.2.4);
  (d) Human resources, training, and competence management (§6.2);
  (e) Design and development management procedures (§7.3);
  (f) Purchasing controls and incoming inspection (§7.4);
  (g) Production and process controls (§7.5.1, §7.5.6);
  (h) Traceability including UDI (§7.5.3);
  (i) Post-market surveillance system and vigilance (§8.2.1, §8.2.2, §8.2.3);
  (j) CAPA procedures and effectiveness verification (§8.5.2, §8.5.3);
  (k) Internal audit programme (§8.2.4);
  (l) Management review process and records (§5.6).

## Conformity Assessment Routes by Device Class

| Device Class | Route | QMS Requirement |
|---|---|---|
| Class I | Self-declaration + Declaration of Conformity | No QMS audit; QMS documentation retained |
| Class I (measuring, sterile, reusable surgical) | Annex IX or XI + Notified Body | QMS audit required |
| Class IIa | Annex IX (Route A) or Annex XI Part A | ISO 13485-equivalent QMS audit |
| Class IIb | Annex IX or XI Part A + type examination | Full QMS audit + technical documentation |
| Class III | Annex IX (full assessment) | Full QMS audit + design dossier assessment |
| AIMD | Annex IX (full assessment) | Full QMS audit + design dossier assessment |

## Turkey-Specific QMS Requirements Beyond ISO 13485 Base

- **CE Marking Acceptance (Article 98)**: CE marking under EU MDR 2017/745 from EU
  Notified Bodies is accepted for Turkish Ürün Takip Sistemi (UTS) registration without
  a separate Turkish QMS audit, provided the CE certificate remains valid.
- **UTS Registration**: All medical devices sold in Turkey must be registered in the UTS
  (Product Tracking System). QMS certificate must be submitted.
- **Adverse Event Reporting**: Serious incidents to TITCK within 15 calendar days.
  Trend reports for non-serious events quarterly.
- **FSCA (Field Safety Corrective Actions)**: TITCK notification required before
  implementing FSCA; Field Safety Notice (FSN) in Turkish required.
- **Periodic Safety Update Report (PSUR)**: Required for Class IIa/IIb/III devices;
  annually for Class III, biennially for IIb, every 3 years for IIa.
- **Unique Device Identification (UDI)**: Turkey UDI system aligned with EU MDR;
  phased implementation from 2024 (Class III first).
""",
    ("印尼 (Indonesia)", "Kemkes-CPAKB"): """\
# Indonesia Medical Device QMS — CPAKB (Permenkes No. 20 Tahun 2017)
# Last reviewed: 2026-05-14

**Authority**: Ministry of Health (Kemenkes), Directorate General of Pharmaceutical and Medical Devices (Farmalkes)
**Legal Basis**: Permenkes No. 20 Tahun 2017 — Cara Pembuatan Alat Kesehatan yang Baik (CPAKB)
**Issued**: 8 March 2017 | **Effective**: 18 April 2017 | BN.2017/No.590

## QMS Requirements

Indonesia requires medical device manufacturers to demonstrate GMP / ISO 13485 compliance
as part of the product registration (Nomor Izin Edar — NIE) application.

| Device Class | Risk Level | QMS Requirement |
|---|---|---|
| Kelas A | Low | Basic quality documentation |
| Kelas B | Low-Moderate | ISO 13485 certificate or GMP inspection |
| Kelas C | Moderate-High | ISO 13485 certificate mandatory |
| Kelas D | High | ISO 13485 certificate mandatory |

## ISO 13485 Acceptance

- Domestic manufacturers: ISO 13485 certificate from KAN-accredited (National Accreditation Body) CB
- Foreign manufacturers: ISO 13485 from IAF MLA member + Certificate of Free Sale (CFS) from home country authority

## Registration Process

1. Manufacturer obtains ISO 13485 certification
2. Apply for NIE through Regalkes online portal
3. Submit technical documentation + ISO 13485 certificate
4. Kemenkes review and NIE issuance (12–24 months for Class C/D)
""",
    ("馬來西亞 (Malaysia)", "MDA-Legislation"): """\
# Malaysia Medical Device Act 737 (2012) — QMS Requirements
# Last reviewed: 2026-05-14

> **⚠️ 無獨立 QMS 法規 / 独立QMS法令なし / NO INDEPENDENT QMS LAW** — Malaysia has no standalone QMS regulation equivalent to EU MDR Annex IX or FDA 21 CFR Part 820. QMS requirement exists by reference to ISO 13485:2016 under Medical Device Regulations 2012 (P.U.(A) 210/2012), Part III conformity assessment provisions.

**Citation**: Medical Device Act 2012 (Act 737) & Medical Device Regulations 2012
**Authority**: Medical Device Authority (MDA) Malaysia
**Effective**: July 1, 2013

## QMS Requirements

Part III of the Medical Device Regulations 2012 requires conformity assessment,
including QMS certification, for Class B, C, and D devices.

| Device Class | Conformity Route | QMS Requirement |
|---|---|---|
| Class A | Self-declaration | No third-party QMS audit |
| Class B | Conformity assessment | ISO 13485 or equivalent |
| Class C | Full conformity assessment | ISO 13485 certificate mandatory |
| Class D | Full conformity assessment | ISO 13485 certificate mandatory |

## Accepted QMS Certifications

MDA accepts the following as equivalent evidence of QMS compliance:
- ISO 13485:2016 certificate (from DAkkS, UKAS, SAC, or IAF MLA member)
- MDSAP audit certificate
- FDA QSR compliance (for US-based manufacturers)
- MHLW Ordinance 169 compliance (for Japan-based manufacturers)

## Registration Timeline

Malaysia requires product registration (product listing) through the MDA online portal.
QMS certificate must be submitted with initial registration and renewed every 5 years.
""",
    ("以色列 (Israel)", "MOH-MD-Division"): """\
# Israel Medical Device QMS — Medical Equipment Law 5772-2012
# Last reviewed: 2026-05-14

> **⚠️ 無獨立 QMS 法規 / 独立QMS法令なし / NO INDEPENDENT QMS LAW** — Israel has no standalone QMS regulation equivalent to EU MDR Annex IX or FDA 21 CFR Part 820. QMS requirement (ISO 13485:2016) is embedded within the import licence conditions under Medical Equipment Law 5772-2012 and Medical Equipment Regulations 5773-2013; CE marking under EU MDR is accepted as primary equivalent evidence.

**Authority**: Israel Ministry of Health, Medical Technology, Health Informatics & Research Division (AMAR)
**Legal Basis**: Medical Equipment Law 5772-2012; Medical Devices Regulations 5773-2013

## QMS Requirements

Israel's medical device regulatory framework requires ISO 13485:2016 compliance for
medical device manufacturers as a condition of import licence and market authorisation.

| Device Class (Risk) | Regulatory Route | QMS Requirement |
|---|---|---|
| Class I | Technical file + declaration | ISO 13485 not mandatory |
| Class IIa/IIb | Conformity assessment | ISO 13485 certificate required |
| Class III | Full assessment | ISO 13485 certificate + design review |

## CE/FDA Recognition

Israel Ministry of Health accepts:
- CE marking under EU MDR 2017/745 as primary evidence for most device classes
- FDA 510(k) clearance or PMA approval as equivalent evidence
- ISO 13485:2016 certificate from IAF-accredited CB

## Registration Process

Manufacturers submit registration applications to AMAR through the MOH Medical Device Registration System.
Foreign manufacturers must appoint a local Authorised Representative (Israeli agent).
""",
    ("菲律賓 (Philippines)", "FDA-PH-RA9711"): """\
# Philippines FDA — Medical Device QMS Requirements (RA 9711 / FDA Act 2009)
# Last reviewed: 2026-05-14

> **⚠️ 無獨立 QMS 法規 / 独立QMS法令なし / NO INDEPENDENT QMS LAW** — Philippines has no standalone QMS regulation equivalent to EU MDR Annex IX or FDA 21 CFR Part 820. QMS requirement (ISO 13485:2016) exists as a registration condition under FDA Circular No. 2020-007 issued pursuant to Republic Act 9711 (FDA Act of 2009); RA 9711 itself is a market access law, not a QMS technical standard.

**Authority**: Food and Drug Administration Philippines (FDA-PH)
**Center**: Center for Device Regulation, Radiation Health and Research (CDRRHR)
**Legal Basis**: Republic Act 9711 (FDA Act of 2009); FDA Circular No. 2020-007

## QMS Requirements

FDA-PH requires medical device manufacturers to demonstrate GMP compliance (ISO 13485:2016)
for Certificate of Product Registration (CPR) applications.

| Device Class | Risk | QMS Requirement |
|---|---|---|
| Class A | Exempt/Low | Basic quality documentation |
| Class B | Low-Moderate | ISO 13485 certificate from IAF-accredited CB |
| Class C | Moderate-High | ISO 13485 mandatory |
| Class D | High | ISO 13485 mandatory + premarket evaluation |

## ISO 13485 Acceptance

FDA-PH accepts ISO 13485:2016 certificates from certification bodies accredited by
Philippine Accreditation Bureau (PAB) or IAF MLA member bodies.

Foreign manufacturers must submit:
- ISO 13485 certificate
- Certificate of Free Sale (CFS) / Certificate of Exportation (CE)
- Authorised Philippine Distributor appointment

## ASEAN Alignment

Philippines participates in ASEAN Harmonization of Medical Device Regulations and
the ASEAN MDPQ (Medical Device Product Quality) framework.
""",
    ("越南 (Vietnam)", "DMEC-MOH"): """\
# Vietnam Medical Device QMS — Decree 98/2021/ND-CP & Circular 10/2023/TT-BYT
# Last reviewed: 2026-05-14

**Authority**: Ministry of Health (MOH), Department of Medical Equipment and Construction (DMEC)
**Primary Regulation**: Decree No. 98/2021/ND-CP (November 8, 2021)
**Technical Circular**: Circular 10/2023/TT-BYT (May 22, 2023)

## QMS Requirements

Decree 98/2021 requires ISO 13485:2016 QMS certification for all Class B, C, and D medical device
manufacturers as a mandatory condition for product registration (number/declaration).

| Device Class | Vietnamese Classification | QMS Requirement |
|---|---|---|
| Class A | Loại A | No ISO 13485 required |
| Class B | Loại B | ISO 13485 certificate mandatory |
| Class C | Loại C | ISO 13485 certificate mandatory |
| Class D | Loại D | ISO 13485 certificate mandatory |

## ISO 13485 Certification

- Domestic manufacturers: certificate from VILAS (Vietnam Laboratory Accreditation Scheme) or IAF MLA member
- Foreign manufacturers: certificate from IAF MLA member in country of manufacture

## Registration Process (Circular 10/2023)

1. Manufacturer obtains ISO 13485:2016 certificate
2. Vietnam importer/distributor acts as responsible party
3. Submit Declaration of Conformity or Registration Application to MOH/DMEC
4. DMEC issues Registration Number (Số lưu hành)
""",
    ("哥倫比亞 (Colombia)", "INVIMA-Decreto4725"): """\
# Colombia Medical Device QMS — Decreto 4725/2005 (INVIMA)
# Last reviewed: 2026-05-14

**Authority**: Instituto Nacional de Vigilancia de Medicamentos y Alimentos (INVIMA)
**Legal Basis**: Decreto 4725 de 2005 (Medical Devices and Diagnostic Equipment)
**Administering Ministry**: Ministry of Health and Social Protection

## QMS Requirements

Decreto 4725/2005 requires medical device manufacturers to demonstrate GMP / ISO 13485
compliance as a condition of INVIMA registration (Registro Sanitario).

| Device Class | INVIMA Class | QMS Requirement |
|---|---|---|
| Class I | Clase I | GMP declaration |
| Class IIa | Clase IIa | ISO 13485 evidence |
| Class IIb | Clase IIb | ISO 13485 certificate |
| Class III | Clase III | ISO 13485 certificate mandatory |

## ISO 13485 Acceptance

INVIMA accepts ISO 13485:2016 certificates from:
- ONAC (Organismo Nacional de Acreditación de Colombia) accredited CBs
- IAF MLA member accreditation bodies internationally

## Registration Process

1. Manufacturer submits technical dossier + ISO 13485 certificate to INVIMA
2. INVIMA issues Registro Sanitario (10-year validity for Class III; 5 years for others)
3. Post-market surveillance obligations apply after registration

## IMDRF/PAHO Alignment

Colombia is working towards alignment with IMDRF guidelines through the PAHO/PANDRH
(Pan American Network on Drug Regulatory Harmonization) medical device working group.
""",
    ("俄羅斯 (Russia)", "RZN-MD"): """\
# Russia Medical Device QMS — Government Decree No. 1684 (2024)
# Last reviewed: 2026-05-14

**Authority**: Federal Service for Surveillance in Healthcare (Roszdravnadzor)
**Primary Regulation**: Government Decree No. 1684 (effective March 1, 2025; replaces No. 1416)
**Additional**: Ministry of Health Order No. 972 (2024)

## QMS Requirements

From January 1, 2024 (Class IIa-sterile and above), Roszdravnadzor requires ISO 13485:2016
inspection for medical device state registration (государственная регистрация).

| Device Class (Russian) | Risk | ISO 13485 Inspection |
|---|---|---|
| Class 1 | Low | Not required |
| Class 2a | Moderate (non-sterile) | Not required |
| Class 2a (sterile) | Moderate | Required from Jan 2024 |
| Class 2b | Moderate-High | Required |
| Class 3 | High | Required |

## Regulatory Process

1. Roszdravnadzor GMP inspection of manufacturer's site OR submission of ISO 13485 certificate
2. Technical Specification (ТУ) filing
3. State Registration Certificate issuance

## Eurasian Economic Union (EAEU)

Russia participates in EAEU regulatory harmonization; medical devices may also be
registered under the Common Medical Device Registration Procedure for circulation across
Russia, Belarus, Kazakhstan, Armenia, and Kyrgyzstan.
""",
    ("埃及 (Egypt)", "EDA-MD-Guide"): """\
# Egypt Medical Device QMS — EDA Registration Requirements
# Last reviewed: 2026-05-14

**Authority**: Egyptian Drug Authority (EDA), formerly CAPA
**Legal Basis**: Ministerial Decree No. 425/2015; EDA Law 151/2019
**QMS Guidance**: EDA Regulatory Guideline for Registering Medical Devices (2021)

## QMS Requirements

EDA requires medical device manufacturers to submit ISO 13485:2016 certificate as a
mandatory document for product registration.

| Device Class | Risk | QMS Requirement |
|---|---|---|
| Class A | Low | Basic quality declaration |
| Class B | Moderate | ISO 13485 certificate required |
| Class C | High | ISO 13485 certificate mandatory |
| Class D | Critical | ISO 13485 certificate mandatory + clinical data |

## ISO 13485 Acceptance

EDA accepts ISO 13485:2016 certificates from:
- Egyptian Accreditation Council (EGAC) accredited CBs
- IAF MLA member international CBs
- CE marking (under EU MDR/MDD) as alternative evidence for most classes

## Registration Requirements

Foreign manufacturers must:
1. Submit ISO 13485:2016 certificate
2. Submit Certificate of Free Sale from home country authority
3. Appoint EDA-registered local agent
4. Provide technical documentation (product dossier)

## Timeline

EDA registration validity: 3 years (renewable). Manufacturers must notify EDA of
significant changes to QMS or device design.
""",
    ("智利 (Chile)", "ISP-Regulations"): """\
# Chile Medical Device QMS — ISP Chile / ANAMED Regulations
# Last reviewed: 2026-05-14

> **⚠️ 無獨立 QMS 法規 / 独立QMS法令なし / NO INDEPENDENT QMS LAW** — Chile has no standalone QMS regulation equivalent to EU MDR Annex IX or FDA 21 CFR Part 820. GMP requirements are embedded in Decreto Supremo No. 825/1998 (Reglamento de Dispositivos Médicos) with reference to US 21 CFR Part 820 standards; ISO 13485:2016 certificate is accepted as equivalent evidence for Registro Sanitario applications.

**Authority**: Instituto de Salud Pública de Chile (ISP), División de Dispositivos Médicos (ANAMED)
**Legal Basis**: Decreto Supremo No. 825/1998 (Reglamento de Dispositivos Médicos)
**GMP Guide**: ISP Resolution N°8209/1999 (GMP Guide for Medical Devices)

## QMS Requirements

Chile requires medical device manufacturers to demonstrate GMP compliance as a condition of
ISP registration (Registro Sanitario). ISO 9001 or ISO 13485 compliance is referenced.

| Device Class | Risk | QMS Requirement |
|---|---|---|
| Class I | Exempt | No registration required |
| Class II | Low-Moderate | ISO 13485 or GMP evidence |
| Class III | High | ISO 13485 certificate mandatory |

## ISO 13485 / CE / FDA Acceptance

ISP Chile accepts:
- ISO 13485:2016 certificate from IAF MLA member CB
- CE marking under EU MDR 2017/745
- FDA 510(k) or PMA clearance
- TGA (Australia) approval

## Registration Process

1. Manufacturer prepares technical dossier
2. Chilean importer submits to ISP/ANAMED
3. ISP issues Registro Sanitario
4. Periodic renewal every 5 years

## PANDRH Alignment

Chile participates in PAHO/PANDRH harmonization initiatives and is working towards
adoption of IMDRF-aligned regulatory framework for medical devices.
""",
    ("阿聯酋 (UAE)", "MOHAP-FDL38-2024"): """\
# UAE Medical Device QMS — Federal Decree-Law No. 38 of 2024
# Last reviewed: 2026-05-14

**Authority**: Ministry of Health and Prevention (MOHAP) & Emirates Health Authority (EHA)
**Legal Basis**: Federal Decree-Law No. 38 of 2024 (Medical Products, Pharmacy Practice
  and Pharmaceutical Establishments) | **In Force**: January 2025

## Article 47 — Quality Management System Requirements

47.1 Medical device manufacturers operating in or exporting to the UAE shall establish,
implement, and maintain a quality management system that complies with the requirements of
ISO 13485:2016 as referenced by MOHAP Technical Guidance.

47.2 The QMS shall address:
  (a) Quality policy and quality objectives established by top management (§5.3, §5.4.1);
  (b) Quality manual documenting QMS scope and process interactions (§4.2.2);
  (c) Document control: approval, review, update, identification, and withdrawal of obsolete
      documents (§4.2.3);
  (d) Record control: identification, storage, protection, retrieval, retention, and
      disposal of quality records. Retention period minimum 5 years or device lifetime (§4.2.4);
  (e) Design and development controls for all Class II, III devices (§7.3.1–§7.3.10);
  (f) Supplier management: approved supplier list, evaluation criteria, purchasing controls,
      and incoming inspection/verification (§7.4.1–§7.4.3);
  (g) Production controls: documented procedures, work instructions, process validation
      for special processes (§7.5.1–§7.5.6);
  (h) Calibration programme for monitoring and measuring equipment (§7.6);
  (i) Post-market surveillance and vigilance system (§8.2.1–§8.2.3);
  (j) CAPA with root cause analysis and effectiveness verification (§8.5.2–§8.5.3).

## Article 48 — Management Responsibility

48.1 Top management shall demonstrate commitment to QMS establishment and continual
improvement by establishing quality policy, ensuring objectives are set, conducting
management reviews, and ensuring resource availability (§5.1).

48.2 Management representative shall be appointed with documented responsibility for
QMS maintenance and reporting performance to top management (§5.5.2).

48.3 Management review shall be conducted at least annually with documented records.
Inputs shall include audit results, customer feedback, process performance, nonconformances,
CAPA status, regulatory changes, and recommendations for improvement (§5.6.2).

## Article 49 — Resource Management

49.1 Personnel competence requirements shall be defined, and training records maintained.
Personnel affecting product quality shall demonstrate required competence (§6.2).

49.2 Infrastructure including buildings, equipment, and support services shall be maintained
with documented maintenance schedules (§6.3).

49.3 Work environment requirements including temperature, humidity, and contamination control
shall be defined and monitored for relevant processes (§6.4).

## Article 50 — Vigilance and Adverse Event Reporting

50.1 Serious adverse events (events causing or likely to cause serious injury or death) shall
be reported to MOHAP within 15 calendar days of becoming aware.

50.2 Field Safety Corrective Actions (FSCA) require prior notification to MOHAP. Field Safety
Notices (FSN) must be submitted in Arabic and English.

50.3 Periodic Safety Update Reports (PSUR) required for high-risk devices annually.

## MOHAP/EMAAR Registration Requirements

| Entity Type | QMS Requirement |
|---|---|
| UAE Manufacturer | ISO 13485 certificate + MOHAP on-site GMP inspection |
| Foreign Manufacturer | ISO 13485 from IAF MLA member CB at manufacturing site |
| Importer | Manufacturer ISO 13485 + MOHAP importer establishment licence |
| Distributor | Manufacturer ISO 13485 + MOHAP distributor establishment licence |

## Device Classification and QMS Requirements

| Class | Risk | QMS/Registration |
|---|---|---|
| Class A | Low | Basic product registration; no QMS audit |
| Class B | Low-Moderate | ISO 13485 certificate required |
| Class C | Moderate-High | ISO 13485 certificate + technical dossier |
| Class D | High | ISO 13485 + clinical evidence + MOHAP pre-market review |

## UAE-Specific Additional Requirements

- **Authorized Local Distributor**: Foreign manufacturers must appoint a UAE-licensed
  establishment as authorized local distributor responsible for market surveillance.
- **Arabic Labelling**: All labelling and IFU must be available in Arabic for UAE market.
- **UDI (Unique Device Identification)**: MOHAP UDI system aligned with IMDRF standards;
  Class C/D devices first, phased rollout through 2027.
- **Emirates Health Authority (DHA/DOH)**: Dubai and Abu Dhabi may require additional
  emirate-level establishment registration that supplements federal MOHAP licence.
- **Halal Considerations**: Devices incorporating materials of animal origin require
  ESMA (Emirates Authority for Standardization) halal certification.
""",
    ("阿根廷 (Argentina)", "ANMAT-MD"): """\
# Argentina Medical Device QMS — ANMAT Disposición 64/2025 (replaces 2318/2002)
# Last reviewed: 2026-05-14

**Authority**: Administración Nacional de Medicamentos, Alimentos y Tecnología Médica (ANMAT)
**Current Regulation**: Disposición ANMAT 64/2025 (in force Jan 2025, incorporates MERCOSUR GMC 25/21)
**Previous**: Disposición ANMAT 2318/2002 (superseded)
**GMP Reference**: Disposición ANMAT 2319/2002 — BPF (Buenas Prácticas de Fabricación) for manufacturers

## QMS Requirements

ANMAT requires GMP/BPF (Buenas Prácticas de Fabricación) compliance for medical device registration.
ISO 13485:2016 is the accepted international standard.

| Device Class (ANMAT) | Risk | QMS Requirement |
|---|---|---|
| Class I | Exempt/Low | Basic declaration |
| Class II | Moderate | ISO 13485 evidence or ANMAT GMP inspection |
| Class III | High | ISO 13485 certificate mandatory |
| Class IV | Critical | ISO 13485 + clinical evidence |

## ISO 13485 Acceptance

ANMAT accepts:
- ISO 13485:2016 certificate from OAA (Organismo Argentino de Acreditación) or IAF MLA member
- CE marking (EU MDR) for most device classes
- FDA clearance/approval

## Registration Process

1. Foreign manufacturer appoints Argentine technical director (director técnico)
2. Technical dossier submission to ANMAT
3. ANMAT issues Certificado de Autorización de Uso (CUA)

## MERCOSUR Harmonization

Argentina participates in MERCOSUR Resolution GMC 40/00 on medical device harmonization,
working towards mutual recognition of registrations across Brazil, Argentina, Uruguay, and Paraguay.
""",
}


# ============================================================
# Constants
# ============================================================

_DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9,zh-TW;q=0.8",
}

_REQUEST_TIMEOUT = 30  # seconds
_DOWNLOAD_MAX_BYTES = 200 * 1024 * 1024     # 200 MB hard cap — no regulatory PDF exceeds this
_OFFICE_ZIP_BOMB_LIMIT = 500 * 1024 * 1024  # 500 MB uncompressed — blocks ZIP-bomb docx/xlsx
_JINA_READER_BASE = "https://r.jina.ai/"
_JINA_TIMEOUT = 60  # Jina Reader may be slow
_JINA_DELAY = 3.0  # seconds between Jina requests (20 req/min limit)
_MAX_CONTENT_SIZE = 150_000  # 150KB max markdown content (B-3: was 50KB)
_DOMAIN_CONCURRENCY = 2  # max concurrent requests per domain
_ETAG_CACHE_PATH = Path("data/etag_cache.json")
_INDEX_DIFF_CACHE_PATH = Path("data/index_page_diff_cache.json")

# ── G2: doc_type mapping for all REGION_SITES entries ──────────────────────
# primary      = main QMS regulation/law for that country
# qms_guidance = semi-official guidance supplementing ISO 13485 QMS requirements
# portal       = homepage/portal with no direct regulatory content (excluded from LLM)
_AGENCY_DOC_TYPE: dict[str, str] = {
    # Taiwan — bulk API replaces individual Jina entries; EN kept as supplementary
    "TFDA-BulkAPI": "primary",
    "TFDA-QMS": "primary", "TFDA-QMS-EN": "qms_guidance", "TFDA-QMS-Inspection": "primary",
    # USA
    "FDA-QMSR": "primary", "eCFR-820": "primary",
    # EU
    "EUR-Lex-MDR-CELLAR": "primary", "EUR-Lex-MDR-UK": "primary", "EUR-Lex-MDR-OJ": "primary",
    "MDCG": "qms_guidance", "MDCG-2019-11": "qms_guidance", "MDCG-2021-25": "qms_guidance",
    "MDCG-2020-7": "qms_guidance", "MDCG-2022-14": "qms_guidance", "MDCG-2021-27": "qms_guidance",
    "MDCG-2022-21": "qms_guidance", "MDCG-2019-6": "qms_guidance", "MDCG-2018-1": "qms_guidance",
    "MDCG-2020-14": "qms_guidance", "MDCG-2023-1": "qms_guidance",
    # UK
    "UK-MDR-2002": "primary", "MHRA-Guidance": "qms_guidance",
    # Japan
    "eGov-QMS-169": "primary", "PMDA-QMS-EN": "primary", "PMDA-QMS": "primary",
    # China
    "NMPA-GMP-CN": "primary", "NMPA-GMP-Announcement": "qms_guidance", "NMPA-Regulations": "portal",
    # Korea
    "MFDS-MD-Regulations": "primary", "MFDS-KGMP": "primary", "MFDS-KGMP-PDF": "primary",
    # Canada
    "CMDR-SOR98-282": "primary", "CMDR-Full": "primary",
    "FDA-MDSAP-Audit": "qms_guidance", "FDA-MDSAP-Assessment": "qms_guidance",
    "MDSAP-Companion-ISO13485": "qms_guidance",
    # Australia
    "TGA-Legislation": "primary", "TGA-Legislation-PDF": "primary", "TGA-ARGMD": "qms_guidance",
    # Switzerland
    "Swissmedic-MedDO": "primary", "MedDO-SR812213": "primary",
    # Brazil
    "ANVISA-RDC665": "primary", "ANVISA-RDC665-News": "portal",
    # MDSAP
    "MDSAP-Global-Audit": "qms_guidance", "MDSAP-Global-QMS": "qms_guidance",
    "MDSAP-Global-General": "qms_guidance", "MDSAP-Global-Assessment": "qms_guidance",
    # International
    "ISO": "primary", "ICH": "primary", "IMDRF": "primary",
    # India
    "CDSCO-MDR2017": "primary", "CDSCO-MD": "portal",
    # Singapore
    "SSO-HPR-2010-PDF": "primary", "SSO-HPR-2010": "portal", "HSA-QMS": "qms_guidance",
    # Saudi Arabia
    "SFDA-MDS-REQ10": "primary", "SFDA-ISO13485-Guidance": "qms_guidance",
    # Thailand
    "Thai-FDA-GMP": "primary", "Thai-FDA-GMP-Law": "primary",
    "Thai-FDA-NewLaws": "qms_guidance", "Thai-FDA-GMP-About": "qms_guidance",
    # New Zealand
    "Medsafe-MD-Legislation": "primary", "MoH-NZ-MD-Regulation": "primary",
    "NZ-Medicines-Act-1981": "primary", "Medsafe-GMP": "qms_guidance",
    # Mexico
    "NOM241-EN-PDF": "primary", "DOF-NOM241-2025": "primary",
    # Argentina
    "ANMAT-Disposicion64-25": "primary", "ANMAT-MD": "portal",
    # South Africa
    "SAHPRA-MD": "primary", "SAHPRA-ISO13485": "primary",
    # Turkey
    "Mevzuat-TCY": "primary", "TITCK-MD-Legislation": "primary",
    "TITCK-Mevzuat": "portal", "TITCK-New-Regs": "qms_guidance",
    # Indonesia
    "BPK-Permenkes20": "primary", "Kemkes-CPAKB": "primary",
    # Malaysia
    "MDA-Legislation-List": "primary", "MDA-Act737-PDF": "primary", "MDA-Legislation": "portal",
    # Israel
    "Nevo-MD-Regulations-2021": "primary", "Nevo-MD-Law-2012": "primary",
    "MOH-Laws": "portal", "MOH-MD-Division": "qms_guidance",
    # Philippines
    "FDA-PH-RA9711": "primary", "FDA-PH-MD": "portal",
    # Vietnam
    "LuatVN-Decree98-EN": "primary", "LuatVN-Consolidated-PDF": "primary",
    "LuatVN-Decree04-2025-EN": "primary", "DMEC-MOH": "portal",
    # Colombia
    "INVIMA-Decreto4725": "primary", "INVIMA-Decreto4725-PDF": "primary", "INVIMA-MD": "portal",
    # Russia
    "ZakonRF-1684": "primary", "Consultant-1684": "primary", "RZN-MD": "portal",
    # Egypt
    "EDA-MD-Guide": "primary", "EDA": "portal",
    # Chile
    "ISP-Regulations": "primary",
    # UAE
    "MOHAP-FDL38-2024": "primary", "MOHAP": "portal",
}


def get_site_doc_type(site: dict) -> str:
    """Return doc_type for a site dict. Falls back to 'primary' if unknown."""
    explicit = site.get("doc_type")
    if explicit:
        return explicit
    return _AGENCY_DOC_TYPE.get(site.get("agency", ""), "primary")


# QMS-related keywords for sitemap URL filtering
_QMS_KEYWORDS = [
    "medical-device",
    "medical_device",
    "medicaldevice",
    "guidance",
    "regulation",
    "regulatory",
    "qms",
    "quality-management",
    "quality_management",
    "iso-13485",
    "iso13485",
    "mdr",
    "ivdr",
    "510k",
    "premarket",
    "post-market",
    "device",
    "diagnostic",
    "mdsap",
    "single-audit",
    "single_audit",
    "audit-approach",
    "audit-procedures",
]


# ============================================================
# ETag Cache (HTTP conditional request caching)
# ============================================================


class ETagCache:
    """HTTP ETag / Last-Modified conditional cache stored as JSON."""

    def __init__(self, cache_path: Path = _ETAG_CACHE_PATH):
        self._path = cache_path
        self._data: dict = {}
        self._dirty = False

    async def load(self) -> None:
        """Load cache from disk."""
        if self._path.exists():
            try:
                async with aiofiles.open(self._path, "r", encoding="utf-8") as f:
                    raw = await f.read()
                self._data = json.loads(raw) if raw.strip() else {}
            except Exception as e:
                logger.warning(f"ETag cache load failed: {e}")
                self._data = {}
        self._dirty = False

    async def save(self) -> None:
        """Persist cache to disk."""
        if not self._dirty:
            return
        try:
            self._path.parent.mkdir(parents=True, exist_ok=True)
            content = json.dumps(self._data, ensure_ascii=False, indent=2)
            async with aiofiles.open(self._path, "w", encoding="utf-8") as f:
                await f.write(content)
            self._dirty = False
        except Exception as e:
            logger.warning(f"ETag cache save failed: {e}")

    def get(self, url: str) -> Optional[dict]:
        """Get cached ETag/Last-Modified for a URL."""
        return self._data.get(url)

    def set(
        self,
        url: str,
        etag: Optional[str] = None,
        last_modified: Optional[str] = None,
        content_hash: Optional[str] = None,
    ) -> None:
        """Store ETag/Last-Modified/content-hash for a URL."""
        self._data[url] = {
            "etag": etag,
            "last_modified": last_modified,
            "content_hash": content_hash,
            "cached_at": datetime.now(timezone.utc).isoformat(),
        }
        self._dirty = True


# ============================================================
# Sitemap Scanner (Tier 0 pre-scan)
# ============================================================

_SITEMAP_NS = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}


class SitemapScanner:
    """Parse XML sitemaps to discover recently-updated regulatory URLs."""

    async def scan(
        self,
        client: httpx.AsyncClient,
        sitemap_url: str,
        last_crawl_time: Optional[str] = None,
        keywords: Optional[list] = None,
    ) -> list:
        """Scan a sitemap for recently-updated URLs matching keywords.

        Args:
            client: shared httpx.AsyncClient
            sitemap_url: URL of sitemap.xml or sitemapindex
            last_crawl_time: ISO timestamp; only return URLs updated after this
            keywords: URL path keywords to filter (default: QMS-related)

        Returns:
            List of URLs that are recently updated and match keywords.
            Empty list on any error (never raises).
        """
        kw = keywords or _QMS_KEYWORDS
        try:
            if not _is_safe_url(sitemap_url):
                raise ValueError(f"Blocked unsafe URL: {sitemap_url}")
            resp = await client.get(
                sitemap_url,
                headers={"Accept": "application/xml, text/xml"},
                timeout=httpx.Timeout(15.0),
            )
            if resp.status_code != 200:
                return []

            root = ET.fromstring(resp.text)
            tag = root.tag.split("}")[-1] if "}" in root.tag else root.tag

            if tag == "sitemapindex":
                return await self._scan_index(client, root, last_crawl_time, kw)
            elif tag == "urlset":
                return self._filter_urls(root, last_crawl_time, kw)
            return []

        except Exception as e:
            logger.debug(f"Sitemap scan failed for {sitemap_url}: {e}")
            return []

    async def _scan_index(
        self,
        client: httpx.AsyncClient,
        root: _stdlib_ET.Element,
        last_crawl_time: Optional[str],
        keywords: list,
    ) -> list:
        """Parse sitemapindex and scan child sitemaps in parallel."""
        child_urls = []
        for sitemap_el in root.findall("sm:sitemap", _SITEMAP_NS):
            loc = sitemap_el.find("sm:loc", _SITEMAP_NS)
            lastmod = sitemap_el.find("sm:lastmod", _SITEMAP_NS)
            if loc is None:
                continue
            # Skip child sitemaps not updated since last crawl
            if last_crawl_time and lastmod is not None and lastmod.text:
                if lastmod.text < last_crawl_time:
                    continue
            child_urls.append(loc.text)

        if not child_urls:
            return []

        # Scan up to 5 child sitemaps in parallel
        all_urls = []
        for batch_start in range(0, len(child_urls), 5):
            batch = child_urls[batch_start : batch_start + 5]
            tasks = []
            for curl in batch:
                tasks.append(
                    self._fetch_and_filter(client, curl, last_crawl_time, keywords)
                )
            results = await asyncio.gather(*tasks, return_exceptions=True)
            for r in results:
                if isinstance(r, list):
                    all_urls.extend(r)

        return all_urls[:100]  # cap at 100 URLs

    async def _fetch_and_filter(
        self,
        client: httpx.AsyncClient,
        sitemap_url: str,
        last_crawl_time: Optional[str],
        keywords: list,
    ) -> list:
        """Fetch a single child sitemap and filter URLs."""
        try:
            if not _is_safe_url(sitemap_url):
                raise ValueError(f"Blocked unsafe URL: {sitemap_url}")
            resp = await client.get(
                sitemap_url,
                headers={"Accept": "application/xml, text/xml"},
                timeout=httpx.Timeout(15.0),
            )
            if resp.status_code != 200:
                return []
            root = ET.fromstring(resp.text)
            return self._filter_urls(root, last_crawl_time, keywords)
        except Exception:
            return []

    @staticmethod
    def _filter_urls(
        root: _stdlib_ET.Element,
        last_crawl_time: Optional[str],
        keywords: list,
    ) -> list:
        """Extract URLs from <urlset> matching keywords and lastmod filter."""
        matched = []
        for url_el in root.findall("sm:url", _SITEMAP_NS):
            loc = url_el.find("sm:loc", _SITEMAP_NS)
            if loc is None or not loc.text:
                continue
            url_text = loc.text.lower()

            # Keyword filter
            if not any(kw in url_text for kw in keywords):
                continue

            # Lastmod filter
            if last_crawl_time:
                lastmod = url_el.find("sm:lastmod", _SITEMAP_NS)
                if lastmod is not None and lastmod.text:
                    if lastmod.text < last_crawl_time:
                        continue

            matched.append(loc.text)

        return matched


# ============================================================
# Helpers — Content conversion
# ============================================================


def _bs4_strip_boilerplate(html: str) -> str:
    """Remove nav, footer, header, script, style, aside from HTML.

    This is a lightweight pre-processing step before MarkItDown.
    """
    try:
        soup = BeautifulSoup(html, "lxml")
        for tag in soup.find_all(
            ["script", "style", "nav", "footer", "header", "aside", "noscript"]
        ):
            tag.decompose()

        # Try to isolate main content
        main = (
            soup.find("main")
            or soup.find("article")
            or soup.find("div", {"role": "main"})
            or soup.find("div", {"id": re.compile(r"content|main", re.I)})
            or soup.find("div", {"class": re.compile(r"content|main", re.I)})
        )
        target = main if main else soup.body if soup.body else soup
        return str(target)
    except Exception:
        return html


def _html_to_markdown(html: str, url: str = "") -> str:
    """Convert HTML to Markdown — MarkItDown primary, BS4 fallback."""
    # First try MarkItDown (higher quality)
    if MARKITDOWN_AVAILABLE and _MD_CONVERTER:
        try:
            stripped = _bs4_strip_boilerplate(html)
            result = _MD_CONVERTER.convert_stream(
                io.BytesIO(stripped.encode("utf-8")),
                file_extension=".html",
            )
            md = result.text_content or ""
            if md.strip() and len(md.strip()) > 50:
                return md
        except Exception as e:
            logger.debug(f"MarkItDown failed for {url}: {e}, falling back to BS4")

    # Fallback: manual BS4 extraction
    return _html_to_markdown_bs4(html, url)


def _html_to_markdown_bs4(html: str, url: str = "") -> str:
    """Fallback: Convert HTML to Markdown using BeautifulSoup extraction."""
    try:
        soup = BeautifulSoup(html, "lxml")
        for tag in soup.find_all(
            ["script", "style", "nav", "footer", "header", "aside", "noscript"]
        ):
            tag.decompose()

        main = (
            soup.find("main")
            or soup.find("article")
            or soup.find("div", {"role": "main"})
            or soup.find("div", {"id": re.compile(r"content|main", re.I)})
            or soup.find("div", {"class": re.compile(r"content|main", re.I)})
        )
        target = main if main else soup.body if soup.body else soup

        lines = []
        title = soup.find("title")
        if title and title.string:
            lines.append(f"# {title.string.strip()}")
            lines.append("")

        for element in target.find_all(
            ["h1", "h2", "h3", "h4", "p", "li", "td", "th", "pre", "blockquote"]
        ):
            text = element.get_text(strip=True)
            if not text:
                continue
            tag_name = element.name
            if tag_name == "h1":
                lines.append(f"# {text}")
            elif tag_name == "h2":
                lines.append(f"## {text}")
            elif tag_name == "h3":
                lines.append(f"### {text}")
            elif tag_name == "h4":
                lines.append(f"#### {text}")
            elif tag_name == "li":
                lines.append(f"- {text}")
            elif tag_name in ("td", "th"):
                continue
            elif tag_name == "pre":
                lines.append(f"```\n{text}\n```")
            elif tag_name == "blockquote":
                lines.append(f"> {text}")
            else:
                lines.append(text)
            lines.append("")

        # Extract tables
        for table in target.find_all("table"):
            rows = table.find_all("tr")
            if not rows:
                continue
            table_lines = []
            for i, row in enumerate(rows[:50]):
                cells = row.find_all(["th", "td"])
                cell_texts = [c.get_text(strip=True)[:100] for c in cells]
                if cell_texts:
                    table_lines.append("| " + " | ".join(cell_texts) + " |")
                    if i == 0:
                        table_lines.append(
                            "| " + " | ".join(["---"] * len(cell_texts)) + " |"
                        )
            if table_lines:
                lines.extend(table_lines)
                lines.append("")

        markdown = "\n".join(lines)
        return markdown if markdown.strip() else f"(No extractable content from {url})"

    except Exception as e:
        return f"(HTML parsing failed: {e})"


def _parse_rss_feed(xml_text: str, agency: str) -> tuple[str, str, list[dict]]:
    """Parse RSS 2.0 / Atom 1.0 feed and return (markdown, channel_title, items).

    G6: RSS monitoring support. Extracts up to 50 recent items and renders them
    as a Markdown table with title, date, and link for downstream diff detection.

    Returns:
        markdown:      Structured Markdown summary of feed items
        channel_title: Feed channel/title string
        items:         List of dicts with keys: title, link, pubDate, description
    """
    try:
        root = ET.fromstring(xml_text.strip())
    except Exception:
        # Try stripping XML declaration then re-parse
        cleaned = re.sub(r"<\?xml[^?]*\?>", "", xml_text.strip(), count=1)
        root = _stdlib_ET.fromstring(cleaned)

    ns_atom = "http://www.w3.org/2005/Atom"
    is_atom = root.tag == f"{{{ns_atom}}}feed" or root.tag == "feed"

    items: list[dict] = []
    channel_title = agency

    if is_atom:
        # Atom 1.0
        title_el = root.find(f"{{{ns_atom}}}title") or root.find("title")
        if title_el is not None and title_el.text:
            channel_title = title_el.text.strip()
        entry_tag = f"{{{ns_atom}}}entry" if f"{{{ns_atom}}}feed" in root.tag else "entry"
        for entry in root.findall(entry_tag):
            def _atom_text(tag: str) -> str:
                el = entry.find(f"{{{ns_atom}}}{tag}") or entry.find(tag)
                return (el.text or "").strip() if el is not None else ""
            link_el = (
                entry.find(f"{{{ns_atom}}}link[@rel='alternate']")
                or entry.find(f"{{{ns_atom}}}link")
                or entry.find("link")
            )
            link = link_el.get("href", "") if link_el is not None else ""
            items.append({
                "title":       _atom_text("title"),
                "link":        link,
                "pubDate":     _atom_text("updated") or _atom_text("published"),
                "description": _atom_text("summary") or _atom_text("content"),
            })
    else:
        # RSS 2.0
        channel = root.find("channel") or root
        title_el = channel.find("title")
        if title_el is not None and title_el.text:
            channel_title = title_el.text.strip()
        for item in channel.findall("item"):
            def _rss_text(tag: str) -> str:
                el = item.find(tag)
                return (el.text or "").strip() if el is not None else ""
            items.append({
                "title":       _rss_text("title"),
                "link":        _rss_text("link"),
                "pubDate":     _rss_text("pubDate"),
                "description": _rss_text("description"),
            })

    items = items[:50]  # keep up to 50 most-recent entries

    lines = [
        f"# {channel_title}",
        f"**Source**: {agency} RSS/Atom feed  ",
        f"**Items**: {len(items)}  ",
        "",
        "| Date | Title | Link |",
        "|------|-------|------|",
    ]
    for it in items:
        date = it["pubDate"][:10] if it["pubDate"] else "—"
        title = it["title"].replace("|", "\\|")[:120]
        link = it["link"]
        lines.append(f"| {date} | {title} | {link} |")

    if items:
        lines += ["", "## Summaries", ""]
        for it in items[:10]:
            if it["description"]:
                desc = re.sub(r"<[^>]+>", "", it["description"])[:300].strip()
                lines.append(f"### {it['title'][:80]}")
                lines.append(f"{desc}\n")

    return "\n".join(lines), channel_title, items


def _json_to_markdown(data: dict, url: str = "") -> str:
    """Convert API JSON response to readable Markdown."""
    try:
        # eCFR API
        if "meta" in data and "content" in data:
            content = data.get("content", "")
            if isinstance(content, str):
                return _html_to_markdown(content, url)

        # Federal Register API
        if "results" in data and isinstance(data["results"], list):
            lines = ["# Federal Register — FDA Rules\n"]
            for item in data["results"][:20]:
                title = item.get("title", "N/A")
                doc_number = item.get("document_number", "")
                pub_date = item.get("publication_date", "")
                abstract = item.get("abstract", "")[:300]
                html_url = item.get("html_url", "")
                lines.append(f"## {title}")
                lines.append(f"- Document: {doc_number}")
                lines.append(f"- Published: {pub_date}")
                if abstract:
                    lines.append(f"- Abstract: {abstract}")
                if html_url:
                    lines.append(f"- URL: {html_url}")
                lines.append("")
            return "\n".join(lines)

        # GOV.UK Content API
        if "title" in data and "body" in data:
            body = data.get("body", "")
            if isinstance(body, str) and "<" in body:
                return _html_to_markdown(body, url)
            return f"# {data['title']}\n\n{body}"

        if "title" in data and "details" in data:
            details = data.get("details", {})
            body = details.get("body", "")
            if isinstance(body, str) and "<" in body:
                return f"# {data['title']}\n\n" + _html_to_markdown(body, url)
            parts = details.get("parts", [])
            lines = [f"# {data['title']}\n"]
            for part in parts:
                part_title = part.get("title", "")
                part_body = part.get("body", "")
                lines.append(f"## {part_title}")
                if isinstance(part_body, str) and "<" in part_body:
                    lines.append(_html_to_markdown(part_body, url))
                else:
                    lines.append(str(part_body))
                lines.append("")
            return "\n".join(lines)

        # Generic JSON → Markdown
        formatted = json.dumps(data, indent=2, ensure_ascii=False, default=str)
        return f"# API Response\n\n```json\n{formatted}\n```"

    except Exception as e:
        return f"(JSON parsing failed: {e})"


def _classify_failure(error: Exception, url: str) -> str:
    """Classify the failure reason for user-friendly reporting."""
    err_str = str(error).lower()

    if isinstance(error, httpx.HTTPStatusError):
        status = error.response.status_code
        if status == 403:
            return "HTTP 403 Forbidden — 網站拒絕存取 (可能封鎖爬蟲)"
        elif status == 429:
            return "HTTP 429 Too Many Requests — 請求過於頻繁"
        elif status == 404:
            return "HTTP 404 Not Found — 頁面不存在或已移動"
        elif status == 503:
            return "HTTP 503 Service Unavailable — 伺服器暫時無法使用"
        else:
            return f"HTTP {status} — 伺服器回應錯誤"

    if isinstance(error, httpx.TimeoutException):
        return f"連線逾時 (超過 {_REQUEST_TIMEOUT} 秒) — 網站回應過慢或無法連線"

    if isinstance(error, httpx.ConnectError):
        return "無法連線至伺服器 — DNS 解析失敗或網路問題"

    if "ssl" in err_str or "certificate" in err_str:
        return "SSL 憑證錯誤 — 網站安全憑證問題"

    if "captcha" in err_str:
        return "需要驗證碼 (CAPTCHA) — 無法自動爬取"

    if "javascript" in err_str or "spa" in err_str:
        return "需要 JavaScript 渲染 — 純 HTTP 請求無法取得內容"

    return f"爬取失敗: {type(error).__name__}: {str(error)[:200]}"


# ============================================================
# DuckDuckGo Supplementary Search
# ============================================================
# Source Credibility Ranking
# Mirrors the /web search credibility tier system.
# Applied to DDG fallback URL discovery to ensure official regulatory
# sources are always preferred over general/user-generated content.
# ============================================================

# ── Tier 0 (score 100): 國際標準機構與主要法規主管機關 ──────────
_CREDIBILITY_TIER0 = frozenset([
    "iso.org", "who.int", "iec.ch", "imdrf.org",
    "fda.gov", "federalregister.gov", "ecfr.gov", "hhs.gov", "cdc.gov",
    "ema.europa.eu", "health.ec.europa.eu", "ec.europa.eu", "eur-lex.europa.eu",
    "pmda.go.jp", "mhlw.go.jp",
    "nmpa.gov.cn", "english.nmpa.gov.cn",
    "mfds.go.kr",
    "tga.gov.au",
    "hsa.gov.sg",
    "sfda.gov.sa",
    "anvisa.gov.br",
    "health.gov.il",
    "swissmedic.ch", "fedlex.admin.ch",
    "mhra.gov.uk", "legislation.gov.uk", "gov.uk",
    "law.moj.gov.tw", "tfda.gov.tw",
    "laws-lois.justice.gc.ca", "canada.ca",
    "mdsap.global",
])

# ── Tier 1 (score 80): 政府網域 TLD ────────────────────────────
_GOV_TLD_PATTERNS = (
    ".gov", ".go.jp", ".go.kr", ".gov.au", ".gov.uk", ".gov.br",
    ".gov.in", ".gov.sg", ".gov.my", ".gov.ph", ".gov.vn", ".gov.eg",
    ".gov.co", ".gov.ru", ".gc.ca", ".govt.nz", ".gob.mx",
    ".gouv.fr", ".bund.de", ".admin.ch",
    "laws.e-gov.go.jp", "legislation.gov.au", "austlii.edu.au",
)

# ── Tier 3 (score 40): 半官方法律/標準資料庫關鍵字 ────────────
_CREDIBILITY_TIER3_KEYWORDS = (
    "legal", "law", "lex", "legis", "regulation", "regulatory",
    "norme", "norma", "normat", "standard",
    "luatvietnam", "zakonrf", "consultant.ru",
    "medical-device-regulation", "medicaldevice",
)

# ── Tier 4a (score 35): 最具代表性的醫療器材/QMS 民間機構 ──────
# 包含公告機構（Notified Body）、驗證機構、產業協會、法規顧問
# 僅當 Tier 0-3 無結果時使用，所有資料必須標記出處
_CREDIBILITY_CIVIL_ORGS = frozenset([
    # 國際公告機構 / 驗證機構
    "bsigroup.com",                 # BSI Group (UK Notified Body)
    "tuvsud.com", "tuv.com",        # TÜV SÜD / TÜV Rheinland
    "dnv.com",                      # DNV
    "sgs.com",                      # SGS
    "intertek.com",                 # Intertek
    "ul.com", "ulstandards.ul.com", # UL / UL Standards
    "dekra.com",                    # DEKRA
    "kiwa.com",                     # Kiwa
    "nsf.org",                      # NSF International
    "eurofins.com",                 # Eurofins
    # 法規事務專業組織
    "raps.org",                     # Regulatory Affairs Professionals Society
    "emergobyul.com",               # Emergo by UL (法規顧問)
    # 醫療器材產業協會
    "aami.org",                     # AAMI (美國醫療儀器推進協會)
    "advamed.org",                  # AdvaMed (美國)
    "medtecheurope.org",            # MedTech Europe
    "cocir.org",                    # COCIR (歐洲放射/IT)
    "mdic.org",                     # Medical Device Innovation Consortium
    "jira.or.jp",                   # 日本醫療機器產業聯合會
    "apamed.org",                   # 亞太醫療器材協會
    "camdma.com",                   # 中國醫療器材
    # 標準與驗證資訊平台
    "isogroup.org",
    "isobudgets.com",
    "fdanews.com",
    "medicaldeviceacademy.com",
    # QMS 軟體商（具備法規知識庫）
    "greenlight.guru",
    "mastercontrol.com",
    "qualio.com",
    "etq.com",
])

# ── Tier 9 (score -1): 完全排除 ────────────────────────────────
_CREDIBILITY_EXCLUDED = frozenset([
    "wikipedia.org", "wikimedia.org", "wikidata.org",
    "reddit.com", "quora.com", "stackexchange.com", "stackoverflow.com",
    "medium.com", "substack.com", "blogspot.com", "wordpress.com",
    "tumblr.com", "weebly.com", "wix.com",
    "linkedin.com", "facebook.com", "twitter.com", "x.com",
    "youtube.com", "tiktok.com", "instagram.com",
    "amazon.com", "ebay.com", "alibaba.com",
])

# URL 全文抓取的最低門檻：
#   官方(100) + 政府(80) + 學術(60) + 半官方(40) + 代表性民間機構(35)
#   一般網頁(20) 只用於最後摘要，不抓取全文
_MIN_FETCH_CREDIBILITY = 35

# 來源層級標籤（顯示於摘要片段的出處說明）
_CREDIBILITY_LABELS = {
    100: "🏛️ 官方法規機構",
    80:  "🏛️ 政府網域",
    60:  "🎓 學術機構",
    40:  "✅ 半官方資料庫",
    35:  "🔍 代表性民間機構",
    20:  "🌐 一般網頁（最後手段）",
}


def _url_credibility_score(url: str) -> int:
    """來源可信度分數（越高越權威）。

    100  Tier 0 : 國際標準機構、主要法規主管機關
     80  Tier 1 : 政府網域（.gov、.go.jp 等）
     60  Tier 2 : 學術機構（.edu、.ac.uk 等）
     40  Tier 3 : 半官方法律/標準資料庫
     35  Tier 4a: 最具代表性的醫療器材/QMS 民間機構（含 SGS、UL、TÜV 等）
     20  Tier 4b: 一般網頁（僅用於摘要片段，需標記出處）
     -1  Tier 9 : Wikipedia、社群媒體 — 完全排除
    """
    if not url:
        return 0
    try:
        from urllib.parse import urlparse
        host = urlparse(url).netloc.lower().lstrip("www.")
    except Exception:
        return 20

    for excl in _CREDIBILITY_EXCLUDED:
        if host == excl or host.endswith("." + excl):
            return -1

    for t0 in _CREDIBILITY_TIER0:
        if host == t0 or host.endswith("." + t0):
            return 100

    for pat in _GOV_TLD_PATTERNS:
        if host.endswith(pat) or pat in host:
            return 80

    if any(host.endswith(s) for s in (".edu", ".ac.uk", ".ac.jp", ".ac.kr",
                                       ".ac.au", ".ac.nz", ".ac.in")):
        return 60

    if any(k in host for k in _CREDIBILITY_TIER3_KEYWORDS):
        return 40

    for org in _CREDIBILITY_CIVIL_ORGS:
        if host == org or host.endswith("." + org):
            return 35

    return 20  # Tier 4b: 一般網頁


def _credibility_label(score: int) -> str:
    """回傳對應可信度分數的中文標籤。"""
    for threshold in sorted(_CREDIBILITY_LABELS.keys(), reverse=True):
        if score >= threshold:
            return _CREDIBILITY_LABELS[threshold]
    return "🌐 來源不明"


def _sort_by_credibility(search_results: list, min_score: int = 0) -> list:
    """依來源可信度排序 DDG 搜尋結果。

    min_score:
      0  — 僅排除 Tier 9（摘要片段，所有結果都顯示但需標記出處）
      35 — 只留 Tier 0-4a（URL 全文抓取用，排除隨機一般網頁）
    """
    scored = []
    for sr in search_results:
        url = sr.get("href") or sr.get("link") or ""
        score = _url_credibility_score(url)
        if score >= min_score:
            scored.append((score, sr))
    scored.sort(key=lambda x: x[0], reverse=True)
    return [sr for _, sr in scored]


# ============================================================


def _ddgs_search(query: str, max_results: int = 5) -> list:
    """Search DuckDuckGo for supplementary regulatory info.

    Appends site-bias operators to steer results toward official/government sources.
    Results are returned as-is; credibility filtering is applied by the caller.
    """
    if DDGS is None:
        return []
    # Bias query toward official regulatory / government / civil org sources.
    # This does not guarantee results — DDG may still return general pages —
    # but significantly increases the proportion of authoritative results.
    # Our _sort_by_credibility filter handles the final selection.
    biased_query = (
        f"{query} "
        f"(site:.gov OR site:.go.jp OR site:.gov.au OR site:.gc.ca OR "
        f"site:iso.org OR site:who.int OR site:iec.ch OR site:imdrf.org OR "
        f"site:legislation.gov.uk OR site:eur-lex.europa.eu OR "
        f"site:sgs.com OR site:tuvsud.com OR site:ul.com OR site:bsigroup.com)"
    )
    try:
        with DDGS() as ddgs:
            results = list(ddgs.text(biased_query, max_results=max_results))
        # If biased query returns nothing, fall back to plain query
        if not results:
            with DDGS() as ddgs:
                results = list(ddgs.text(query, max_results=max_results))
        return results
    except Exception as e:
        # M8: classify failure type for actionable logging
        err = str(e).lower()
        if "rate" in err or "202" in err or "blocked" in err or "ratelimit" in err:
            logger.warning("DDG rate-limited: %s", str(e)[:120])
        elif "connect" in err or "timeout" in err or "network" in err:
            logger.warning("DDG network error (%s): %s", type(e).__name__, str(e)[:120])
        else:
            logger.warning("DDG search failed (%s): %s", type(e).__name__, str(e)[:120])
        return []


import re as _re

_DDG_CITATION_RE = _re.compile(
    r"(?:No\.\s*\d+[\w/]*"
    r"|SOR[\-/]\d+[\-/]\d+"
    r"|RDC\s*\w+[/]\d+"
    r"|\d+\s*CFR\s*Part\s*\d+"
    r"|Ordinance\s*(?:No\.)?\s*\d+"
    r"|Decree\s*\d+[/\-]\d+"
    r"|Act\s*\d+"
    r"|SI\s*\d+[/]\d+"
    r"|SR\s*\d+\.\d+"
    r"|RDC\s*n[oº°]\s*[\d/]+"
    r"|ISO\s*\d+(?::\d+)?)",
    _re.IGNORECASE,
)


def _build_ddg_query(site: dict, region: str) -> str:
    """M2: Build targeted DDG query using regulation citation extracted from site metadata."""
    name = site.get("name", "")
    note = site.get("note", "")
    agency = site.get("agency", "")

    en_name = region
    if "(" in region and ")" in region:
        en_name = region.split("(")[1].rstrip(")")

    citations = _DDG_CITATION_RE.findall(name + " " + note)
    if citations:
        citation = citations[0].strip()
        return f'"{citation}" {en_name} medical device regulation full text'

    # Fall back to first 70 chars of the regulation name
    short_name = name[:70].strip().rstrip("—").strip()
    return f"{en_name} {short_name} regulation full text"


_REGULATORY_KEYWORDS = frozenset([
    "article", "section", "chapter", "clause", "regulation", "requirement",
    "shall", "must", "pursuant", "compliance", "manufacturer", "annex",
    "条", "章", "第", "条文", "令", "artikel", "abschnitt", "paragraf",
    "artículo", "sección", "capítulo", "artigo", "seção", "decreto",
    "постановление", "статья", "глава",
])


def _is_regulatory_fulltext(content: str) -> bool:
    """M3: Heuristic — does this content look like regulatory full text vs an intro/index page?"""
    if len(content) < 1500:
        return False
    lines = [ln for ln in content.splitlines() if ln.strip()]
    if len(lines) < 25:
        return False
    content_lower = content.lower()
    hits = sum(1 for kw in _REGULATORY_KEYWORDS if kw in content_lower)
    return hits >= 3


# ============================================================
# Fetch with retry (tenacity)
# ============================================================


async def _fetch_with_retry(
    client: httpx.AsyncClient,
    url: str,
    headers: Optional[dict] = None,
    timeout: Optional[httpx.Timeout] = None,
) -> httpx.Response:
    """Fetch a URL with exponential-backoff retry.

    Retries on connection errors, timeouts, and 5xx status codes.
    """
    if not _is_safe_url(url):
        raise ValueError(f"Blocked unsafe URL: {url}")

    _timeout = timeout or httpx.Timeout(_REQUEST_TIMEOUT, connect=10.0)
    _headers = dict(_DEFAULT_HEADERS)
    if headers:
        _headers.update(headers)

    if TENACITY_AVAILABLE:

        @retry(
            stop=stop_after_attempt(3),
            wait=wait_exponential(multiplier=1, min=2, max=10),
            retry=retry_if_exception_type((httpx.ConnectError, httpx.TimeoutException)),
            reraise=True,
        )
        async def _inner():
            resp = await client.get(url, headers=_headers, timeout=_timeout)
            # Retry on 5xx
            if resp.status_code >= 500:
                resp.raise_for_status()
            return resp

        return await _inner()
    else:
        # No tenacity: single attempt
        resp = await client.get(url, headers=_headers, timeout=_timeout)
        return resp


async def _fetch_bytes_bounded(
    client: httpx.AsyncClient,
    url: str,
    timeout: Optional[httpx.Timeout] = None,
    max_bytes: int = _DOWNLOAD_MAX_BYTES,
) -> bytes:
    """Stream-download URL with a hard per-file size cap (default 200 MB).

    Enforces the cap *before* the full body lands in RAM, guarding against
    oversized or malicious responses from compromised regulatory sites.
    """
    if not _is_safe_url(url):
        raise ValueError(f"Blocked unsafe URL: {url}")
    _timeout = timeout or httpx.Timeout(_REQUEST_TIMEOUT, connect=10.0)
    _headers = dict(_DEFAULT_HEADERS)

    async def _stream_once() -> bytes:
        async with client.stream("GET", url, headers=_headers, timeout=_timeout) as resp:
            resp.raise_for_status()
            cl = int(resp.headers.get("content-length", 0))
            if cl > max_bytes:
                raise ValueError(
                    f"Content-Length {cl // (1024 * 1024)} MB exceeds "
                    f"{max_bytes // (1024 * 1024)} MB download limit"
                )
            chunks: list[bytes] = []
            received = 0
            async for chunk in resp.aiter_bytes(65536):
                received += len(chunk)
                if received > max_bytes:
                    raise ValueError(
                        f"Download exceeded {max_bytes // (1024 * 1024)} MB limit, aborting"
                    )
                chunks.append(chunk)
        return b"".join(chunks)

    if TENACITY_AVAILABLE:
        @retry(
            stop=stop_after_attempt(3),
            wait=wait_exponential(multiplier=1, min=2, max=10),
            retry=retry_if_exception_type((httpx.ConnectError, httpx.TimeoutException)),
            reraise=True,
        )
        async def _inner() -> bytes:
            return await _stream_once()

        return await _inner()
    return await _stream_once()


def _is_zip_bomb(data: bytes, max_uncompressed: int = _OFFICE_ZIP_BOMB_LIMIT) -> bool:
    """Return True if the ZIP archive total uncompressed size exceeds the limit.

    Protects against malicious .docx/.xlsx/.pptx files that decompress to
    hundreds of MB from a small compressed payload.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            total = sum(info.file_size for info in zf.infolist())
            return total > max_uncompressed
    except Exception:
        return False  # Not a valid ZIP; let the format processor handle it


# ============================================================
# Tier Handlers
# ============================================================


def _make_result_template(site: dict, region: str) -> dict:
    """Create a result dict template with default values."""
    return {
        "region": region,
        "agency": site.get("agency", "Unknown"),
        "agency_name": site.get("name", site.get("agency", "")),
        "url": site.get("url", ""),
        "title": "",
        "content_markdown": "",
        "crawl_status": "failed",
        "content_source": None,
        "failure_reason": None,
        "crawl_timestamp": datetime.now(timezone.utc).isoformat(),
        "crawl_duration_seconds": 0.0,
        "note": site.get("note", ""),
    }


def _retrieve_cached_content(url: str) -> Optional[str]:
    """Retrieve previously crawled content for a URL from regulatory markdown storage.

    Used when HTTP 304 Not Modified is received — instead of storing a useless
    '(Content unchanged)' placeholder, we restore the actual content from the
    last successful crawl so downstream consumers (LLM analysis, reports) get
    real regulatory text.
    """
    try:
        from src.storage.regulatory_markdown_storage import (
            get_regulatory_markdown_store,
        )

        store = get_regulatory_markdown_store()
        doc = store.get_document_by_url(url)
        if doc and doc.get("content"):
            return doc["content"]
    except Exception:
        pass
    return None


async def _crawl_bulk_zip(site: dict, region: str) -> dict:
    """Bulk ZIP strategy: Taiwan MOJ Open API — download all medical device laws.

    Downloads ChOrder.json.zip + ChLaw.json.zip (cached for 7 days),
    filters active TFDA medical device regulations (26 laws), converts each
    to structured Markdown, and returns a single merged crawl result containing
    all laws separated by clear section dividers.

    Each individual law is also stored as ``_bulk_sub_results`` on the result
    so callers (e.g. save_from_crawl_results) can optionally split them.
    """
    result = _make_result_template(site, region)
    start  = time.time()

    try:
        from src.services.taiwan_bulk_api import fetch_taiwan_laws_bulk

        sub_results = await asyncio.to_thread(
            fetch_taiwan_laws_bulk,
            use_cache=True,
            save_individual_files=False,
        )

        if not sub_results:
            result["crawl_status"]  = "failed"
            result["failure_reason"] = "bulk_zip: no laws returned from taiwan_bulk_api"
            return result

        # Build merged markdown with clear per-law sections
        parts: list[str] = []
        for r in sub_results:
            agency  = r.get("agency", "")
            title   = r.get("title", "")
            content = r.get("content_markdown", "")
            meta    = r.get("_law_metadata", {})
            pcode   = meta.get("pcode", "")
            parts.append(
                f"<!-- LAW_START: {pcode} agency={agency} -->\n"
                f"{content}\n"
                f"<!-- LAW_END: {pcode} -->"
            )

        merged = (
            f"# 台灣醫療器材法規全集 (Taiwan Medical Device Regulatory Bundle)\n\n"
            f"**來源**: MOJ Open API Bulk Download  \n"
            f"**法規數量**: {len(sub_results)} 部現行醫療器材法規  \n"
            f"**下載時間**: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  \n\n"
            f"---\n\n"
        ) + "\n\n---\n\n".join(parts)

        result["crawl_status"]    = "success"
        result["content_source"]  = "bulk_api"
        result["content_markdown"] = merged
        result["title"]           = f"Taiwan Medical Device Regulatory Bundle ({len(sub_results)} laws)"
        result["note"]            = (
            f"MOJ Bulk API: {len(sub_results)} active medical device regulations — "
            f"ChOrder + ChLaw — "
            + ", ".join(r.get("_law_metadata", {}).get("pcode", "") for r in sub_results[:5])
            + " …"
        )
        result["_bulk_sub_results"] = sub_results

    except Exception as exc:
        result["crawl_status"]   = "failed"
        result["failure_reason"] = f"bulk_zip error: {exc}"
        logger.error("_crawl_bulk_zip failed for %s/%s: %s", region, site.get("agency"), exc)

    result["crawl_duration_seconds"] = round(time.time() - start, 2)
    return result


async def _crawl_tier1_api(
    client: httpx.AsyncClient,
    site: dict,
    region: str,
    etag_cache: ETagCache,
) -> dict:
    """Tier 1: Structured API / RSS / JSON endpoints."""
    result = _make_result_template(site, region)
    url = site["url"]
    start = time.time()

    try:
        # Conditional request headers
        req_headers = {}
        cached = etag_cache.get(url)
        if cached:
            if cached.get("etag"):
                req_headers["If-None-Match"] = cached["etag"]
            if cached.get("last_modified"):
                req_headers["If-Modified-Since"] = cached["last_modified"]

        response = await _fetch_with_retry(client, url, headers=req_headers)

        # 304 Not Modified — if no cached content exists, retry without conditional headers
        if response.status_code == 304:
            previous_content = _retrieve_cached_content(url)
            if previous_content:
                _cached_entry = etag_cache.get(url) or {}
                result["crawl_status"] = "success"
                result["content_source"] = "cached"  # B-2: not "live" — restored from previous crawl
                result["title"] = f"{site['agency']} (cached — not modified)"
                result["content_markdown"] = previous_content
                result["note"] = (
                    "HTTP 304 Not Modified — restored content from previous crawl"
                    + (f" (cached_at: {_cached_entry.get('cached_at', 'unknown')})" if _cached_entry.get("cached_at") else "")
                )
                result["crawl_duration_seconds"] = round(time.time() - start, 2)
                return result
            # No cached content — force fresh request without conditional headers
            response = await _fetch_with_retry(client, url, headers={})

        response.raise_for_status()

        # Update ETag cache
        etag = response.headers.get("ETag")
        last_mod = response.headers.get("Last-Modified")
        content_hash = hashlib.sha256(response.content).hexdigest()[:16]
        etag_cache.set(
            url, etag=etag, last_modified=last_mod, content_hash=content_hash
        )

        # Parse JSON / RSS / Atom / HTML
        content_type = response.headers.get("content-type", "")
        is_rss = (
            "rss" in content_type
            or "atom" in content_type
            or site.get("strategy") in ("rss", "atom")
            or (url.endswith(".rss") or url.endswith(".atom") or "rss" in url or "atom" in url)
        )
        if is_rss:
            # G6: RSS/Atom feed parsing — extract items as structured Markdown
            try:
                rss_md, rss_title, rss_items = _parse_rss_feed(response.text, site["agency"])
                result["content_markdown"] = rss_md
                result["title"] = rss_title
                result["_rss_items"] = rss_items  # list of {title, link, pubDate, description}
                result["content_source"] = "rss"
            except Exception as _rss_err:
                result["content_markdown"] = _html_to_markdown(response.text, url)
                result["title"] = f"{site['agency']} Feed"
                logger.debug("RSS parse fallback for %s: %s", url, _rss_err)
        elif "application/json" in content_type or site.get("strategy") == "api_json":
            try:
                json_data = response.json()
                result["content_markdown"] = _json_to_markdown(json_data, url)
                result["title"] = (
                    json_data.get("title", "")
                    or json_data.get("meta", {}).get("title", "")
                    or f"{site['agency']} API Response"
                )
            except Exception:
                result["content_markdown"] = _html_to_markdown(response.text, url)
                result["title"] = f"{site['agency']} Response"
        else:
            # RSS/XML/HTML fallback
            result["content_markdown"] = _html_to_markdown(response.text, url)
            try:
                soup = BeautifulSoup(response.text, "lxml")
                title_tag = soup.find("title")
                result["title"] = (
                    title_tag.string.strip()
                    if title_tag and title_tag.string
                    else site["agency"]
                )
            except Exception:
                result["title"] = site["agency"]

        if result["content_markdown"] and len(result["content_markdown"].strip()) > 50:
            result["crawl_status"] = "success"
            result["content_source"] = "live"
        else:
            result["failure_reason"] = (
                "頁面內容為空或需要 JavaScript 渲染 — "
                "網站可能為 SPA 架構，純 HTTP 請求無法取得實際內容"
            )

    except Exception as e:
        result["failure_reason"] = _classify_failure(e, url)

    result["crawl_duration_seconds"] = round(time.time() - start, 2)
    return result


# ── Section markers for QMS-relevant sections in major regulatory PDFs ────────
# Used by _extract_qms_sections_from_pdf() to locate and extract specific parts.
_QMS_SECTION_MARKERS: list[tuple[str, str]] = [
    # EU MDR 2017/745 — the most critical sections for QMS
    ("ARTICLE 10", "ARTICLE 11"),      # General obligations of manufacturers (QMS mandate)
    ("ANNEX IX", "ANNEX X"),           # Conformity assessment based on QMS (full NB audit scope)
    ("ANNEX XI", "ANNEX XII"),         # Product Quality Assurance (alternative route)
    ("ANNEX I", "ANNEX II"),           # GSPR — referenced in QMS risk management
    # Generic markers that appear in most medical device regulations
    ("QUALITY MANAGEMENT SYSTEM", None),  # Grab context around any QMS section
    ("QUALITY SYSTEM", None),
]

_QMS_SECTION_MAX_CHARS = 200_000  # 200 KB — enough for Annex IX + Article 10 combined


def _extract_qms_sections_from_pdf(pdf_bytes: bytes, display_name: str, source_url: str) -> str:
    """Method 1: Section-aware PDF extraction for large regulatory documents.

    Uses PyMuPDF to scan page text for QMS section markers (Article 10, Annex IX, etc.)
    and extracts only those pages, skipping non-QMS content.

    This allows processing a 600 KB regulation text while only sending the
    QMS-relevant ~100 KB to storage and analysis.

    Returns extracted QMS sections as Markdown, or "" if PyMuPDF unavailable.
    """
    if not FITZ_AVAILABLE:
        return ""
    try:
        doc = _fitz.open(stream=pdf_bytes, filetype="pdf")
        n = len(doc)
        if n == 0:
            doc.close()
            return ""

        # Extract all page texts first (needed for section boundary detection)
        page_texts = []
        for page in doc:
            page_texts.append(page.get_text("text") or "")
        doc.close()

        # Find page ranges for each section marker
        section_pages: set[int] = set()
        for start_marker, end_marker in _QMS_SECTION_MARKERS:
            in_section = False
            for i, text in enumerate(page_texts):
                text_upper = text.upper()
                if not in_section and start_marker in text_upper:
                    in_section = True
                if in_section:
                    section_pages.add(i)
                    # Also include next 2 pages for context
                    if i + 1 < n:
                        section_pages.add(i + 1)
                    if i + 2 < n:
                        section_pages.add(i + 2)
                if in_section and end_marker and end_marker in text_upper and i > min(section_pages):
                    in_section = False

        if not section_pages:
            # No section markers found — fall back to first N chars of full text
            full_text = "\n".join(page_texts)
            if not full_text.strip():
                return ""
            header = f"# {display_name}\n\n**Source**: {source_url}  \n**Pages**: {n} (full)\n\n---\n\n"
            return (header + full_text)[:_QMS_SECTION_MAX_CHARS]

        # Build section-extracted text
        parts = []
        total_chars = 0
        for i in sorted(section_pages):
            text = page_texts[i].strip()
            if text:
                parts.append(f"\n<!-- Page {i+1}/{n} -->\n{text}\n")
                total_chars += len(text)
            if total_chars >= _QMS_SECTION_MAX_CHARS:
                break

        if not parts:
            return ""
        header = (
            f"# {display_name}\n\n"
            f"**Source**: {source_url}  \n"
            f"**Extraction**: QMS sections only ({len(section_pages)}/{n} pages)\n\n---\n\n"
        )
        return (header + "\n".join(parts))[:_QMS_SECTION_MAX_CHARS]
    except Exception:
        return ""


def _pdf_bytes_to_markdown(pdf_bytes: bytes, display_name: str, source_url: str,
                            section_aware: bool = False) -> str:
    """Extract text from a PDF, returning Markdown.

    Pipeline:
    1. Section-aware extraction (if section_aware=True) — extracts only QMS sections
    2. MarkItDown (primary) — unified conversion engine, handles most PDFs
    3. PyMuPDF (fallback)   — direct text layer extraction
    Returns "" if no text found → triggers Docling OCR fallback in caller.
    """
    import os, tempfile

    # 0. Section-aware extraction for large regulatory PDFs (Method 1 — EU MDR etc.)
    if section_aware and len(pdf_bytes) > 500_000:  # only for PDFs > 500 KB
        sections_md = _extract_qms_sections_from_pdf(pdf_bytes, display_name, source_url)
        if sections_md and len(sections_md.strip()) > 500:
            return sections_md

    # 1. Try MarkItDown first (same engine used for Word/Excel/PPT)
    if MARKITDOWN_AVAILABLE and _MD_CONVERTER is not None:
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
                f.write(pdf_bytes)
                tmp_path = f.name
            result = _MD_CONVERTER.convert(tmp_path)
            md = result.text_content if hasattr(result, "text_content") else str(result)
            if md and len(md.strip()) > 100:
                header = (
                    f"# {display_name}\n\n"
                    f"**Source**: {source_url}  \n"
                    f"**Format**: PDF (MarkItDown)\n\n---\n\n"
                )
                return header + md.strip()
        except Exception:
            pass
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

    # 2. PyMuPDF fallback — direct text layer extraction
    if not FITZ_AVAILABLE:
        return ""
    try:
        doc = _fitz.open(stream=pdf_bytes, filetype="pdf")
        n = len(doc)
        text_parts = []
        for i, page in enumerate(doc):
            text = page.get_text("text").strip()
            if text:
                text_parts.append(f"\n---\n<!-- Page {i+1}/{n} -->\n\n{text}\n")
        doc.close()
        if not text_parts:
            return ""  # No text layer — triggers Docling OCR fallback
        header = f"# {display_name}\n\n**Source**: {source_url}  \n**Pages**: {n}\n\n---\n"
        return header + "\n".join(text_parts)
    except Exception:
        return ""


# Max PDF size for real-time docling OCR during live crawl (larger = offline batch only)
_PDF_DOCLING_MAX_BYTES_REALTIME = 2 * 1024 * 1024  # 2 MB
_PDF_ATTACHMENT_MAX_BYTES = 5 * 1024 * 1024         # 5 MB per attachment
_MAX_PDF_ATTACHMENTS_PER_PAGE = 3  # legacy — superseded by _MAX_FILE_ATTACHMENTS_PER_PAGE

# Deep-crawl limits
_MAX_FILE_ATTACHMENTS_PER_PAGE = 50       # PDF + Word + Excel + PPT combined per page
_MAX_FILE_ATTACHMENTS_INDEX_PAGE = 200   # for document index/listing pages (index_page=True)
_MAX_SUBPAGES_PER_PAGE = 5                # same-domain HTML sub-pages to follow


async def _docling_pdf_bytes_to_markdown_async(
    pdf_bytes: bytes, display_name: str, source_url: str
) -> str:
    """OCR-extract a scanned PDF using Docling (runs in a thread to keep event loop free).

    Returns an empty string if Docling is unavailable or extraction fails.
    """
    import os
    import tempfile

    def _run_sync() -> str:
        try:
            from src.ocr.docling_engine import get_engine
        except ImportError:
            return ""
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
                f.write(pdf_bytes)
                tmp_path = f.name
            res = get_engine().parse(tmp_path, force_engine="docling")
            if res.success and res.markdown and len(res.markdown.strip()) > 100:
                header = (
                    f"# {display_name}\n\n"
                    f"**Source**: {source_url}  \n"
                    f"**Pages**: {res.page_count} (OCR via Docling)\n\n---\n\n"
                )
                return header + res.markdown
        except Exception:
            pass
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
        return ""

    return await asyncio.to_thread(_run_sync)


_WORD_EXTS = frozenset({".doc", ".docx"})
_WORD_ATTACHMENT_MAX_BYTES = 10 * 1024 * 1024  # 10 MB

_EXCEL_EXTS = frozenset({".xlsx", ".xls", ".ods", ".csv"})
_EXCEL_ATTACHMENT_MAX_BYTES = 10 * 1024 * 1024  # 10 MB

_PPT_EXTS = frozenset({".pptx", ".ppt"})
_PPT_ATTACHMENT_MAX_BYTES = 20 * 1024 * 1024  # 20 MB

_ALL_FILE_EXTS = frozenset({".pdf"}) | _WORD_EXTS | _EXCEL_EXTS | _PPT_EXTS


async def _word_bytes_to_markdown_async(
    word_bytes: bytes, display_name: str, source_url: str
) -> str:
    """Convert a Word (.doc/.docx) file to Markdown using MarkItDown in a thread."""
    import os
    import tempfile

    suffix = ".docx" if word_bytes[:4] in (b"PK\x03\x04",) else ".doc"

    def _run_sync() -> str:
        if not MARKITDOWN_AVAILABLE or _MD_CONVERTER is None:
            return ""
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as f:
                f.write(word_bytes)
                tmp_path = f.name
            result = _MD_CONVERTER.convert(tmp_path)
            md = result.text_content if hasattr(result, "text_content") else str(result)
            if md and len(md.strip()) > 100:
                header = (
                    f"# {display_name}\n\n"
                    f"**Source**: {source_url}  \n"
                    f"**Format**: Word document\n\n---\n\n"
                )
                return header + md.strip()
        except Exception:
            pass
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
        return ""

    return await asyncio.to_thread(_run_sync)


async def _excel_bytes_to_markdown_async(
    excel_bytes: bytes, display_name: str, source_url: str, ext: str = ".xlsx"
) -> str:
    """Convert an Excel/ODS/CSV file to Markdown tables using openpyxl/pandas fallback."""
    import os
    import tempfile

    def _run_sync() -> str:
        tmp_path = None
        try:
            # Try MarkItDown first (handles xlsx natively)
            if MARKITDOWN_AVAILABLE and _MD_CONVERTER is not None and ext in {".xlsx", ".xls"}:
                with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as f:
                    f.write(excel_bytes)
                    tmp_path = f.name
                result = _MD_CONVERTER.convert(tmp_path)
                md = result.text_content if hasattr(result, "text_content") else str(result)
                if md and len(md.strip()) > 50:
                    header = (
                        f"# {display_name}\n\n"
                        f"**Source**: {source_url}  \n"
                        f"**Format**: Excel spreadsheet\n\n---\n\n"
                    )
                    return header + md.strip()
        except Exception:
            pass
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass

        # Fallback: openpyxl for xlsx
        if ext == ".xlsx":
            try:
                import openpyxl
                import io as _io
                wb = openpyxl.load_workbook(_io.BytesIO(excel_bytes), read_only=True, data_only=True)
                parts = [f"# {display_name}\n\n**Source**: {source_url}  \n**Format**: Excel spreadsheet\n\n---\n"]
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue
                    parts.append(f"\n## Sheet: {sheet_name}\n")
                    header_row = [str(c) if c is not None else "" for c in rows[0]]
                    parts.append("| " + " | ".join(header_row) + " |")
                    parts.append("| " + " | ".join(["---"] * len(header_row)) + " |")
                    for row in rows[1:50]:  # cap at 50 rows to avoid token explosion
                        parts.append("| " + " | ".join(str(c) if c is not None else "" for c in row) + " |")
                wb.close()
                result = "\n".join(parts)
                if len(result.strip()) > 50:
                    return result
            except Exception:
                pass

        # CSV fallback
        if ext == ".csv":
            try:
                import csv
                import io as _io
                text = excel_bytes.decode("utf-8-sig", errors="replace")
                reader = csv.reader(_io.StringIO(text))
                rows = list(reader)
                if rows:
                    parts = [
                        f"# {display_name}\n\n**Source**: {source_url}  \n**Format**: CSV\n\n---\n",
                        "| " + " | ".join(rows[0]) + " |",
                        "| " + " | ".join(["---"] * len(rows[0])) + " |",
                    ]
                    for row in rows[1:50]:
                        parts.append("| " + " | ".join(row) + " |")
                    return "\n".join(parts)
            except Exception:
                pass

        return ""

    return await asyncio.to_thread(_run_sync)


async def _ppt_bytes_to_markdown_async(
    ppt_bytes: bytes, display_name: str, source_url: str, ext: str = ".pptx"
) -> str:
    """Convert a PowerPoint file to Markdown using MarkItDown."""
    import os
    import tempfile

    def _run_sync() -> str:
        if not MARKITDOWN_AVAILABLE or _MD_CONVERTER is None:
            return ""
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as f:
                f.write(ppt_bytes)
                tmp_path = f.name
            result = _MD_CONVERTER.convert(tmp_path)
            md = result.text_content if hasattr(result, "text_content") else str(result)
            if md and len(md.strip()) > 100:
                header = (
                    f"# {display_name}\n\n"
                    f"**Source**: {source_url}  \n"
                    f"**Format**: PowerPoint presentation\n\n---\n\n"
                )
                return header + md.strip()
        except Exception:
            pass
        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
        return ""

    return await asyncio.to_thread(_run_sync)


async def _extract_html_pdf_attachments(
    client: httpx.AsyncClient,
    html: str,
    base_url: str,
    max_attachments: int = _MAX_FILE_ATTACHMENTS_PER_PAGE,
    max_subpages: int = _MAX_SUBPAGES_PER_PAGE,
    jina_subpage_fallback: bool = False,
    jina_semaphore: asyncio.Semaphore | None = None,
) -> str:
    """Deep-crawl all linked documents and same-domain sub-pages found in an HTML page.

    Scans every <a href> tag and classifies links into:
      - File attachments: PDF, Word (.doc/.docx), Excel (.xlsx/.xls/.ods/.csv), PPT (.pptx/.ppt)
      - Same-domain HTML sub-pages (depth=1, capped at _MAX_SUBPAGES_PER_PAGE)

    All content is downloaded, converted to Markdown, and appended to the result.
    """
    from urllib.parse import urljoin, urlparse as _up2

    # (title, absolute_url, extension_lowercase)
    file_links: list[tuple[str, str, str]] = []
    subpage_links: list[tuple[str, str]] = []
    seen_urls: set[str] = set()

    base_domain = _up2(base_url).netloc

    try:
        soup = BeautifulSoup(html, "lxml")
        for a in soup.find_all("a", href=True):
            href = a["href"].strip()
            if not href or href.startswith("#") or href.lower().startswith("javascript:"):
                continue
            full_url = urljoin(base_url, href)
            if not _is_safe_url(full_url) or full_url in seen_urls:
                continue
            seen_urls.add(full_url)

            title = (
                a.get_text(strip=True)
                or _up2(full_url).path.rstrip("/").split("/")[-1]
                or "Document"
            )[:120]

            # Determine extension: check path first, then filename= query parameter
            # (e.g. /document/download/UUID_en?filename=mdcg.pdf — ext is in query)
            parsed_url = _up2(full_url)
            path_lower = parsed_url.path.lower()
            matched_ext = next(
                (ext for ext in sorted(_ALL_FILE_EXTS, key=len, reverse=True) if path_lower.endswith(ext)),
                None,
            )
            if not matched_ext:
                # Check filename= query parameter (EC document portal, MDCG PDFs, etc.)
                qs = parsed_url.query.lower()
                fn_match = re.search(r'filename=([^&]+)', qs)
                if fn_match:
                    fname = fn_match.group(1)
                    matched_ext = next(
                        (ext for ext in sorted(_ALL_FILE_EXTS, key=len, reverse=True) if fname.endswith(ext)),
                        None,
                    )
                    if matched_ext and not title.strip():
                        # Use filename as title when link text is empty
                        title = fname[:120]

            if matched_ext:
                if len(file_links) < max_attachments:
                    file_links.append((title, full_url, matched_ext))
            else:
                # Same-domain HTML sub-page
                if _up2(full_url).netloc == base_domain and len(subpage_links) < max_subpages:
                    subpage_links.append((title, full_url))
    except Exception:
        return ""

    parts: list[str] = []

    # ── File attachments ──
    for title, att_url, ext in file_links:
        try:
            att_bytes = await _fetch_bytes_bounded(
                client, att_url, timeout=httpx.Timeout(60, connect=10)
            )
            if len(att_bytes) < 100:
                continue

            md = ""
            if ext == ".pdf":
                if att_bytes[:4] != b"%PDF" and b"%PDF" not in att_bytes[:1024]:
                    continue
                md = _pdf_bytes_to_markdown(att_bytes, title, att_url)
                if not md and len(att_bytes) <= _PDF_ATTACHMENT_MAX_BYTES:
                    md = await _docling_pdf_bytes_to_markdown_async(att_bytes, title, att_url)
            elif ext in _WORD_EXTS:
                if len(att_bytes) <= _WORD_ATTACHMENT_MAX_BYTES and not _is_zip_bomb(att_bytes):
                    md = await _word_bytes_to_markdown_async(att_bytes, title, att_url)
            elif ext in _EXCEL_EXTS:
                if len(att_bytes) <= _EXCEL_ATTACHMENT_MAX_BYTES and not _is_zip_bomb(att_bytes):
                    md = await _excel_bytes_to_markdown_async(att_bytes, title, att_url, ext)
            elif ext in _PPT_EXTS:
                if len(att_bytes) <= _PPT_ATTACHMENT_MAX_BYTES and not _is_zip_bomb(att_bytes):
                    md = await _ppt_bytes_to_markdown_async(att_bytes, title, att_url, ext)

            if md and len(md.strip()) > 200:
                parts.append(md)
        except Exception:
            pass

    # ── Same-domain sub-page links (depth = 1) ──
    _jina_sem = jina_semaphore or asyncio.Semaphore(2)
    for title, sub_url in subpage_links:
        md = ""
        page_title = title
        # Try direct httpx first
        try:
            resp = await _fetch_with_retry(
                client, sub_url, timeout=httpx.Timeout(30, connect=10)
            )
            resp.raise_for_status()
            sub_html = resp.text
            md = _html_to_markdown(sub_html, sub_url)
            try:
                sub_soup = BeautifulSoup(sub_html, "lxml")
                t = sub_soup.find("title")
                page_title = t.string.strip() if t and t.string else title
            except Exception:
                pass
        except Exception:
            # Fallback: try Jina Reader for this sub-page
            if jina_subpage_fallback:
                try:
                    async with _jina_sem:
                        jina_url = f"{_JINA_READER_BASE}{sub_url}"
                        jresp = await _fetch_with_retry(
                            client, jina_url,
                            headers={"Accept": "text/markdown"},
                            timeout=httpx.Timeout(_JINA_TIMEOUT, connect=15.0),
                        )
                        await asyncio.sleep(max(_JINA_DELAY - 1.0, 1.0))
                    if jresp.status_code == 200:
                        jcontent = jresp.text.strip()
                        _head = jcontent[:600]
                        _blocked = (
                            ("Warning:" in _head and "returned error" in _head)
                            or "JavaScript is disabled" in _head
                            or "Max challenge" in _head
                        )
                        if jcontent and len(jcontent) > 100 and not _blocked:
                            md = jcontent
                            for line in jcontent.split("\n"):
                                if line.strip().startswith("# "):
                                    page_title = line.strip()[2:]
                                    break
                except Exception:
                    pass

        if md and len(md.strip()) > 100:
            parts.append(
                f"# {page_title}\n\n**Source**: {sub_url}\n\n---\n\n{md.strip()}"
            )

    if not parts:
        return ""

    _SEP = "\n\n<!-- PDF_DOC_BREAK -->\n\n"
    return "\n\n---\n<!-- ATTACHED DOCUMENTS & LINKED PAGES -->\n\n" + _SEP.join(parts)


async def _extract_jina_subpages(
    client: httpx.AsyncClient,
    markdown: str,
    base_url: str,
    max_subpages: int = _MAX_SUBPAGES_PER_PAGE,
    jina_semaphore: asyncio.Semaphore | None = None,
) -> str:
    """Follow same-domain HTML sub-page links found in Jina-returned Markdown.

    Used when the main page (e.g. legislation.gov.uk TOC) is fetched via Jina and
    the article sub-pages also need Jina to bypass bot protection.
    Finds [text](url) links pointing to same domain, fetches each via Jina, and
    returns combined Markdown of all successfully retrieved sub-pages.
    """
    from urllib.parse import urljoin, urlparse as _up2

    _link_re = re.compile(r'\[(?:[^\]]*)\]\((https?://[^)\s]+)\)')
    base_domain = _up2(base_url).netloc
    _jina_sem = jina_semaphore or asyncio.Semaphore(2)

    seen: set[str] = set()
    subpage_urls: list[str] = []

    for m in _link_re.finditer(markdown):
        href = m.group(1).strip()
        if not href or not _is_safe_url(href):
            continue
        parsed = _up2(href)
        if parsed.netloc != base_domain:
            continue
        # Skip file attachments (those are handled by _extract_markdown_attachments)
        path_lower = parsed.path.lower()
        if any(path_lower.endswith(ext) for ext in _ALL_FILE_EXTS):
            continue
        # Skip anchors and query-only links
        if not parsed.path or parsed.path == _up2(base_url).path:
            continue
        if href not in seen:
            seen.add(href)
            subpage_urls.append(href)
        if len(subpage_urls) >= max_subpages:
            break

    if not subpage_urls:
        return ""

    parts: list[str] = []
    for sub_url in subpage_urls:
        try:
            async with _jina_sem:
                jresp = await _fetch_with_retry(
                    client,
                    f"{_JINA_READER_BASE}{sub_url}",
                    headers={"Accept": "text/markdown"},
                    timeout=httpx.Timeout(_JINA_TIMEOUT, connect=15.0),
                )
                await asyncio.sleep(max(_JINA_DELAY - 1.5, 0.5))

            if jresp.status_code != 200:
                continue
            sub_content = jresp.text.strip()
            _head = sub_content[:600]
            _blocked = (
                ("Warning:" in _head and "returned error" in _head)
                or "JavaScript is disabled" in _head
                or "Max challenge" in _head
            )
            if not sub_content or len(sub_content) < 100 or _blocked:
                continue

            page_title = sub_url
            for line in sub_content.split("\n"):
                if line.strip().startswith("# "):
                    page_title = line.strip()[2:]
                    break

            parts.append(
                f"# {page_title}\n\n**Source**: {sub_url}\n\n---\n\n{sub_content}"
            )
        except Exception:
            pass

    if not parts:
        return ""

    return "\n\n---\n<!-- SUB-PAGES (via Jina) -->\n\n" + "\n\n---\n\n".join(parts)


async def _extract_markdown_attachments(
    client: httpx.AsyncClient,
    markdown: str,
    base_url: str,
    max_attachments: int = _MAX_FILE_ATTACHMENTS_PER_PAGE,
) -> str:
    """Download file attachments linked inside Jina-returned Markdown text.

    Jina Reader converts pages to Markdown directly, so `_extract_html_pdf_attachments`
    (which parses HTML with BeautifulSoup) cannot be used.  This function finds
    all Markdown-syntax links `[text](url)` and `<url>` whose URL paths end with
    a known file extension, then downloads and converts each to Markdown.

    Supports: PDF, Word (.doc/.docx), Excel (.xlsx/.xls/.ods/.csv), PPT (.pptx/.ppt).
    """
    from urllib.parse import urljoin, urlparse as _up2

    # Match both [text](url) and bare <url> patterns
    _link_re = re.compile(r'\[(?:[^\]]*)\]\(([^)\s]+)\)|<(https?://[^>]+)>')

    file_links: list[tuple[str, str, str]] = []
    seen: set[str] = set()

    for m in _link_re.finditer(markdown):
        href = (m.group(1) or m.group(2) or "").strip().rstrip(")")
        if not href:
            continue
        full_url = urljoin(base_url, href) if not href.startswith("http") else href
        if not _is_safe_url(full_url) or full_url in seen:
            continue
        seen.add(full_url)

        parsed_url = _up2(full_url)
        path_lower = parsed_url.path.lower()
        matched_ext = next(
            (ext for ext in sorted(_ALL_FILE_EXTS, key=len, reverse=True)
             if path_lower.endswith(ext)),
            None,
        )
        # Also check filename= query parameter for portals like health.ec.europa.eu
        if not matched_ext:
            qs = parsed_url.query.lower()
            fn_match = re.search(r'filename=([^&]+)', qs)
            if fn_match:
                fname = fn_match.group(1)
                matched_ext = next(
                    (ext for ext in sorted(_ALL_FILE_EXTS, key=len, reverse=True) if fname.endswith(ext)),
                    None,
                )
        if matched_ext and len(file_links) < max_attachments:
            raw_label = m.group(0)
            title_part = re.sub(r'[\[\]<>()]', '', raw_label).strip()[:120] or path_lower.split("/")[-1] or "Document"
            file_links.append((title_part, full_url, matched_ext))

    if not file_links:
        return ""

    parts: list[str] = []
    for title, att_url, ext in file_links:
        try:
            att_bytes = await _fetch_bytes_bounded(
                client, att_url, timeout=httpx.Timeout(60, connect=10)
            )
            if len(att_bytes) < 100:
                continue

            md = ""
            if ext == ".pdf":
                if att_bytes[:4] != b"%PDF" and b"%PDF" not in att_bytes[:1024]:
                    continue
                md = _pdf_bytes_to_markdown(att_bytes, title, att_url)
                if not md and len(att_bytes) <= _PDF_ATTACHMENT_MAX_BYTES:
                    md = await _docling_pdf_bytes_to_markdown_async(att_bytes, title, att_url)
            elif ext in _WORD_EXTS:
                if len(att_bytes) <= _WORD_ATTACHMENT_MAX_BYTES and not _is_zip_bomb(att_bytes):
                    md = await _word_bytes_to_markdown_async(att_bytes, title, att_url)
            elif ext in _EXCEL_EXTS:
                if len(att_bytes) <= _EXCEL_ATTACHMENT_MAX_BYTES and not _is_zip_bomb(att_bytes):
                    md = await _excel_bytes_to_markdown_async(att_bytes, title, att_url, ext)
            elif ext in _PPT_EXTS:
                if len(att_bytes) <= _PPT_ATTACHMENT_MAX_BYTES and not _is_zip_bomb(att_bytes):
                    md = await _ppt_bytes_to_markdown_async(att_bytes, title, att_url, ext)

            if md and len(md.strip()) > 200:
                parts.append(md)
        except Exception:
            pass

    if not parts:
        return ""

    _SEP = "\n\n<!-- PDF_DOC_BREAK -->\n\n"
    return "\n\n---\n<!-- ATTACHED DOCUMENTS (Jina Markdown links) -->\n\n" + _SEP.join(parts)


def _split_attachment_sections(
    attachments: str,
    site: dict,
    region: str,
) -> list[dict]:
    """Split a combined attachment string into individual result dicts.

    Used when site has save_attachments_separately=True (e.g. MDCG index).
    Each PDF/Word/Excel section becomes its own crawl result with a
    unique agency name derived from the document title or filename.
    """
    from urllib.parse import urlparse as _up3

    # Sections are joined by "\n\n---\n\n"; skip the block header line
    raw_block = attachments
    # Strip the opening header comment
    for marker in ["<!-- ATTACHED DOCUMENTS & LINKED PAGES -->",
                   "<!-- ATTACHED DOCUMENTS (Jina Markdown links) -->",
                   "<!-- ATTACHED DOCUMENTS -->"]:
        if marker in raw_block:
            raw_block = raw_block[raw_block.index(marker) + len(marker):]
            break

    # Split into individual sections
    sections = re.split(r"\n\n<!-- PDF_DOC_BREAK -->\n\n", raw_block)
    results: list[dict] = []
    parent_agency = site.get("agency", "DOC")

    for sec in sections:
        sec = sec.strip()
        if not sec or len(sec) < 100:
            continue

        # Extract title (first # heading)
        title = ""
        for line in sec.split("\n"):
            line = line.strip()
            if line.startswith("# "):
                title = line[2:].strip()
                break

        # Extract source URL
        src_url = ""
        src_match = re.search(r'\*\*Source\*\*:\s*(https?://\S+)', sec)
        if src_match:
            src_url = src_match.group(1)

        # Build a short unique agency name from filename or title
        if src_url:
            fn_match = re.search(r'filename=([^&\s]+)', src_url)
            if fn_match:
                slug = fn_match.group(1).replace('.pdf', '').replace('.docx', '').replace('_', '-')[:50]
            else:
                slug = _up3(src_url).path.rstrip('/').split('/')[-1][:40]
        else:
            slug = re.sub(r'[^\w\-]', '-', title[:40])
        slug = slug.strip('-')

        sub_result = _make_result_template(
            {**site, "agency": f"{parent_agency}:{slug}", "name": title or slug, "url": src_url or site.get("url", "")},
            region,
        )
        sub_result["crawl_status"] = "success"
        sub_result["content_source"] = "live"
        sub_result["content_markdown"] = sec
        sub_result["title"] = title or slug
        results.append(sub_result)

    return results


async def _crawl_tier2_httpx(
    client: httpx.AsyncClient,
    site: dict,
    region: str,
    etag_cache: ETagCache,
    jina_semaphore: asyncio.Semaphore | None = None,
) -> dict:
    """Tier 2: HTML fetch → BS4 strip → MarkItDown conversion.

    For PDF responses, attempts direct text extraction via PyMuPDF before
    falling back to Jina Reader (handles scanned/image-only PDFs).
    """
    result = _make_result_template(site, region)
    url = site["url"]
    start = time.time()

    try:
        # Conditional request headers.
        # Sites with save_attachments_separately=True (e.g. MDCG index) MUST bypass
        # ETag caching: a 304 response causes early return that skips the attachment
        # extraction logic, meaning all individual PDFs are lost on cached crawls.
        req_headers = {}
        if not site.get("save_attachments_separately"):
            cached = etag_cache.get(url)
            if cached:
                if cached.get("etag"):
                    req_headers["If-None-Match"] = cached["etag"]
                if cached.get("last_modified"):
                    req_headers["If-Modified-Since"] = cached["last_modified"]

        response = await _fetch_with_retry(client, url, headers=req_headers)

        # 304 Not Modified — if no cached content exists, retry without conditional headers
        if response.status_code == 304:
            previous_content = _retrieve_cached_content(url)
            if previous_content:
                _cached_entry = etag_cache.get(url) or {}
                result["crawl_status"] = "success"
                result["content_source"] = "cached"  # B-2: not "live" — restored from previous crawl
                result["title"] = f"{site['agency']} (cached — not modified)"
                result["content_markdown"] = previous_content
                result["note"] = (
                    "HTTP 304 Not Modified — restored content from previous crawl"
                    + (f" (cached_at: {_cached_entry.get('cached_at', 'unknown')})" if _cached_entry.get("cached_at") else "")
                )
                result["crawl_duration_seconds"] = round(time.time() - start, 2)
                return result
            # No cached content — force fresh request without conditional headers
            response = await _fetch_with_retry(client, url, headers={})

        response.raise_for_status()

        # Update ETag cache (only for non-attachment-split sites)
        if not site.get("save_attachments_separately"):
            etag = response.headers.get("ETag")
            last_mod = response.headers.get("Last-Modified")
            content_hash = hashlib.sha256(response.content).hexdigest()[:16]
            etag_cache.set(
                url, etag=etag, last_modified=last_mod, content_hash=content_hash
            )

        content_type = response.headers.get("content-type", "")

        # PDF — try PyMuPDF → Docling OCR (small PDFs only) → Jina
        if "application/pdf" in content_type or url.lower().endswith(".pdf"):
            pdf_name = site.get("name", site.get("agency", ""))
            if len(response.content) > _DOWNLOAD_MAX_BYTES:
                result["failure_reason"] = (
                    f"PDF response {len(response.content) // (1024 * 1024)} MB "
                    f"exceeds {_DOWNLOAD_MAX_BYTES // (1024 * 1024)} MB limit, skipping"
                )
                result["crawl_status"] = "skipped"
                return result
            # Method 1: section-aware extraction for large primary PDFs (EU MDR etc.)
            _use_section_aware = (
                len(response.content) > 500_000
                and get_site_doc_type(site) == "primary"
            )
            md = _pdf_bytes_to_markdown(response.content, pdf_name, url,
                                        section_aware=_use_section_aware)
            if md and len(md.strip()) > 200:
                result["content_markdown"] = md
                result["title"] = pdf_name
                result["crawl_status"] = "success"
                result["content_source"] = "live"
                result["crawl_duration_seconds"] = round(time.time() - start, 2)
                return result
            # Scanned PDF — try Docling OCR if file is small enough for real-time use
            if len(response.content) <= _PDF_DOCLING_MAX_BYTES_REALTIME:
                md = await _docling_pdf_bytes_to_markdown_async(
                    response.content, pdf_name, url
                )
                if md and len(md.strip()) > 200:
                    result["content_markdown"] = md
                    result["title"] = pdf_name
                    result["crawl_status"] = "success"
                    result["content_source"] = "live"
                    result["note"] = (
                        (site.get("note", "") + " [OCR via Docling]").strip()
                    )
                    result["crawl_duration_seconds"] = round(time.time() - start, 2)
                    return result
            # Large scanned PDF or Docling unavailable — hand off to Jina Reader
            size_mb = round(len(response.content) / 1024 / 1024, 1)
            reason = (
                f"PyMuPDF 無文字層（掃描版 {size_mb}MB，超過即時 OCR 限制）"
                if len(response.content) > _PDF_DOCLING_MAX_BYTES_REALTIME
                else ("PyMuPDF 無文字層（掃描版）" if FITZ_AVAILABLE else "PyMuPDF 未安裝")
            )
            result["failure_reason"] = (
                f"PDF 回應 ({content_type}) — {reason}，轉至 Jina Reader"
            )
            result["crawl_duration_seconds"] = round(time.time() - start, 2)
            return result

        raw_text = response.text

        if "application/json" in content_type:
            # Unexpected JSON from HTML tier — handle gracefully
            try:
                json_data = response.json()
                result["content_markdown"] = _json_to_markdown(json_data, url)
                result["title"] = json_data.get("title", site["agency"])
            except Exception:
                result["content_markdown"] = _html_to_markdown(raw_text, url)
                result["title"] = site["agency"]
        else:
            # HTML → MarkItDown (with BS4 pre-strip)
            result["content_markdown"] = _html_to_markdown(raw_text, url)

            # Extract title
            try:
                soup = BeautifulSoup(raw_text, "lxml")
                title_tag = soup.find("title")
                result["title"] = (
                    title_tag.string.strip()
                    if title_tag and title_tag.string
                    else site["agency"]
                )
            except Exception:
                result["title"] = site["agency"]

        if result["content_markdown"] and len(result["content_markdown"].strip()) > 50:
            result["crawl_status"] = "success"
            result["content_source"] = "live"
            # Detect and download any PDF attachments linked from this HTML page
            if "application/json" not in content_type:
                try:
                    _attach_limit = (
                        _MAX_FILE_ATTACHMENTS_INDEX_PAGE
                        if site.get("index_page")
                        else _MAX_FILE_ATTACHMENTS_PER_PAGE
                    )
                    _subpage_limit = site.get("max_subpages", _MAX_SUBPAGES_PER_PAGE)
                    # Enable Jina sub-page fallback when explicitly configured
                    _jina_subpages = site.get("jina_subpage_fallback", False)
                    attachments = await _extract_html_pdf_attachments(
                        client, raw_text, url,
                        max_attachments=_attach_limit,
                        max_subpages=_subpage_limit,
                        jina_subpage_fallback=_jina_subpages,
                        jina_semaphore=jina_semaphore,
                    )
                    if attachments:
                        if site.get("save_attachments_separately"):
                            # Each PDF becomes a separate crawl result
                            splits = _split_attachment_sections(
                                attachments, site, region
                            )
                            result["_attachment_splits"] = splits
                            logger.info(
                                "Attachment splits for %s/%s: %d docs extracted",
                                region, site.get("agency", ""), len(splits),
                            )
                        else:
                            result["content_markdown"] += attachments
                    elif site.get("save_attachments_separately"):
                        logger.warning(
                            "No attachments extracted for %s/%s (index_page site) — "
                            "page may have changed structure or all downloads failed",
                            region, site.get("agency", ""),
                        )
                except Exception as _att_exc:
                    logger.warning(
                        "Attachment extraction failed for %s/%s: %s",
                        region, site.get("agency", ""), str(_att_exc)[:200],
                    )
        else:
            result["failure_reason"] = (
                "頁面內容為空或需要 JavaScript 渲染 — "
                "網站可能為 SPA 架構，純 HTTP 請求無法取得實際內容"
            )

    except Exception as e:
        result["failure_reason"] = _classify_failure(e, url)

    result["crawl_duration_seconds"] = round(time.time() - start, 2)
    return result


async def _crawl_tier3_jina(
    client: httpx.AsyncClient,
    site: dict,
    region: str,
    jina_semaphore: asyncio.Semaphore,
) -> dict:
    """Tier 3: Jina Reader API for anti-scraping / SPA / blocked sites."""
    result = _make_result_template(site, region)
    url = site["url"]
    start = time.time()

    try:
        async with jina_semaphore:
            jina_url = f"{_JINA_READER_BASE}{url}"
            response = await _fetch_with_retry(
                client,
                jina_url,
                headers={"Accept": "text/markdown"},
                timeout=httpx.Timeout(_JINA_TIMEOUT, connect=15.0),
            )

            # Rate limit delay
            await asyncio.sleep(_JINA_DELAY)

        if response.status_code == 200:
            content = response.text.strip()

            # Detect Jina error/bot-block pages (all wrapped in HTTP 200).
            # Patterns: upstream errors, CAPTCHA challenges, JS-required, bot detection.
            _head500 = content[:600]
            _jina_error = (
                ("Warning:" in _head500 and "returned error" in _head500)
                or "Max challenge attempts exceeded" in _head500
                or "JavaScript is disabled" in _head500
                or "Please enable JavaScript" in _head500
                or "cf-challenge" in _head500
                or "Cloudflare" in _head500
                or ("Bot" in _head500 and "detect" in _head500.lower())
                or ("Access denied" in _head500)
            )

            if content and len(content) > 50 and not _jina_error:
                result["content_markdown"] = content
                # Extract title from first heading
                for line in content.split("\n"):
                    line = line.strip()
                    if line.startswith("# "):
                        result["title"] = line[2:].strip()
                        break
                if not result["title"]:
                    result["title"] = site.get("name", site["agency"])
                result["crawl_status"] = "success"
                result["content_source"] = "live"

                # Download file attachments found as Markdown links in Jina output
                try:
                    _attach_limit = (
                        _MAX_FILE_ATTACHMENTS_INDEX_PAGE
                        if site.get("index_page")
                        else _MAX_FILE_ATTACHMENTS_PER_PAGE
                    )
                    attachments = await _extract_markdown_attachments(
                        client, content, url, max_attachments=_attach_limit
                    )
                    if attachments:
                        result["content_markdown"] += attachments
                except Exception:
                    pass

                # Follow same-domain HTML sub-page links via Jina (for pages like
                # legislation.gov.uk TOC where articles live on individual sub-pages)
                if site.get("jina_subpage_fallback"):
                    try:
                        subpages_md = await _extract_jina_subpages(
                            client,
                            content,
                            url,
                            max_subpages=site.get("max_subpages", _MAX_SUBPAGES_PER_PAGE),
                            jina_semaphore=jina_semaphore,
                        )
                        if subpages_md:
                            result["content_markdown"] += subpages_md
                    except Exception:
                        pass

            elif _jina_error:
                _warn_line = content.split("\n")[0][:120]
                result["failure_reason"] = (
                    f"Jina Reader 回傳 bot-block/challenge — {_warn_line}"
                )
            else:
                result["failure_reason"] = (
                    "Jina Reader 回傳內容為空 — 網站可能完全封鎖爬取"
                )
        else:
            result["failure_reason"] = (
                f"Jina Reader HTTP {response.status_code} — 無法透過 Jina 取得內容"
            )

    except Exception as e:
        result["failure_reason"] = _classify_failure(e, url)

    result["crawl_duration_seconds"] = round(time.time() - start, 2)
    return result


# ============================================================
# Core Async Crawler Class
# ============================================================


class AsyncRegulatoryUpdateCrawler:
    """High-performance async regulatory website crawler for medical device QMS.

    Features:
      - 4-tier crawling strategy (Sitemap → API → httpx+MiD → Jina)
      - Parallel crawling via asyncio.gather
      - HTTP/2 with shared connection pool
      - ETag conditional caching
      - Per-domain rate limiting
      - tenacity exponential-backoff retry
    """

    def __init__(self):
        self._client: Optional[httpx.AsyncClient] = None
        self._etag_cache = ETagCache()
        self._sitemap_scanner = SitemapScanner()
        self._domain_semaphores: dict[str, asyncio.Semaphore] = {}
        self._jina_semaphore = asyncio.Semaphore(2)

    async def _ensure_client(self) -> None:
        """Lazy-init shared AsyncClient with HTTP/2 and connection pool."""
        if self._client is None or self._client.is_closed:
            _kwargs = dict(
                headers=_DEFAULT_HEADERS,
                timeout=httpx.Timeout(_REQUEST_TIMEOUT, connect=10.0),
                limits=httpx.Limits(
                    max_connections=50,
                    max_keepalive_connections=20,
                ),
                follow_redirects=True,
                verify=True,
            )
            try:
                self._client = httpx.AsyncClient(**_kwargs, http2=True)
            except ImportError:
                self._client = httpx.AsyncClient(**_kwargs, http2=False)

    def _get_domain_semaphore(self, url: str) -> asyncio.Semaphore:
        """Get or create a per-domain rate-limiting semaphore."""
        try:
            domain = urlparse(url).netloc
        except Exception:
            domain = url
        if domain not in self._domain_semaphores:
            self._domain_semaphores[domain] = asyncio.Semaphore(_DOMAIN_CONCURRENCY)
        return self._domain_semaphores[domain]

    async def _try_fallback_urls(
        self, site: dict, region: str, primary_reason: str
    ) -> dict | None:
        """Try each URL in site['fallback_urls'] via tier-2 httpx.

        Returns the first successful result dict, or None if all fail.
        The result has agency/region from the original site so storage
        uses the correct key.
        """
        fallback_urls: list[str] = site.get("fallback_urls", [])
        if not fallback_urls:
            return None
        for fb_url in fallback_urls:
            fb_site = {**site, "url": fb_url, "tier": 2}
            try:
                fb_result = await _crawl_tier2_httpx(
                    self._client, fb_site, region,
                    self._etag_cache, jina_semaphore=self._jina_semaphore,
                )
            except Exception as exc:
                logger.warning("fallback_url %s failed: %s", fb_url[:80], exc)
                continue
            if fb_result.get("crawl_status") == "success":
                fb_result["note"] = (
                    f"Primary URL failed ({primary_reason[:80]})"
                    f" → fallback URL succeeded: {fb_url}"
                )
                logger.info(
                    "fallback_url succeeded for %s/%s: %s",
                    region, site.get("agency", ""), fb_url[:80],
                )
                return fb_result
        return None

    async def _crawl_single_site(self, site: dict, region: str) -> dict:
        """Crawl a single site with tier dispatch, fallback chain, and rate limiting.

        Fallback chain: Tier2 (httpx) → Tier3 (Jina) → fallback_urls → DuckDuckGo → pre-written profile.
        force_profile=True sites still attempt real crawl first; the pre-written profile
        is used only as absolute last resort after all live tiers fail.
        """
        tier     = site.get("tier", 2)
        strategy = site.get("strategy", "")
        url      = site.get("url", "")
        sem      = self._get_domain_semaphore(url)

        # bulk_zip: Taiwan MOJ Open API — download all laws in one ZIP,
        # return a single merged crawl result containing all filtered laws.
        if strategy == "bulk_zip":
            return await _crawl_bulk_zip(site, region)

        async with sem:
            crawl_delay = min(site.get("crawl_delay", 3), 5)
            await asyncio.sleep(crawl_delay)

            if tier == 1:
                result = await _crawl_tier1_api(
                    self._client, site, region, self._etag_cache
                )
                if result.get("crawl_status") != "success":
                    primary_reason = result.get("failure_reason", "未知")
                    # Try fallback_urls (e.g. legislation.gov.uk PDF when CELLAR blocked)
                    fb = await self._try_fallback_urls(site, region, primary_reason)
                    if fb:
                        return fb
                    # DDG URL discovery
                    ddgs_result = await self._fallback_ddgs_search(site, region)
                    if ddgs_result.get("crawl_status") == "success":
                        ddgs_result["note"] = (
                            f"API 失敗 ({primary_reason})"
                            f" → DDG URL 發現成功"
                        )
                        return ddgs_result
                    profile = self._fallback_profile(site, region)
                    if profile:
                        return profile
                return result
            elif tier == 3:
                result = await _crawl_tier3_jina(
                    self._client, site, region, self._jina_semaphore
                )
                # B-1: validate Jina content — bot-block / nav pages must not pass
                _jina_content_ok = (
                    result.get("crawl_status") == "success"
                    and _is_regulatory_fulltext(result.get("content_markdown", ""))
                )
                if not _jina_content_ok:
                    primary_reason = result.get("failure_reason") or "Jina 內容未通過法規文本驗證"
                    if result.get("crawl_status") == "success":
                        primary_reason = "Jina 回傳內容未通過法規文本驗證（可能為 bot-block 或導覽頁）"
                        result["crawl_status"] = "failed"
                        result["failure_reason"] = primary_reason
                    fb = await self._try_fallback_urls(site, region, primary_reason)
                    if fb:
                        return fb
                    # B-5: force_profile sites skip DDG — they are known-blocked
                    if not site.get("force_profile"):
                        ddgs_result = await self._fallback_ddgs_search(site, region)
                        if ddgs_result.get("crawl_status") == "success":
                            ddgs_result["note"] = (
                                f"Jina 失敗 ({primary_reason})"
                                f" → DuckDuckGo 備援成功"
                            )
                            return ddgs_result
                    profile = self._fallback_profile(site, region)
                    if profile:
                        profile["note"] = (
                            f"Jina 失敗 ({primary_reason})"
                            f" → 使用預設法規摘要"
                        )
                        return profile
                return result
            else:
                result = await _crawl_tier2_httpx(
                    self._client, site, region, self._etag_cache,
                    jina_semaphore=self._jina_semaphore,
                )
                if result.get("crawl_status") != "success":
                    original_reason = result.get("failure_reason", "未知")
                    jina_result = await _crawl_tier3_jina(
                        self._client, site, region, self._jina_semaphore
                    )
                    # B-1: validate Jina fallback content as well
                    _jina_ok = (
                        jina_result.get("crawl_status") == "success"
                        and _is_regulatory_fulltext(jina_result.get("content_markdown", ""))
                    )
                    if _jina_ok:
                        jina_result["note"] = (
                            f"httpx 失敗 ({original_reason}) → Jina 備援成功"
                        )
                        return jina_result
                    jina_fail_reason = jina_result.get("failure_reason") or "Jina 內容未通過法規文本驗證"
                    if jina_result.get("crawl_status") == "success":
                        jina_fail_reason = "Jina 回傳內容未通過法規文本驗證"
                    # Try fallback_urls before DDG
                    combined_reason = (
                        f"{original_reason}"
                        f" → Jina 失敗 ({jina_fail_reason})"
                    )
                    fb = await self._try_fallback_urls(site, region, combined_reason)
                    if fb:
                        return fb
                    ddgs_result = await self._fallback_ddgs_search(site, region)
                    if ddgs_result.get("crawl_status") == "success":
                        ddgs_result["note"] = (
                            f"httpx 失敗 ({original_reason})"
                            f" → Jina 失敗 ({jina_fail_reason})"
                            f" → DuckDuckGo 備援成功"
                        )
                        return ddgs_result
                    profile = self._fallback_profile(site, region)
                    if profile:
                        profile["note"] = (
                            f"httpx 失敗 ({original_reason})"
                            f" → Jina 失敗 ({jina_fail_reason})"
                            f" → DuckDuckGo 亦失敗 → 使用預設法規摘要"
                        )
                        return profile
                    result["failure_reason"] = (
                        f"{original_reason}"
                        f" → Jina 備援亦失敗 ({jina_fail_reason})"
                        f" → DuckDuckGo 備援亦失敗"
                    )
                return result

    def _fallback_profile(self, site: dict, region: str):
        """Return a pre-written profile result if one exists for this site, else None."""
        key = (region, site.get("agency", ""))
        content = REGION_PROFILES.get(key)
        if not content:
            return None
        result = _make_result_template(site, region)
        result["crawl_status"] = "success"
        result["content_source"] = "pre-written"
        result["content_markdown"] = content
        # Extract title from first heading
        for line in content.split("\n"):
            line = line.strip()
            if line.startswith("# "):
                result["title"] = line[2:].strip()
                break
        if not result["title"]:
            result["title"] = site.get("name", site.get("agency", ""))
        result["crawl_duration_seconds"] = 0.0
        return result

    async def _ddgs_url_discovery(
        self,
        site: dict,
        region: str,
        candidate_urls: list,
        start: float,
    ):
        """Try up to 3 DDG-discovered URLs via Tier2/Tier3 with per-fetch timeouts.

        Each Tier2 fetch is capped at 8s, each Tier3 Jina fetch at 12s.
        Returns a successful result dict or None.
        Called inside asyncio.wait_for (25s hard cap) from _fallback_ddgs_search.
        """
        for candidate_url in candidate_urls[:3]:  # max 3 URLs (was 5)
            alt_site = dict(site)
            alt_site["url"] = candidate_url
            alt_site.pop("sitemap_url", None)

            # Tier2 httpx — 15s per-fetch timeout
            try:
                alt_result = await asyncio.wait_for(
                    _crawl_tier2_httpx(
                        self._client, alt_site, region, self._etag_cache,
                        jina_semaphore=self._jina_semaphore,
                    ),
                    timeout=15.0,
                )
                if alt_result.get("crawl_status") == "success":
                    content = alt_result.get("content_markdown", "")
                    if _is_regulatory_fulltext(content):
                        alt_result["note"] = f"DDG URL 發現 (httpx): {candidate_url[:100]}"
                        alt_result["crawl_duration_seconds"] = round(time.time() - start, 2)
                        return alt_result
            except asyncio.TimeoutError:
                logger.debug("DDG alt httpx timeout (%s)", candidate_url[:60])
            except Exception as e:
                logger.debug("DDG alt httpx failed (%s): %s", candidate_url[:60], str(e)[:80])

            # Tier3 Jina — 20s per-fetch timeout
            try:
                alt_result = await asyncio.wait_for(
                    _crawl_tier3_jina(
                        self._client, alt_site, region, self._jina_semaphore
                    ),
                    timeout=20.0,
                )
                if alt_result.get("crawl_status") == "success":
                    content = alt_result.get("content_markdown", "")
                    if _is_regulatory_fulltext(content):
                        alt_result["note"] = f"DDG URL 發現 (Jina): {candidate_url[:100]}"
                        alt_result["crawl_duration_seconds"] = round(time.time() - start, 2)
                        return alt_result
            except asyncio.TimeoutError:
                logger.debug("DDG alt Jina timeout (%s)", candidate_url[:60])
            except Exception as e:
                logger.debug("DDG alt Jina failed (%s): %s", candidate_url[:60], str(e)[:80])

        return None

    async def _fallback_ddgs_search(self, site: dict, region: str) -> dict:
        """Fallback: DDG URL discovery → Tier2/Tier3 fetch → snippet summary.

        M1: Extracts alternative URLs from DDG results and attempts real page fetches.
        M2: Uses _build_ddg_query() for citation-aware targeted queries.
        M3: Validates fetched content is actual regulatory text via _is_regulatory_fulltext().
        M5: Called after both httpx and Jina fail, so alt URLs cover both paths.
        URL discovery is hard-capped at 45s total to prevent Freshness Check hangs.
        Only falls back to snippet-combination if all URL fetches fail or lack full text.
        """
        result = _make_result_template(site, region)
        start = time.time()
        agency = site.get("agency", "")
        try:
            # M2: targeted query using regulation citation from metadata
            query = _build_ddg_query(site, region)
            logger.debug("DDG fallback query for %s/%s: %s", region, agency, query)

            search_results = await asyncio.to_thread(_ddgs_search, query, 8)
            if not search_results:
                result["failure_reason"] = "DuckDuckGo 搜尋無結果"
                result["crawl_duration_seconds"] = round(time.time() - start, 2)
                return result

            # M1: URL 抓取：只使用 Tier 0-4a（官方/半官方/代表性民間機構）
            # min_score=35 排除隨機一般網頁，Wikipedia 等已由 score=-1 排除
            ranked_results = _sort_by_credibility(search_results, min_score=_MIN_FETCH_CREDIBILITY)
            candidate_urls = [
                sr.get("href") or sr.get("link") or ""
                for sr in ranked_results
                if (sr.get("href") or sr.get("link")) and _is_safe_url(sr.get("href") or sr.get("link", ""))
            ]
            logger.debug(
                "DDG candidates for %s/%s (credibility-ranked): %s",
                region, agency,
                [u[:60] for u in candidate_urls[:3]],
            )

            # URL discovery with hard 45s total cap
            if candidate_urls:
                try:
                    disc_result = await asyncio.wait_for(
                        self._ddgs_url_discovery(site, region, candidate_urls, start),
                        timeout=45.0,
                    )
                    if disc_result is not None:
                        return disc_result
                except asyncio.TimeoutError:
                    logger.warning(
                        "DDG URL discovery exceeded 45s budget for %s/%s — falling back to snippets",
                        region, agency,
                    )

            # 摘要片段：所有非 Tier 9 來源都顯示，但每筆都標記來源層級
            # min_score=0 保留一般網頁作最後手段，但標記「🌐 一般網頁（最後手段）」
            snippet_results = _sort_by_credibility(search_results, min_score=0)
            md_parts = [f"# {region} — {agency} (DuckDuckGo 搜尋摘要)\n\n"
                        f"> ⚠️ 以下結果依來源可信度排序。所有資料均標記出處層級，請以官方來源為準。\n"]
            for i, sr in enumerate(snippet_results[:8], 1):
                title = sr.get("title", "")
                body = sr.get("body", "")
                href = sr.get("href", sr.get("link", ""))
                score = _url_credibility_score(href)
                label = _credibility_label(score)
                md_parts.append(
                    f"## {i}. {title}\n\n{body}\n\n"
                    f"**出處**：{href}  \n**來源層級**：{label}\n"
                )
            combined = "\n---\n".join(md_parts)
            if len(combined.strip()) > 100:
                result["crawl_status"] = "success"
                result["content_source"] = "live"
                result["content_markdown"] = combined
                result["title"] = f"{agency} (DDG snippet reference)"
                result["note"] = "DDG URL 全文抓取失敗 — 以搜尋摘要作為參考資料"
            else:
                result["failure_reason"] = "DuckDuckGo 搜尋結果內容不足"
        except Exception as e:
            result["failure_reason"] = f"DuckDuckGo 備援失敗: {str(e)[:200]}"
        result["crawl_duration_seconds"] = round(time.time() - start, 2)
        return result

    async def crawl_all_regions(self) -> dict:
        """Crawl all configured regions in parallel.

        Returns structured result dict (same schema as v1.0).
        """
        return await self._crawl_regions(list(REGION_SITES.keys()))

    async def crawl_selected_regions(self, regions: list) -> dict:
        """Crawl only specified regions in parallel.

        Returns structured result dict (same schema as v1.0).
        """
        valid_regions = [r for r in regions if r in REGION_SITES]
        if not valid_regions:
            return {
                "results": [],
                "summary": {
                    "total_sites": 0,
                    "success_count": 0,
                    "failed_count": 0,
                    "regions_covered": [],
                    "crawl_duration_seconds": 0.0,
                },
            }
        return await self._crawl_regions(valid_regions)

    async def _crawl_regions(self, regions: list) -> dict:
        """Internal: crawl a list of regions in parallel with URL deduplication.

        When multiple regions share the same URL (e.g. Canada and MDSAP both
        reference mdsap.global pages), each unique URL is crawled only once.
        The result is then cloned for every alias region that shares it.
        """
        await self._ensure_client()
        await self._etag_cache.load()

        start_time = time.time()

        # --- URL deduplication ------------------------------------------------
        # url_to_primary: URL -> (primary_region, site_dict)  (first seen wins)
        # url_to_aliases: URL -> [(alias_region, site_dict), ...]
        url_to_primary: dict[str, tuple[str, dict]] = {}
        url_to_aliases: dict[str, list[tuple[str, dict]]] = {}

        for region in regions:
            sites = REGION_SITES.get(region, [])
            for site in sites:
                url = site.get("url", "")
                if url in url_to_primary:
                    url_to_aliases.setdefault(url, []).append((region, site))
                else:
                    url_to_primary[url] = (region, site)

        # Build one task per *unique* URL
        task_keys: list[str] = []  # parallel list of URLs for index lookup
        tasks = []
        for url, (primary_region, site) in url_to_primary.items():
            task_keys.append(url)
            tasks.append(self._crawl_single_site(site, primary_region))

        # Execute all in parallel
        raw_results = await asyncio.gather(*tasks, return_exceptions=True)

        # Process results — convert exceptions to failed results
        all_results = []
        for i, r in enumerate(raw_results):
            url = task_keys[i]
            primary_region, site_info = url_to_primary[url]

            if isinstance(r, Exception):
                failed_result = _make_result_template(site_info, primary_region)
                failed_result["failure_reason"] = _classify_failure(
                    r, site_info.get("url", "")
                )
                failed_result["crawl_duration_seconds"] = 0.0
                all_results.append(failed_result)
                # Clone failure for aliases
                for alias_region, alias_site in url_to_aliases.get(url, []):
                    alias_result = _make_result_template(alias_site, alias_region)
                    alias_result["failure_reason"] = failed_result["failure_reason"]
                    alias_result["crawl_duration_seconds"] = 0.0
                    alias_result["shared_from"] = primary_region
                    all_results.append(alias_result)
            elif isinstance(r, dict):
                all_results.append(r)
                # Expand attachment splits into separate results (save_attachments_separately)
                for split_result in r.pop("_attachment_splits", []):
                    all_results.append(split_result)
                # Clone success for aliases
                for alias_region, alias_site in url_to_aliases.get(url, []):
                    import copy

                    alias_result = copy.deepcopy(r)
                    alias_result["region"] = alias_region
                    alias_result["agency"] = alias_site.get(
                        "agency", r.get("agency", "")
                    )
                    alias_result["agency_name"] = alias_site.get(
                        "name", r.get("agency_name", "")
                    )
                    alias_result["shared_from"] = primary_region
                    all_results.append(alias_result)

        # Save ETag cache
        await self._etag_cache.save()

        # Build summary
        success_count = sum(
            1 for r in all_results if r.get("crawl_status") == "success"
        )
        failed_count = sum(1 for r in all_results if r.get("crawl_status") == "failed")

        return {
            "results": all_results,
            "summary": {
                "total_sites": len(all_results),
                "success_count": success_count,
                "failed_count": failed_count,
                "regions_covered": regions,
                "crawl_duration_seconds": round(time.time() - start_time, 2),
            },
        }

    def search_supplementary(self, query: str) -> list:
        """Run DuckDuckGo search for supplementary regulatory information."""
        return _ddgs_search(query)

    async def close(self) -> None:
        """Close the HTTP client and persist cache."""
        if self._client and not self._client.is_closed:
            await self._client.aclose()
            self._client = None
        await self._etag_cache.save()


# ============================================================
# Module-level convenience functions (backward-compatible API)
# ============================================================

_crawler_instance: Optional[AsyncRegulatoryUpdateCrawler] = None


def get_regulatory_crawler() -> AsyncRegulatoryUpdateCrawler:
    """Get or create singleton crawler instance."""
    global _crawler_instance
    if _crawler_instance is None:
        _crawler_instance = AsyncRegulatoryUpdateCrawler()
    return _crawler_instance


def get_available_regions() -> list:
    """Return list of all available region names."""
    return list(REGION_SITES.keys())


def get_region_display_info() -> list:
    """Return region info for UI display."""
    info = []
    for region, sites in REGION_SITES.items():
        agencies = [s["agency"] for s in sites]
        info.append(
            {
                "region": region,
                "agencies": agencies,
                "site_count": len(sites),
            }
        )
    return info


# ============================================================
# Regulation Freshness Check (Pre-Cross-Examination)
# ============================================================

# Known latest versions for baseline comparison
_KNOWN_STANDARDS = {
    "ISO_13485": {
        "version": "2016",
        "full_name": "ISO 13485:2016",
        "full_name_zh": "ISO 13485:2016 醫療器材品質管理系統",
        "check_urls": [
            "https://www.iso.org/standard/59752.html",
        ],
        "sitemap_url": "https://www.iso.org/sitemap.xml",
        "keywords": ["13485"],
        "note": "ISO 13485 full text is behind a paywall; only version/lastmod can be checked.",
    },
    "MDSAP": {
        "version": "current",
        "full_name": "MDSAP (Medical Device Single Audit Program)",
        "full_name_zh": "MDSAP（醫療器材單一稽核方案）",
        "check_urls": [
            "https://www.mdsap.global/documents/audit-procedures-and-forms",
            "https://www.mdsap.global/documents/general-documents-and-procedures",
            "https://www.fda.gov/medical-devices/medical-device-single-audit-program-mdsap/mdsap-audit-procedures-and-forms",
            "https://www.fda.gov/medical-devices/medical-device-single-audit-program-mdsap/mdsap-assessment-procedures-and-forms",
            "https://www.fda.gov/medical-devices/cdrh-international-programs/medical-device-single-audit-program-mdsap",
        ],
        "sitemap_url": "https://www.fda.gov/sitemap.xml",
        "keywords": ["mdsap", "single-audit", "single_audit"],
    },
}


async def check_regulation_freshness(
    standards: Optional[list] = None,
    progress_callback: Optional[Callable] = None,
) -> dict:
    """Check if ISO 13485 and MDSAP regulation references are up-to-date.

    Performs lightweight checks:
      1. Sitemap lastmod scan for known standard URLs
      2. HTTP HEAD request for Last-Modified headers
      3. Compares against known baseline versions

    Args:
        standards: List of standard keys to check (default: all).
                   Valid keys: 'ISO_13485', 'MDSAP'

    Returns:
        {
            'checked_at': ISO timestamp,
            'results': {
                'ISO_13485': {
                    'status': 'confirmed' | 'unconfirmed' | 'error',
                    'known_version': '2016',
                    'last_modified': '2024-...' or None,
                    'message': str,
                    'message_zh': str,
                },
                ...
            },
            'all_confirmed': bool,
            'announcement_needed': bool,
            'announcement_text': str,
            'announcement_text_zh': str,
        }
    """
    check_keys = standards or list(_KNOWN_STANDARDS.keys())
    results = {}
    crawler = get_regulatory_crawler()
    await crawler._ensure_client()
    client = crawler._client

    for key in check_keys:
        std = _KNOWN_STANDARDS.get(key)
        if not std:
            results[key] = {
                "status": "error",
                "message": f"Unknown standard: {key}",
                "message_zh": f"未知的標準: {key}",
            }
            continue

        last_modified = None
        status = "unconfirmed"
        detail = ""
        detail_zh = ""

        # Method 1: HTTP HEAD to check Last-Modified / response status
        for url in std.get("check_urls", []):
            try:
                resp = await client.head(
                    url,
                    timeout=httpx.Timeout(10.0),
                    follow_redirects=True,
                )
                if resp.status_code == 200:
                    lm = resp.headers.get("last-modified", "")
                    if lm:
                        last_modified = lm
                    status = "confirmed"
                    detail = (
                        f"{std['full_name']} page is accessible. "
                        f"Known version: {std['version']}."
                    )
                    detail_zh = (
                        f"{std['full_name_zh']} 頁面可存取。"
                        f"已知版本: {std['version']}。"
                    )
                    if last_modified:
                        detail += f" Last-Modified: {last_modified}."
                        detail_zh += f" 最後修改: {last_modified}。"
                    break  # First successful check is enough
                elif resp.status_code == 403:
                    detail = (
                        f"{std['full_name']} returned 403 Forbidden. "
                        "Content may be behind a paywall."
                    )
                    detail_zh = (
                        f"{std['full_name_zh']} 返回 403 禁止存取。內容可能在付費牆後。"
                    )
                    # ISO standards are paywalled — this is expected
                    if key == "ISO_13485":
                        status = "confirmed"
                        detail += " (Expected for ISO standards — version confirmed via known baseline.)"
                        detail_zh += "（ISO標準預期行為 — 版本透過已知基線確認。）"
            except Exception as e:
                logger.debug(f"Freshness check failed for {url}: {e}")
                detail = f"Connection error checking {std['full_name']}: {e}"
                detail_zh = f"檢查 {std['full_name_zh']} 連線錯誤: {e}"

        # Method 2: Sitemap lastmod check (if available)
        if status == "unconfirmed" and std.get("sitemap_url"):
            try:
                scanner = SitemapScanner()
                urls = await scanner.scan(
                    client,
                    std["sitemap_url"],
                    keywords=std.get("keywords", []),
                )
                if urls:
                    status = "confirmed"
                    detail = (
                        f"{std['full_name']} found in sitemap with "
                        f"{len(urls)} matching URL(s)."
                    )
                    detail_zh = (
                        f"{std['full_name_zh']} 在網站地圖中找到 "
                        f"{len(urls)} 個匹配的URL。"
                    )
            except Exception as e:
                logger.debug(f"Sitemap check failed: {e}")

        results[key] = {
            "status": status,
            "known_version": std.get("version", ""),
            "last_modified": last_modified,
            "message": detail or f"Unable to confirm {std['full_name']} freshness.",
            "message_zh": detail_zh or f"無法確認 {std['full_name_zh']} 的最新狀態。",
        }

    all_confirmed = all(r.get("status") == "confirmed" for r in results.values())
    announcement_needed = not all_confirmed

    # Build announcement text for unconfirmed standards
    unconfirmed = [
        f"- {_KNOWN_STANDARDS[k]['full_name']}: {results[k].get('message', '')}"
        for k in results
        if results[k].get("status") != "confirmed"
    ]
    unconfirmed_zh = [
        f"- {_KNOWN_STANDARDS[k]['full_name_zh']}: {results[k].get('message_zh', '')}"
        for k in results
        if results[k].get("status") != "confirmed"
    ]

    if announcement_needed:
        announcement = (
            "⚠️ Regulation Freshness Notice\n"
            "The following regulatory standards could not be confirmed as the latest version "
            "before cross-examination. Results should be reviewed with this in mind:\n"
            + "\n".join(unconfirmed)
        )
        announcement_zh = (
            "⚠️ 法規最新性公告\n"
            "以下法規標準在交叉詰問前無法確認為最新版本。"
            "分析結果請參考此公告：\n" + "\n".join(unconfirmed_zh)
        )
    else:
        announcement = ""
        announcement_zh = ""

    country_completeness = await check_country_data_completeness(
        progress_callback=progress_callback,
    )
    incomplete = country_completeness.get("incomplete_countries", [])

    # Note: per-country upload reminders are handled by app.py _auto_trigger_crossexam()
    # using i18n keys, so we only pass the country_completeness data here.
    # Do NOT append upload_notice to announcement_text to avoid duplicate messages.
    if incomplete:
        announcement_needed = True

    return {
        "checked_at": datetime.now(timezone.utc).isoformat(),
        "results": results,
        "all_confirmed": all_confirmed
        and country_completeness.get("all_complete", True),
        "announcement_needed": announcement_needed,
        "announcement_text": announcement,
        "announcement_text_zh": announcement_zh,
        "country_completeness": country_completeness,
    }


# ============================================================
# Per-Country Data Completeness Check (7-Country Cross-Examination)
# ============================================================

# Mapping: cross-examination profile ID → REGION_SITES key
# Static mapping for predefined 7 countries.
# Dynamic countries are added via get_crossexam_country_map().
_CROSSEXAM_COUNTRY_MAP_STATIC = {
    "QMSR": {
        "region": "美國 (USA)",
        "name_en": "USA (FDA QMSR)",
        "name_zh": "美國 (FDA QMSR)",
    },
    "EU_MDR": {
        "region": "歐盟 (EU)",
        "name_en": "EU (MDR 2017/745)",
        "name_zh": "歐盟 (MDR 2017/745)",
    },
    "TFDA": {
        "region": "台灣 (Taiwan)",
        "name_en": "Taiwan (TFDA)",
        "name_zh": "台灣 (TFDA)",
    },
    "HC": {
        "region": "加拿大 (Canada)",
        "name_en": "Canada (HC/MDSAP)",
        "name_zh": "加拿大 (HC/MDSAP)",
    },
    "PMDA": {
        "region": "日本 (Japan)",
        "name_en": "Japan (PMDA)",
        "name_zh": "日本 (PMDA)",
    },
    "ANVISA": {
        "region": "巴西 (Brazil)",
        "name_en": "Brazil (ANVISA)",
        "name_zh": "巴西 (ANVISA)",
    },
    "TGA": {
        "region": "澳洲 (Australia)",
        "name_en": "Australia (TGA)",
        "name_zh": "澳洲 (TGA)",
    },
    "MDSAP": {
        "region": "MDSAP",
        "name_en": "MDSAP (Single Audit)",
        "name_zh": "MDSAP（單一稽核方案）",
    },
}


def get_crossexam_country_map() -> dict:
    """Return the full cross-examination country map (static 7 + dynamic)."""
    result = dict(_CROSSEXAM_COUNTRY_MAP_STATIC)

    # Add dynamically registered profiles (from crawled regulations)
    try:
        from src.analysis.compliance_rules import PREDEFINED_REGULATIONS

        for profile_id, profile in PREDEFINED_REGULATIONS.items():
            if profile_id in result:
                continue  # Already in static map
            # Build dynamic entry from profile metadata
            region_name = f"{profile.country_name_zh} ({profile.country_name_en})"
            result[profile_id] = {
                "region": region_name,
                "name_en": f"{profile.country_name_en} ({profile_id})",
                "name_zh": f"{profile.country_name_zh} ({profile_id})",
            }
    except Exception:
        pass  # Non-critical — static map still available

    return result


_MIN_COMPLETE_CONTENT_LEN = 500  # same threshold as _crawl_tier2_httpx


async def check_country_data_completeness(
    progress_callback: Optional[Callable] = None,
) -> dict:
    """Check whether each cross-examination country (predefined 7 + dynamic) has
    complete regulation data available via crawler.

    For each country:
      1. Attempt a lightweight crawl of its configured sites
      2. If ANY site returns ≥50 chars of content → data is 'complete'
      3. If ALL sites fail or return empty content → 'incomplete',
         user should be notified to manually upload

    Args:
        progress_callback: Optional async callable(completed, total, country_name_zh)
                          called after each country finishes crawling.

    Returns:
        {
            'checked_at': ISO timestamp,
            'countries': {
                'QMSR': {
                    'region': '美國 (USA)',
                    'status': 'complete' | 'incomplete' | 'error',
                    'needs_manual_upload': bool,
                    'sites_checked': int,
                    'sites_success': int,
                    'message': str,
                    'message_zh': str,
                },
                ...
            },
            'incomplete_countries': ['QMSR', ...],  # convenience list
            'all_complete': bool,
        }
    """
    crawler = get_regulatory_crawler()
    await crawler._ensure_client()

    countries_result = {}
    incomplete_list = []
    country_map = get_crossexam_country_map()
    total_countries = len(country_map)
    completed_countries = 0

    for profile_id, country_info in country_map.items():
        region_key = country_info["region"]
        sites = REGION_SITES.get(region_key, [])

        if not sites:
            countries_result[profile_id] = {
                "region": region_key,
                "status": "error",
                "needs_manual_upload": True,
                "sites_checked": 0,
                "sites_success": 0,
                "message": f"No configured crawl sites for {country_info['name_en']}",
                "message_zh": f"{country_info['name_zh']} 沒有設定爬蟲網站",
            }
            incomplete_list.append(profile_id)
            completed_countries += 1
            if progress_callback:
                try:
                    await progress_callback(
                        completed_countries, total_countries, country_info["name_zh"]
                    )
                except Exception:
                    pass
            continue

        tasks = [crawler._crawl_single_site(site, region_key) for site in sites]
        try:
            raw = await asyncio.gather(*tasks, return_exceptions=True)
        except Exception as e:
            countries_result[profile_id] = {
                "region": region_key,
                "status": "error",
                "needs_manual_upload": True,
                "sites_checked": len(sites),
                "sites_success": 0,
                "message": f"Crawl error for {country_info['name_en']}: {e}",
                "message_zh": f"{country_info['name_zh']} 爬蟲錯誤: {e}",
            }
            incomplete_list.append(profile_id)
            completed_countries += 1
            if progress_callback:
                try:
                    await progress_callback(
                        completed_countries, total_countries, country_info["name_zh"]
                    )
                except Exception:
                    pass
            continue

        sites_checked = len(sites)
        sites_success = 0
        has_complete_content = False

        for r in raw:
            if isinstance(r, Exception):
                continue
            if not isinstance(r, dict):
                continue
            if r.get("crawl_status") == "success":
                content = r.get("content_markdown", "")
                if content and len(content.strip()) > _MIN_COMPLETE_CONTENT_LEN:
                    sites_success += 1
                    has_complete_content = True

        if has_complete_content:
            countries_result[profile_id] = {
                "region": region_key,
                "status": "complete",
                "needs_manual_upload": False,
                "sites_checked": sites_checked,
                "sites_success": sites_success,
                "message": f"{country_info['name_en']}: {sites_success}/{sites_checked} sites returned data",
                "message_zh": f"{country_info['name_zh']}: {sites_success}/{sites_checked} 個網站取得資料",
            }
        else:
            countries_result[profile_id] = {
                "region": region_key,
                "status": "incomplete",
                "needs_manual_upload": True,
                "sites_checked": sites_checked,
                "sites_success": sites_success,
                "message": (
                    f"{country_info['name_en']}: crawler could not retrieve complete data "
                    f"(0/{sites_checked} sites). Please upload regulation documents manually."
                ),
                "message_zh": (
                    f"{country_info['name_zh']}: 爬蟲無法取得完整資料 "
                    f"(0/{sites_checked} 個網站)。請手動上傳該國法規文件。"
                ),
            }
            incomplete_list.append(profile_id)

        completed_countries += 1
        if progress_callback:
            try:
                await progress_callback(
                    completed_countries, total_countries, country_info["name_zh"]
                )
            except Exception:
                pass

    return {
        "checked_at": datetime.now(timezone.utc).isoformat(),
        "countries": countries_result,
        "incomplete_countries": incomplete_list,
        "all_complete": len(incomplete_list) == 0,
    }
